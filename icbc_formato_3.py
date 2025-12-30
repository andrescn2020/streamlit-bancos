import streamlit as st
import io
import pdfplumber
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule

# Regex para caracteres ilegales en Excel
ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

def clean_for_excel(text):
    if not text: return ""
    text = str(text)
    text = ILLEGAL_CHARACTERS_RE.sub("", text)
    return text.strip()

def parse_importe(importe_str):
    try:
        clean = importe_str.replace(".", "").replace(",", ".")
        return float(clean)
    except:
        return 0.0

def procesar_icbc_formato_3(archivo_pdf):
    """Procesa ICBC Formato 3 (Resumen de Transferencias)"""
    st.info("Procesando archivo ICBC (Formato 3)...")

    try:
        texto_completo = ""
        with pdfplumber.open(io.BytesIO(archivo_pdf.read())) as pdf:
            for page in pdf.pages:
                texto_completo += page.extract_text() + "\n"
        
        lineas = texto_completo.splitlines()
        
        # --- Metadata ---
        titular_global = "Sin Especificar"
        periodo_global = "Sin Especificar"
        anio_global = "2024" # Default fallback
        
        # Titular: Primera línea no vacía que no sea tecnica
        for l in lineas[:5]:
            if l.strip() and not "P ER I OD O" in l:
                titular_global = l.strip()
                break
        
        # Periodo con Regex "P ER I OD O 0 1 / 06 / 20 2 5..."
        # El texto viene con espacios extraños: "P ER I OD O"
        for l in lineas[:20]:
            clean_l = l.replace(" ", "")
            if "PERIODO" in clean_l:
                # Buscar fechas dd/mm/yyyy
                fechas = re.findall(r"(\d{2}/\d{2}/\d{4})", clean_l)
                if len(fechas) >= 2:
                    periodo_global = f"Del {fechas[0]} al {fechas[1]}"
                    anio_global = fechas[0].split("/")[-1] # Tomar año del inicio
                    break

        # --- Movimientos ---
        # Formato esperado linea: "05-06 VARIOS ..."
        regex_linea = r"^(\d{2}-\d{2})\s+(.*)$"
        regex_importe = r"(\d{1,3}(?:\.\d{3})*,\d{2})"
        
        movimientos = []
        
        for l in lineas:
            l = l.strip()
            # Ignorar encabezados parecidos
            if "FECHA" in l or "HOJA N" in l: continue
            
            match_inicio = re.match(regex_linea, l)
            if match_inicio:
                fecha_dia_mes = match_inicio.group(1)
                resto = match_inicio.group(2)
                
                # Buscar importes al final de la linea
                importes = re.findall(regex_importe, resto)
                
                if importes:
                    # Asumimos logica: 
                    # Si Header dice "DEBITOS CREDITOS" y solo hay 1 numero al final,
                    # Analizamos contexto. Pero en el ejemplo dado:
                    # "3.385.000,00" (Positivo) y son transferencias entrantes (ORDENANTE URSSINO).
                    # Asumiremos CRÉDITO por defecto si hay un solo monto y es positivo contextualmente.
                    # Para robustez: En este reporte no hay signo negativo explicito visible en el ejemplo.
                    
                    monto_str = importes[-1]
                    importe = parse_importe(monto_str)
                    
                    # Decidir signo:
                    # Si hubiera 2 montos, [0]=Debito, [1]=Credito
                    # Si hay 1 monto, ¿Es Debito o Credito?
                    # "RESUMEN DE TRANSFERENCIAS" mezcla ambos?
                    # El ejemplo muestra 3.385.000,00 al final.
                    # Asumiremos Crédito (Positivo) por ahora basado en los datos vistos.
                    # TODO: Si aparecen Debitos, podrían estar alineados distinto o tener otra columna.
                    
                    # Limpiar descripcion (quitar el monto del final)
                    descripcion = resto.replace(monto_str, "").strip()
                    
                    # Fecha completa
                    fecha = f"{fecha_dia_mes.replace('-', '/')}/{anio_global}"
                    
                    movimientos.append({
                        "Fecha": fecha,
                        "Descripcion": descripcion,
                        "Importe": importe
                    })
        
        if not movimientos:
            st.error("No se encontraron movimientos en este archivo.")
            return None

        df = pd.DataFrame(movimientos)
         # Ordenar por fecha (el PDF parece estar cronologico ascendente 05 -> 13 -> 19)
        # No necesitamos invertir si ya viene ascendente
        
        # --- EXCEL ---
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Transf"
        ws.sheet_view.showGridLines = False
        
        color_bg_main = "C5001A"
        color_txt_main = "FFFFFF"
        
        thin_border = Border(left=Side(style='thin', color="A6A6A6"), 
                             right=Side(style='thin', color="A6A6A6"), 
                             top=Side(style='thin', color="A6A6A6"), 
                             bottom=Side(style='thin', color="A6A6A6"))
        
        fill_head = PatternFill(start_color="C5001A", end_color="C5001A", fill_type="solid")
        fill_row = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        # Header Global
        ws.merge_cells("A1:E1")
        tit = ws["A1"]
        tit.value = f"REPORTE TRANSFERENCIAS ICBC - {clean_for_excel(titular_global)}"
        tit.font = Font(size=14, bold=True, color=color_txt_main)
        tit.fill = fill_head
        tit.alignment = Alignment(horizontal="center", vertical="center")
        
        ws["A3"] = "TITULAR"
        ws["B3"] = clean_for_excel(titular_global)
        ws["A4"] = "PERÍODO"
        ws["B4"] = clean_for_excel(periodo_global)
        
        for r in ["3","4"]:
            ws[f"A{r}"].font = Font(bold=True)
            ws[f"B{r}"].font = Font(bold=True)

        # Tabla Unificada
        row = 7
        headers = ["Fecha", "Descripción", "Débitos", "Créditos"]
        cols = ["A", "B", "C", "D"]
        
        for c, h in zip(cols, headers):
            cell = ws[f"{c}{row}"]
            cell.value = h
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        row += 1
        
        total_deb = 0.0
        total_cred = 0.0
        
        for mov in movimientos:
            ws[f"A{row}"] = mov["Fecha"]
            ws[f"B{row}"] = mov["Descripcion"]
            
            # Lógica Heurística (según ejemplos del usuario):
            # 3.385.000 (Credito) tiene "ORD. :" en descripcion
            # 500.000 (Debito) NO tiene "ORD. :"
            is_credito = "ORD." in mov["Descripcion"] or "ORDENANTE" in mov["Descripcion"]
            
            if is_credito:
                ws[f"D{row}"] = mov["Importe"]
                ws[f"D{row}"].number_format = '"$ "#,##0.00'
                total_cred += mov["Importe"]
            else:
                ws[f"C{row}"] = mov["Importe"]
                ws[f"C{row}"].number_format = '"$ "#,##0.00'
                total_deb += mov["Importe"]
                
            # Estilos
            for c in cols:
                ws[f"{c}{row}"].border = thin_border
                ws[f"{c}{row}"].fill = fill_row
            
            row += 1
            
        # Totales
        ws[f"B{row}"] = "TOTALES"
        ws[f"B{row}"].font = Font(bold=True)
        ws[f"B{row}"].alignment = Alignment(horizontal='right')
        
        ws[f"C{row}"] = total_deb
        ws[f"C{row}"].number_format = '"$ "#,##0.00'
        ws[f"C{row}"].font = Font(bold=True)
        
        ws[f"D{row}"] = total_cred
        ws[f"D{row}"].number_format = '"$ "#,##0.00'
        ws[f"D{row}"].font = Font(bold=True)
        
        # Anchos
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18

        wb.save(output)
        output.seek(0)
        return output.getvalue()
        
    except Exception as e:
        import traceback
        st.error(f"Error procesando ICBC Formato 3: {e}")
        print(traceback.format_exc())
        return None
