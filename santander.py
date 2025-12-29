import streamlit as st
import io
import PyPDF2
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule

# Regex para caracteres ilegales en Excel
ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

def clean_for_excel(text):
    """Elimina caracteres ilegales para Excel y espacios extra"""
    if not text: return ""
    text = str(text)
    text = ILLEGAL_CHARACTERS_RE.sub("", text)
    return text.strip()

def procesar_santander_rio(archivo_pdf):
    """Procesa archivos PDF de Santander Rio con Estilo Dashboard Multi-Moneda"""
    st.info("Procesando archivo de Santander Rio...")

    try:
        # Reinicializar el archivo para lectura
        archivo_pdf.seek(0)
        
        # Abrir el PDF usando PyPDF2
        reader = PyPDF2.PdfReader(io.BytesIO(archivo_pdf.read()))
        texto_completo = "".join(page.extract_text() + "\n" for page in reader.pages)
        lineas_raw = texto_completo.splitlines()

        # 1. Metadatos (Titular, Periodo)
        titular_global = "Sin Especificar"
        periodo_global = "Sin Especificar"
        
        # Titular: Linea anterior a "CUIT:" o "CUIL:"
        for i, l in enumerate(lineas_raw[:20]):
            if "CUIT:" in l or "CUIL:" in l:
                if i > 0:
                    titular_global = lineas_raw[i-1].strip()
                break
        
        # Periodo: "Desde: 27/01/23" ... "Hasta: 02/03/23"
        f_desde = None
        f_hasta = None
        for l in lineas_raw[:30]:
            match_d = re.search(r"Desde:\s*(\d{2}/\d{2}/\d{2,4})", l)
            if match_d: f_desde = match_d.group(1)
            match_h = re.search(r"Hasta:\s*(\d{2}/\d{2}/\d{2,4})", l)
            if match_h: f_hasta = match_h.group(1)
        
        if f_desde and f_hasta:
            periodo_global = f"Del {f_desde} al {f_hasta}"

        # --- DELIMITAR SECCIONES ---
        idx_pesos = None
        idx_dolares = None
        idx_fin_pesos = None # Fin de pesos puede ser inicio dolares o fin documento
        idx_fin_dolares = None

        for i, l in enumerate(lineas_raw):
            if "Movimientos en pesos" in l and idx_pesos is None:
                idx_pesos = i
            if "Movimientos en dólares" in l and idx_dolares is None:
                idx_dolares = i
            if ("Así usaste tu dinero este mes" in l or "Detalle impositivo" in l) and idx_fin_dolares is None and idx_dolares is not None:
                idx_fin_dolares = i
            # Si no hay dolares, el fin de pesos puede ser "Así usaste..."
            if ("Así usaste tu dinero este mes" in l or "Detalle impositivo" in l) and idx_fin_pesos is None and idx_pesos is not None and idx_dolares is None:
                idx_fin_pesos = i

        # Ajustar rangos
        lineas_pesos = []
        lineas_dolares = []

        if idx_pesos is not None:
            # Fin de pesos es idx_dolares si existe, sino idx_fin_pesos, sino fin archivo
            end_p = idx_dolares if idx_dolares is not None else (idx_fin_pesos if idx_fin_pesos is not None else len(lineas_raw))
            lineas_pesos = lineas_raw[idx_pesos+1 : end_p]
        
        if idx_dolares is not None:
            end_d = idx_fin_dolares if idx_fin_dolares is not None else len(lineas_raw)
            lineas_dolares = lineas_raw[idx_dolares+1 : end_d]

        # --- FUNCION EXTRACTION (REUTILIZADA) ---
        def extraer_datos_seccion(lineas):
            movimientos_text = []
            linea_actual = ""
            saldo_ini = 0.0
            saldo_fin = 0.0
            
            # Pre-procesado para unir líneas
            for l in lineas:
                # Extraer saldos si aparecen en la sección
                if "Saldo Inicial" in l:
                    matches = re.findall(r"(-?)\$\s?([\d\.]+,\d{2})|(-?)U\$S\s?([\d\.]+,\d{2})", l)
                    # matches devuelve tuplas con grupos vacios, hay que filtrar
                    for m in matches:
                        # m = ('-', '1.200,00', '', '') para pesos
                        # m = ('', '', '-', '100,00') para dolares
                        val_str = m[1] if m[1] else m[3]
                        sign_str = m[0] if m[1] else m[2]
                        if val_str:
                            try:
                                num = float(val_str.replace(".", "").replace(",", "."))
                                if sign_str == "-": num *= -1
                                saldo_ini = num
                            except: pass

                if "Saldo total" in l:
                    matches = re.findall(r"(-?)\$\s?([\d\.]+,\d{2})|(-?)U\$S\s?([\d\.]+,\d{2})", l)
                    for m in matches:
                        val_str = m[1] if m[1] else m[3]
                        sign_str = m[0] if m[1] else m[2]
                        if val_str:
                            try: 
                                num = float(val_str.replace(".", "").replace(",", "."))
                                if sign_str == "-": num *= -1
                                saldo_fin = num
                            except: pass
                    continue # NO unir la linea de Saldo Total al movimiento anterior

                # Unir lineas de movimientos
                if re.match(r"\d{2}/\d{2}/\d{2}", l):
                    if linea_actual: movimientos_text.append(linea_actual.strip())
                    linea_actual = l
                else:
                    linea_actual += " " + l
            if linea_actual: movimientos_text.append(linea_actual.strip())

            # Parsear Movimientos
            parsed_data = []
            for mov in movimientos_text:
                if "Movimientos en" in mov: continue 
                if "Saldo Inicial" in mov: continue # Filtrar Saldo Inicial siempre
                
                fecha = mov[:8]
                resto = mov[8:]
                
                # Heurística: Buscar importe al final. 
                # Patrón: (opcional signo) (Símbolo $ o U$S) (numero) (Saldo final acumulado)
                # El saldo final acumulado suele estar al final de la linea.
                # Ejemplo: ... -$ 100.000,00 $ 6.376.160,86
                
                # Limpieza basica de moneda para facilitar regex unico
                mov_clean = mov.replace("U$S", "$").replace("U$s", "$")

                # Buscamos todos los montos monetarios
                montos = re.findall(r"([+-]?\$\s*[\d\.,]+)", mov_clean)
                
                importe = 0.0
                desc = ""
                
                if len(montos) >= 2:
                    # Asumimos: Penultimo es el importe del movimiento, Ultimo es el saldo
                    str_imp = montos[-2]
                    # Limpiar
                    signo = -1 if "-" in str_imp else 1
                    clean_imp = str_imp.replace("$", "").replace("-", "").strip().replace(".", "").replace(",", ".")
                    try:
                        importe = float(clean_imp) * signo
                    except: importe = 0.0
                    
                    # Descripción es todo lo que hay antes del importe
                    # Buscamos donde empieza el str_imp en la linea original/limpia
                    # Ojo: rfind podría fallar si hay montos iguales en la descripción.
                    # Usamos split o regex inverso.
                    
                    # Metodo seguro: Regex que capture Fecha + Descrip + Importe + Saldo
                    # Pero la descripcion es muy sucia.
                    
                    # Usaremos el indice del penultimo match
                    # Esto es aproximado pero la logica anterior era peor.
                    pass # Ya tenemos importe
                    
                    # Descripcion: eliminar fecha inicial y la parte de los montos finales
                    # Aproximacion: Cortar donde aparece el texto del importe
                    idx_imp = mov_clean.rfind(str_imp) 
                    if idx_imp != -1:
                        desc = mov[:idx_imp] # Incluye fecha en los primeros 8 chars
                        desc = desc[8:].strip() # Quitar fecha
                    else:
                        desc = resto

                    # Limpieza extra: Quitar numeros pegados al inicio (ej: 77367269Transferencia)
                    desc = re.sub(r'^\d+', '', desc).strip()

                    parsed_data.append((fecha, clean_for_excel(desc), importe))

                elif len(montos) == 1:
                    # Solo hay un monto, puede ser saldo inicial o algo raro.
                    if "Saldo Inicial" in mov: continue 
                    # Si es un movimiento sin saldo acumulado visible? Raro en este banco.
                    pass
            
            return parsed_data, saldo_ini, saldo_fin

        # Procesar
        datos_pesos, saldo_ini_pesos, saldo_fin_pesos = extraer_datos_seccion(lineas_pesos)
        datos_dolares, saldo_ini_dolares, saldo_fin_dolares = extraer_datos_seccion(lineas_dolares)
        
        # --- GENERACIÓN EXCEL MULTI-HOJA ---
        output = io.BytesIO()
        wb = Workbook()
        # Eliminar hoja default
        wb.remove(wb.active)
        
        # Estilos
        color_bg_main = "EC0000" 
        color_txt_main = "FFFFFF"
        thin_border = Border(left=Side(style='thin', color="A6A6A6"), right=Side(style='thin', color="A6A6A6"), top=Side(style='thin', color="A6A6A6"), bottom=Side(style='thin', color="A6A6A6"))
        fill_head_deb = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        fill_col_deb = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
        fill_row_deb = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")
        fill_head_cred = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        fill_col_cred = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")
        fill_row_cred = PatternFill(start_color="F2F9F1", end_color="F2F9F1", fill_type="solid")
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        red_font = Font(color='9C0006', bold=True)

        def crear_hoja_dashboard(wb, nombre_hoja, datos, s_ini, s_fin, formato_moneda='"$ "#,##0.00'):
            ws = wb.create_sheet(title=nombre_hoja)
            ws.sheet_view.showGridLines = False
            
            df = pd.DataFrame(datos, columns=["Fecha", "Descripcion", "Importe"])
            
            creditos = df[df["Importe"] > 0].copy()
            debitos = df[df["Importe"] < 0].copy()
            debitos["Importe"] = debitos["Importe"].abs() # Positivo para mostrar
            
            if df.empty:
                creditos = pd.DataFrame(columns=["Fecha", "Descripcion", "Importe"])
                debitos = pd.DataFrame(columns=["Fecha", "Descripcion", "Importe"])

            # Header
            ws.merge_cells("A1:G1")
            tit = ws["A1"]
            tit.value = f"REPORTE SANTANDER ({nombre_hoja}) - {clean_for_excel(titular_global)}"
            tit.font = Font(size=14, bold=True, color=color_txt_main)
            tit.fill = PatternFill(start_color=color_bg_main, end_color=color_bg_main, fill_type="solid")
            tit.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 25

            # Metadata
            ws["A3"] = "SALDO INICIAL"
            ws["A3"].font = Font(bold=True, size=10, color="666666")
            ws["B3"] = s_ini
            ws["B3"].number_format = formato_moneda
            ws["B3"].font = Font(bold=True, size=11)
            ws["B3"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

            ws["A4"] = "SALDO FINAL"
            ws["A4"].font = Font(bold=True, size=10, color="666666")
            ws["B4"] = s_fin
            ws["B4"].number_format = formato_moneda
            ws["B4"].font = Font(bold=True, size=11)
            ws["B4"].border = Border(bottom=Side(style='thin', color="DDDDDD"))
            
            ws["D3"] = "TITULAR"; 
            ws.merge_cells("E3:G3"); ws["E3"] = clean_for_excel(titular_global)
            ws["E3"].alignment = Alignment(horizontal='center')

            ws["D4"] = "PERÍODO"; 
            ws.merge_cells("E4:G4"); ws["E4"] = clean_for_excel(periodo_global)
            ws["E4"].alignment = Alignment(horizontal='center')
            
            ws["D6"] = "CONTROL DE SALDOS"
            
            # Control Formula Placeholder
            ws["D7"] = 0
            ws["D7"].font = Font(bold=True, size=12); ws["D7"].border = thin_border
            ws.conditional_formatting.add('D7', CellIsRule(operator='notEqual', formula=['0'], stopIfTrue=True, fill=red_fill, font=red_font))

            # Tablas
            f_header = 10
            # Creditos
            ws.merge_cells(f"A{f_header}:C{f_header}"); ws[f"A{f_header}"] = "CRÉDITOS"
            ws[f"A{f_header}"].fill = fill_head_cred; ws[f"A{f_header}"].font = Font(bold=True, color="FFFFFF")
            ws[f"A{f_header}"].alignment = Alignment(horizontal="center", vertical="center")
            # Debitos
            ws.merge_cells(f"E{f_header}:G{f_header}"); ws[f"E{f_header}"] = "DÉBITOS"
            ws[f"E{f_header}"].fill = fill_head_deb; ws[f"E{f_header}"].font = Font(bold=True, color="FFFFFF")
            ws[f"E{f_header}"].alignment = Alignment(horizontal="center", vertical="center")
            
            # Subheaders
            for col, txt in zip(["A","B","C", "E","F","G"], ["Fecha","Descripción","Importe", "Fecha","Descripción","Importe"]):
                ws[f"{col}{f_header+1}"] = txt
                ws[f"{col}{f_header+1}"].border = thin_border
                ws[f"{col}{f_header+1}"].alignment = Alignment(horizontal='center')
                if col in ["A","B","C"]: ws[f"{col}{f_header+1}"].fill = fill_col_cred
                else: ws[f"{col}{f_header+1}"].fill = fill_col_deb
            
            # Llenar Creditos
            row = f_header + 2
            start_cred = row
            if creditos.empty:
                ws[f"A{row}"] = "SIN MOVIMIENTOS"; ws.merge_cells(f"A{row}:C{row}")
                ws[f"A{row}"].alignment = Alignment(horizontal='center'); ws[f"A{row}"].font = Font(italic=True, color="666666")
                row += 1
            else:
                for _, r in creditos.iterrows():
                    ws[f"A{row}"] = r["Fecha"]; ws[f"B{row}"] = r["Descripcion"]; ws[f"C{row}"] = r["Importe"]
                    ws[f"C{row}"].number_format = formato_moneda
                    for c in ["A","B","C"]: ws[f"{c}{row}"].border = thin_border; ws[f"{c}{row}"].fill = fill_row_cred
                    row += 1
            
            total_cred_row = row
            ws.merge_cells(f"A{total_cred_row}:B{total_cred_row}")
            ws[f"A{total_cred_row}"] = "TOTAL CRÉDITOS"
            ws[f"A{total_cred_row}"].font = Font(bold=True); ws[f"A{total_cred_row}"].alignment = Alignment(horizontal='right')
            ws[f"C{total_cred_row}"] = f"=SUM(C{start_cred}:C{total_cred_row-1})"
            ws[f"C{total_cred_row}"].number_format = formato_moneda; ws[f"C{total_cred_row}"].font = Font(bold=True)
            for c in ["A","B","C"]: ws[f"{c}{total_cred_row}"].border = thin_border
            
            # Llenar Debitos
            row = f_header + 2
            start_deb = row
            if debitos.empty:
                ws[f"E{row}"] = "SIN MOVIMIENTOS"; ws.merge_cells(f"E{row}:G{row}")
                ws[f"E{row}"].alignment = Alignment(horizontal='center'); ws[f"E{row}"].font = Font(italic=True, color="666666")
                row += 1
            else:
                for _, r in debitos.iterrows():
                    ws[f"E{row}"] = r["Fecha"]; ws[f"F{row}"] = r["Descripcion"]; ws[f"G{row}"] = r["Importe"]
                    ws[f"G{row}"].number_format = formato_moneda
                    for c in ["E","F","G"]: ws[f"{c}{row}"].border = thin_border; ws[f"{c}{row}"].fill = fill_row_deb
                    row += 1
            
            total_deb_row = row
            ws.merge_cells(f"E{total_deb_row}:F{total_deb_row}")
            ws[f"E{total_deb_row}"] = "TOTAL DÉBITOS"
            ws[f"E{total_deb_row}"].font = Font(bold=True); ws[f"E{total_deb_row}"].alignment = Alignment(horizontal='right')
            ws[f"G{total_deb_row}"] = f"=SUM(G{start_deb}:G{total_deb_row-1})"
            ws[f"G{total_deb_row}"].number_format = formato_moneda; ws[f"G{total_deb_row}"].font = Font(bold=True)
            for c in ["E","F","G"]: ws[f"{c}{total_deb_row}"].border = thin_border

            # Update Control Formula final
            ws["D7"] = f"=ROUND(B3+C{total_cred_row}-G{total_deb_row}-B4, 2)"
            ws["D7"].number_format = formato_moneda
            
            # Anchos
            ws.column_dimensions["B"].width = 40; ws.column_dimensions["F"].width = 40
            ws.column_dimensions["C"].width = 18; ws.column_dimensions["G"].width = 18

        # Crear hoja Pesos
        crear_hoja_dashboard(wb, "Pesos", datos_pesos, saldo_ini_pesos, saldo_fin_pesos, formato_moneda='"$ "#,##0.00')
        
        # Crear hoja Dolares
        if datos_dolares or saldo_ini_dolares != 0 or saldo_fin_dolares != 0:
            crear_hoja_dashboard(wb, "Dolares", datos_dolares, saldo_ini_dolares, saldo_fin_dolares, formato_moneda='"U$S "#,##0.00')

        wb.save(output)
        output.seek(0)
        return output.getvalue()

    except Exception as e:
        import traceback
        st.error(f"Error al procesar el archivo: {str(e)}")
        print(traceback.format_exc())
        return None
