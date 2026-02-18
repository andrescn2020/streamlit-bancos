"""
Procesador de extractos PDF de Banco Comafi.
Extrae movimientos de todas las cuentas (CC Pesos, CA Pesos, CA Dólares)
y genera un Excel dashboard con créditos/débitos separados y fórmulas de control.
"""

import streamlit as st
import pdfplumber
import pandas as pd
import re
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ── Utilidades ──────────────────────────────────────────────

def clean_for_excel(text):
    """Elimina caracteres ilegales para Excel."""
    if not text:
        return ""
    return re.sub(r'[\000-\010]|[\013-\014]|[\016-\037]', '', str(text)).strip()


def parse_ar_number(s):
    """Convierte '1.234.567,89' → 1234567.89 (formato argentino).
    También maneja signo negativo al final: '63.670,58-' → -63670.58"""
    neg = s.endswith('-')
    if neg:
        s = s[:-1]
    s = s.replace(".", "").replace(",", ".")
    val = float(s)
    return -val if neg else val


# Regex para montos argentinos: 1.234.567,89 ó 567,89 ó 0,00  (opcionalmente con '-' al final)
RE_MONTO = re.compile(r'(?<!\d)(\d{1,3}(?:\.\d{3})*,\d{2}-?)(?!\d)')

# Regex para fecha DD/MM/YY
RE_FECHA = re.compile(r'^(\d{2}/\d{2}/\d{2})\s+(.+)')

# Regex para detectar encabezado de sección de cuenta
RE_CUENTA_HEADER = re.compile(
    r'(CUENTA CORRIENTE BANCARIA|CUENTA CORRIENTE ESPECIAL|CAJA DE AHORROS|CUENTA DE LA SEGURIDAD SOCIAL)\s+EN\s+(PESOS|DOLARES|DÓLARES)',
    re.IGNORECASE
)

# Regex para número de cuenta  NNNN-NNNNN-N
RE_NRO_CUENTA = re.compile(r'(\d{4}-\d{5}-\d)')

# Regex para saldo al cierre
RE_SALDO_AL = re.compile(r'Saldo al:\s*\d{2}/\d{2}/\d{4}\s+([\d.,]+)', re.IGNORECASE)

# Keywords que indican CRÉDITO
KEYWORDS_CREDITO = [
    "TRANSFERENCIA RECIBIDA",
    "TRANSF. INMEDIATA RECIBIDA",
    "CRÉDITO RESCATE",
    "CREDITO RESCATE",
    "ACREDITACION DE PLAZO FIJO",
    "ACRED. INTERESES",
    "COBRANZA BURSATIL",
    "TRANSFERENCIA RECIBIDA DATANET",
    "ANSES",
    "DEPOSITO DE CHEQUES",
    "DEV. IMP.",
    "AJUSTE BANELCO",
    "DEPÓSITO EFECTIVO",
    "DEPOSITO EFECTIVO",
    "DEPOSITO DE EFECTIVO",
]

# Keywords que indican DÉBITO
KEYWORDS_DEBITO = [
    "PAGO ELECTRONICO",
    "PAGO DE SERVICIOS",
    "TRANSF. INMEDIATA ENVIADA",
    "TRANSFERENCIA ENVIADA",
    "TRANSF. INMEDIATA ENVIADA DIFE",
    "TRANSFERENCIA TERCEROS",
    "TRANSF INMED SUELDOS",
    "DEBITO INMED ENVIADO",
    "DEBITO INMEDIATO ENVIADO",
    "DEBITO DEBIN",
    "DEBITO AUTOM",
    "DEBITO TARJETA",
    "DÉBITO SUSCRIPCIÓN",
    "DEBITO SUSCRIPCION",
    "DÉBITO POR RECAUDACIÓN",
    "DEBITO POR RECAUDACION",
    "COMISIÓN",
    "COMISION",
    "IMPUESTO",
    "IVA",
    "PERCEPCION",
    "CONSTITUCION DE PLAZO FIJO",
    "EXTRACCION CAJERO",
    "CHEQUE DE CAMARA RECHAZADO",
    "TRANSF INMED. CUENTAS PROPIAS",
    "TRANSFERENCIA SUELDOS",
    "IMP. IB",
    "INTERESES POR ACUERDO",
    "IMPUESTO A LOS SELLOS",
    "EXTRACCION DE EFECTIVO",
]


def clasificar_movimiento(concepto):
    """Devuelve 'credito', 'debito' o None según keywords."""
    upper = concepto.upper()
    for kw in KEYWORDS_CREDITO:
        if kw in upper:
            return "credito"
    for kw in KEYWORDS_DEBITO:
        if kw in upper:
            return "debito"
    return None


def _last_known_saldo(movs, saldo_ini):
    """Recorre los movimientos hacia atrás buscando el último Saldo conocido (no None)."""
    for mov in reversed(movs):
        if mov["Saldo"] is not None:
            return mov["Saldo"]
    return saldo_ini


# ── Parser principal ────────────────────────────────────────

def procesar_comafi(archivo_pdf):
    """
    Lee el PDF de Banco Comafi, extrae los movimientos de cada cuenta
    y genera un Excel dashboard con estilos Comafi.
    """
    st.info("Procesando Banco Comafi…")

    try:
        archivo_pdf.seek(0)

        # ── Extraer texto completo ──
        all_text = ""
        raw_bytes = archivo_pdf.read()
        with pdfplumber.open(io.BytesIO(raw_bytes)) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    all_text += text + "\n"

        lines = all_text.splitlines()

        # ── FASE 1: Metadata global ──
        titular = "S/D"
        periodo = ""

        # Titular: está en la primera página, línea con "Hoja:1/"
        for line in lines[:20]:
            m = re.search(r'^(.+?)\s+Hoja:\s*1/', line)
            if m:
                titular = m.group(1).strip()
                break

        # Periodo: "ENERO - 2025" etc.
        meses_map = {
            "ENERO": "01", "FEBRERO": "02", "MARZO": "03", "ABRIL": "04",
            "MAYO": "05", "JUNIO": "06", "JULIO": "07", "AGOSTO": "08",
            "SEPTIEMBRE": "09", "OCTUBRE": "10", "NOVIEMBRE": "11", "DICIEMBRE": "12"
        }
        for line in lines[:10]:
            m = re.search(r'(ENERO|FEBRERO|MARZO|ABRIL|MAYO|JUNIO|JULIO|AGOSTO|SEPTIEMBRE|OCTUBRE|NOVIEMBRE|DICIEMBRE)\s*-\s*(\d{4})', line, re.IGNORECASE)
            if m:
                periodo = f"{m.group(1).capitalize()} {m.group(2)}"
                break

        # ── FASE 2: Identificar secciones de cuentas ──
        # Cada cuenta empieza con un header como:
        #   "Servicio Comafi ... CUENTA CORRIENTE BANCARIA EN PESOS ."
        #   o "Servicio Comafi ... CAJA DE AHORROS EN PESOS ."
        #   o "Servicio Comafi ... CAJA DE AHORROS EN DOLARES ."
        # Seguido de "NRO. NNNN-NNNNN-N ..."
        # Y luego "DETALLE DE MOVIMIENTOS"

        cuentas_info = {}   # {nro_cuenta: {tipo, moneda, movimientos, saldo_ini, saldo_fin}}

        current_account = None
        in_movements = False
        last_date = ""
        pending_continuation = False  # para líneas de continuación

        i = 0
        while i < len(lines):
            line = lines[i].strip()
            i += 1

            if not line:
                continue

            # ── Detectar encabezado de sección de cuenta ──
            m_header = RE_CUENTA_HEADER.search(line)
            if m_header:
                tipo_cuenta = m_header.group(1).strip().title()
                moneda_raw = m_header.group(2).strip().upper()
                moneda = "Dólares" if moneda_raw in ("DOLARES", "DÓLARES") else "Pesos"

                # Buscar NRO en las siguientes líneas
                nro_cuenta = None
                for j in range(i, min(i + 5, len(lines))):
                    m_nro = re.search(r'NRO\.?\s*(\d{4}-\d{5}-\d)', lines[j], re.IGNORECASE)
                    if not m_nro:
                        m_nro = re.search(r'Número\s+(\d{4}-\d{5}-\d)', lines[j], re.IGNORECASE)
                    if m_nro:
                        nro_cuenta = m_nro.group(1)
                        break

                if nro_cuenta and nro_cuenta not in cuentas_info:
                    cuentas_info[nro_cuenta] = {
                        "tipo": tipo_cuenta,
                        "moneda": moneda,
                        "movimientos": [],
                        "saldo_ini": 0.0,
                        "saldo_fin": 0.0,
                    }
                if nro_cuenta:
                    current_account = nro_cuenta
                in_movements = False
                continue

            # ── Detectar inicio de "DETALLE DE MOVIMIENTOS" ──
            if "DETALLE DE MOVIMIENTOS" in line.upper():
                in_movements = True
                continue

            # Saltar líneas de guiones y encabezados de columna
            if line.startswith("---") or line.startswith("___"):
                continue

            if re.match(r'^Fecha\s+Conceptos\s+Referencias', line, re.IGNORECASE):
                continue

            # ── Detectar fin de movimientos ──
            # "Saldo al: DD/MM/YYYY ..."
            m_saldo_fin = RE_SALDO_AL.search(line)
            if m_saldo_fin and current_account and current_account in cuentas_info:
                saldo_final = parse_ar_number(m_saldo_fin.group(1))
                cuentas_info[current_account]["saldo_fin"] = saldo_final
                in_movements = False
                continue

            # ══════════════════════════════════════════════════════════
            # IMPORTANTE: Parsear movimientos ANTES del chequeo de
            # secciones, porque keywords como "PAGO DE SERVICIOS"
            # aparecen tanto en headers de sección como en conceptos
            # de movimientos. Las líneas con fecha siempre son movs.
            # ══════════════════════════════════════════════════════════

            if in_movements and current_account:
                # ── Líneas a ignorar dentro de movimientos ──
                if line.upper().startswith("TRANSPORTE"):
                    continue
                if line.upper().startswith("SIN MOVIMIENTOS"):
                    continue
                # Encabezados de página repetidos (ej: "138.065 - 3/7 - 02 ...")
                if re.match(r'^\d+\.\d+\s*-\s*\d+/\d+', line):
                    continue
                if re.match(r'^Hoja:\s*\d+/\d+', line, re.IGNORECASE):
                    continue
                # Barcodes / códigos basura
                if line.startswith('<') or re.match(r'^\d{10,}$', line):
                    continue
                # Líneas tipo "Página N" (del texto pegado, no del PDF real)
                if re.match(r'^Página\s+\d+', line, re.IGNORECASE):
                    continue

                # ── Línea con fecha = nuevo movimiento ──
                m_fecha = RE_FECHA.match(line)

                if m_fecha:
                    fecha_str = m_fecha.group(1)
                    resto = m_fecha.group(2)

                    # Saldo Anterior
                    if "SALDO ANTERIOR" in resto.upper():
                        montos = RE_MONTO.findall(resto)
                        if montos:
                            saldo_ini = parse_ar_number(montos[-1])
                            cuentas_info[current_account]["saldo_ini"] = saldo_ini
                        last_date = fecha_str
                        continue

                    last_date = fecha_str

                    # Extraer montos de la línea
                    montos = RE_MONTO.findall(resto)

                    # Limpiar concepto (quitar montos del final)
                    concepto = resto
                    for m in reversed(montos):
                        idx = concepto.rfind(m)
                        if idx != -1:
                            concepto = concepto[:idx]
                    concepto = concepto.strip()

                    if len(montos) >= 2:
                        # Último = saldo, penúltimo = importe
                        importe = parse_ar_number(montos[-2])
                        saldo = parse_ar_number(montos[-1])
                        tipo = clasificar_movimiento(concepto)

                        # Si no se pudo clasificar por keyword, usar saldo anterior
                        if tipo is None:
                            movs = cuentas_info[current_account]["movimientos"]
                            saldo_prev = _last_known_saldo(movs, cuentas_info[current_account]["saldo_ini"])
                            diff_deb = abs((saldo_prev - importe) - saldo)
                            diff_cred = abs((saldo_prev + importe) - saldo)
                            tipo = "credito" if diff_cred < diff_deb else "debito"

                        cuentas_info[current_account]["movimientos"].append({
                            "Fecha": fecha_str,
                            "Descripcion": concepto,
                            "Debito": importe if tipo == "debito" else 0.0,
                            "Credito": importe if tipo == "credito" else 0.0,
                            "Saldo": saldo,
                        })

                    elif len(montos) == 1:
                        # Solo importe, sin saldo en esta línea
                        importe = parse_ar_number(montos[0])
                        tipo = clasificar_movimiento(concepto)

                        if tipo is None:
                            tipo = "debito"  # default conservador

                        cuentas_info[current_account]["movimientos"].append({
                            "Fecha": fecha_str,
                            "Descripcion": concepto,
                            "Debito": importe if tipo == "debito" else 0.0,
                            "Credito": importe if tipo == "credito" else 0.0,
                            "Saldo": None,  # se completará cuando aparezca
                        })

                    else:
                        # Línea con fecha pero sin monto → inicio de multi-línea
                        cuentas_info[current_account]["movimientos"].append({
                            "Fecha": fecha_str,
                            "Descripcion": concepto,
                            "Debito": 0.0,
                            "Credito": 0.0,
                            "Saldo": None,
                        })
                    continue  # movimiento procesado, siguiente línea

                # ── Línea de continuación (sin fecha) ──
                movs = cuentas_info[current_account]["movimientos"]
                if movs:
                    montos = RE_MONTO.findall(line)

                    # Limpiar texto de continuación (quitar montos)
                    texto_cont = line
                    for m in reversed(montos):
                        idx = texto_cont.rfind(m)
                        if idx != -1:
                            texto_cont = texto_cont[:idx]
                    texto_cont = texto_cont.strip()

                    last_mov = movs[-1]

                    # Agregar texto al concepto
                    if texto_cont and not texto_cont.startswith("---"):
                        last_mov["Descripcion"] += " " + texto_cont

                    if len(montos) >= 2:
                        importe = parse_ar_number(montos[-2])
                        saldo = parse_ar_number(montos[-1])

                        tipo = clasificar_movimiento(last_mov["Descripcion"])
                        if tipo is None:
                            saldo_prev = _last_known_saldo(movs[:-1], cuentas_info[current_account]["saldo_ini"])
                            diff_deb = abs((saldo_prev - importe) - saldo)
                            diff_cred = abs((saldo_prev + importe) - saldo)
                            tipo = "credito" if diff_cred < diff_deb else "debito"

                        last_mov["Debito"] = importe if tipo == "debito" else 0.0
                        last_mov["Credito"] = importe if tipo == "credito" else 0.0
                        last_mov["Saldo"] = saldo

                    elif len(montos) == 1 and last_mov["Debito"] == 0 and last_mov["Credito"] == 0:
                        importe = parse_ar_number(montos[0])
                        tipo = clasificar_movimiento(last_mov["Descripcion"])
                        if tipo is None:
                            tipo = "debito"
                        last_mov["Debito"] = importe if tipo == "debito" else 0.0
                        last_mov["Credito"] = importe if tipo == "credito" else 0.0
                    continue  # continuación procesada

            # ── Detección de fin de sección (solo para líneas NO-movimiento) ──
            # Se evalúa DESPUÉS de intentar parsear como movimiento, así
            # keywords como "PAGO DE SERVICIOS" dentro de un concepto de
            # transacción no cortan la sección prematuramente.
            if any(kw in line.upper() for kw in [
                "IMPUESTOS DEBITADOS", "TRANSFERENCIAS ELECTRONICAS",
                "VISA DEBITO", "TARJETAS DE CREDITO", "PLAZO FIJO",
                "BONIFICACIONES Y PROMOCIONES", "CAJA DE SEGURIDAD",
                "ACUERDOS VIGENTES", "FONDOS COMUNES", "DETALLE DE CHEQUES",
                "RESUMEN DE SALDO",
                "Los depósitos en pesos"
            ]):
                in_movements = False
                continue

        # ── Completar saldos faltantes (running balance) ──
        for nro, info in cuentas_info.items():
            movs = info["movimientos"]
            running = info["saldo_ini"]
            for mov in movs:
                running = running + mov["Credito"] - mov["Debito"]
                if mov["Saldo"] is None:
                    mov["Saldo"] = round(running, 2)
                else:
                    running = mov["Saldo"]  # reset al saldo conocido

        # ── FASE 3: Generar Excel ──
        if not any(info["movimientos"] for info in cuentas_info.values()):
            st.warning("No se extrajeron movimientos de ninguna cuenta.")
            return None

        output = io.BytesIO()
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        # Estilos Comafi (azul oscuro)
        color_comafi = "003366"
        fill_header = PatternFill(start_color=color_comafi, end_color=color_comafi, fill_type="solid")
        font_header = Font(color="FFFFFF", bold=True, size=12)
        font_bold = Font(bold=True)
        thin_border = Border(
            left=Side(style='thin', color="A6A6A6"),
            right=Side(style='thin', color="A6A6A6"),
            top=Side(style='thin', color="A6A6A6"),
            bottom=Side(style='thin', color="A6A6A6"),
        )
        fill_deb_h = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        fill_deb_c = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
        fill_cred_h = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        fill_cred_c = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")

        for nro_cuenta, info in cuentas_info.items():
            movs = info["movimientos"]
            tipo = info["tipo"]
            moneda = info["moneda"]

            # Nombre de hoja
            if moneda == "Dólares":
                safe_name = f"USD {nro_cuenta}"
                fmt_moneda = '"u$s "#,##0.00'
            else:
                safe_name = f"ARS {nro_cuenta}"
                fmt_moneda = '"$ "#,##0.00'

            ws = wb.create_sheet(title=clean_for_excel(safe_name)[:31])
            ws.sheet_view.showGridLines = False

            ws.column_dimensions['A'].width = 14
            ws.column_dimensions['B'].width = 55
            ws.column_dimensions['C'].width = 18
            ws.column_dimensions['D'].width = 4
            ws.column_dimensions['E'].width = 14
            ws.column_dimensions['F'].width = 55
            ws.column_dimensions['G'].width = 18
            ws.column_dimensions['I'].width = 22
            ws.column_dimensions['J'].width = 28

            # ── Título ──
            ws.merge_cells("A1:G1")
            ws["A1"] = f"REPORTE COMAFI — {tipo} en {moneda} — {nro_cuenta}"
            ws["A1"].fill = fill_header
            ws["A1"].font = font_header
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 28

            s_ini = info["saldo_ini"]
            s_fin = info["saldo_fin"]

            ws["A3"] = "SALDO INICIAL"
            ws["A3"].font = Font(bold=True, color="666666")
            ws["B3"] = s_ini
            ws["B3"].number_format = fmt_moneda
            ws["B3"].font = font_bold

            ws["A4"] = "SALDO FINAL"
            ws["A4"].font = Font(bold=True, color="666666")
            ws["B4"] = s_fin
            ws["B4"].number_format = fmt_moneda
            ws["B4"].font = font_bold

            ws["I3"] = "TITULAR"
            ws["I3"].font = Font(bold=True, color="666666")
            ws["J3"] = titular
            ws["J3"].font = font_bold

            ws["I4"] = "PERIODO"
            ws["I4"].font = Font(bold=True, color="666666")
            ws["J4"] = periodo
            ws["J4"].font = font_bold

            ws["I5"] = "CUENTA"
            ws["I5"].font = Font(bold=True, color="666666")
            ws["J5"] = nro_cuenta
            ws["J5"].font = font_bold

            if not movs:
                ws.merge_cells("A7:G7")
                ws["A7"] = "NO HUBO MOVIMIENTOS EN ESTE PERIODO"
                ws["A7"].font = Font(italic=True, color="666666", size=11)
                ws["A7"].alignment = Alignment(horizontal="center")
                continue

            df = pd.DataFrame(movs)
            creditos = df[df["Credito"] > 0]
            debitos = df[df["Debito"] > 0]

            fila = 7
            headers = ["Fecha", "Descripción", "Importe"]

            # ── CRÉDITOS (A-C) ──
            ws.merge_cells(f"A{fila}:C{fila}")
            ws[f"A{fila}"] = "CRÉDITOS"
            ws[f"A{fila}"].fill = fill_cred_h
            ws[f"A{fila}"].font = Font(color="FFFFFF", bold=True)
            ws[f"A{fila}"].alignment = Alignment(horizontal="center")

            for idx_h, h in enumerate(headers):
                c = ws.cell(row=fila + 1, column=1 + idx_h, value=h)
                c.fill = fill_cred_c
                c.font = font_bold
                c.border = thin_border
                c.alignment = Alignment(horizontal="center")

            r_cred = fila + 2
            start_cred = r_cred
            if creditos.empty:
                ws[f"A{r_cred}"] = "SIN MOVIMIENTOS"
                ws[f"A{r_cred}"].font = Font(italic=True, color="999999")
                r_cred += 1
            else:
                for _, row in creditos.iterrows():
                    ws[f"A{r_cred}"] = row["Fecha"]
                    ws[f"A{r_cred}"].alignment = Alignment(horizontal="center")
                    ws[f"B{r_cred}"] = clean_for_excel(row["Descripcion"])
                    ws[f"C{r_cred}"] = row["Credito"]
                    ws[f"C{r_cred}"].number_format = fmt_moneda
                    for col in ["A", "B", "C"]:
                        ws[f"{col}{r_cred}"].border = thin_border
                        if r_cred % 2 == 0:
                            ws[f"{col}{r_cred}"].fill = fill_cred_c
                    r_cred += 1

            ws[f"B{r_cred}"] = "TOTAL CRÉDITOS"
            ws[f"B{r_cred}"].font = font_bold
            ws[f"B{r_cred}"].alignment = Alignment(horizontal="right")
            ref_cred = f"C{r_cred}"
            if not creditos.empty:
                ws[f"C{r_cred}"] = f"=SUM(C{start_cred}:C{r_cred - 1})"
            else:
                ws[f"C{r_cred}"] = 0
            ws[f"C{r_cred}"].number_format = fmt_moneda
            ws[f"C{r_cred}"].font = font_bold

            # ── DÉBITOS (E-G) ──
            ws.merge_cells(f"E{fila}:G{fila}")
            ws[f"E{fila}"] = "DÉBITOS"
            ws[f"E{fila}"].fill = fill_deb_h
            ws[f"E{fila}"].font = Font(color="FFFFFF", bold=True)
            ws[f"E{fila}"].alignment = Alignment(horizontal="center")

            for idx_h, h in enumerate(headers):
                c = ws.cell(row=fila + 1, column=5 + idx_h, value=h)
                c.fill = fill_deb_c
                c.font = font_bold
                c.border = thin_border
                c.alignment = Alignment(horizontal="center")

            r_deb = fila + 2
            start_deb = r_deb
            if debitos.empty:
                ws[f"E{r_deb}"] = "SIN MOVIMIENTOS"
                ws[f"E{r_deb}"].font = Font(italic=True, color="999999")
                r_deb += 1
            else:
                for _, row in debitos.iterrows():
                    ws[f"E{r_deb}"] = row["Fecha"]
                    ws[f"E{r_deb}"].alignment = Alignment(horizontal="center")
                    ws[f"F{r_deb}"] = clean_for_excel(row["Descripcion"])
                    ws[f"G{r_deb}"] = row["Debito"]
                    ws[f"G{r_deb}"].number_format = fmt_moneda
                    for col in ["E", "F", "G"]:
                        ws[f"{col}{r_deb}"].border = thin_border
                        if r_deb % 2 == 0:
                            ws[f"{col}{r_deb}"].fill = fill_deb_c
                    r_deb += 1

            ws[f"F{r_deb}"] = "TOTAL DÉBITOS"
            ws[f"F{r_deb}"].font = font_bold
            ws[f"F{r_deb}"].alignment = Alignment(horizontal="right")
            ref_deb = f"G{r_deb}"
            if not debitos.empty:
                ws[f"G{r_deb}"] = f"=SUM(G{start_deb}:G{r_deb - 1})"
            else:
                ws[f"G{r_deb}"] = 0
            ws[f"G{r_deb}"].number_format = fmt_moneda
            ws[f"G{r_deb}"].font = font_bold

            # ── CONTROL ──
            ws["I7"] = "CONTROL (debe ser 0)"
            ws["I7"].font = Font(bold=True, color="666666")
            ws["I7"].alignment = Alignment(horizontal="center")

            formula = f"=ROUND(B3 + {ref_cred} - {ref_deb} - B4, 2)"
            ws["I8"] = formula
            ws["I8"].font = Font(bold=True, size=14)
            ws["I8"].alignment = Alignment(horizontal="center")
            ws["I8"].number_format = fmt_moneda
            ws["I8"].border = thin_border

        wb.save(output)
        output.seek(0)
        st.success(f"✅ Procesamiento Comafi completado — {len(cuentas_info)} cuenta(s) encontradas.")
        return output.getvalue()

    except Exception as e:
        st.error(f"Error Crítico: {str(e)}")
        import traceback
        st.code(traceback.format_exc())
        return None
