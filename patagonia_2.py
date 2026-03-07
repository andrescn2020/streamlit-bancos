import streamlit as st
import pdfplumber
import re
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule

ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

def clean_for_excel(text):
    if not text: return ""
    text = str(text)
    text = ILLEGAL_CHARACTERS_RE.sub("", text)
    return text.strip()

def parse_numero_ar(s):
    if not s: return 0.0
    s = s.strip()
    neg = False
    if s.endswith("-"):
        neg = True
        s = s[:-1]
    elif s.startswith("-"):
        neg = True
        s = s[1:]
    s = s.replace(".", "").replace(",", ".")
    try:
        val = float(s)
        return -val if neg else val
    except:
        return 0.0

def convertir_fecha_corta(fecha_str):
    """Convierte D/MM/YY o DD/MM/YY a DD/MM/YYYY"""
    try:
        partes = fecha_str.split("/")
        dia = partes[0].zfill(2)
        mes = partes[1].zfill(2)
        anio = partes[2]
        if len(anio) == 2:
            anio = "20" + anio
        return f"{dia}/{mes}/{anio}"
    except:
        return fecha_str


def find_signs(amounts, target_sum):
    """Encuentra la combinación de signos (+1/-1) para cada importe
    tal que la suma con signos sea igual a target_sum.
    Usa brute-force para grupos pequeños (hasta ~20 entradas)."""
    n = len(amounts)
    if n == 0:
        return []
    if n > 20:
        return None  # Demasiadas combinaciones

    for mask in range(2**n):
        total = 0.0
        for i in range(n):
            if mask & (1 << i):
                total += amounts[i]
            else:
                total -= amounts[i]
        if abs(total - target_sum) < 0.02:
            signs = []
            for i in range(n):
                signs.append(1 if mask & (1 << i) else -1)
            return signs
    return None


def generar_hoja(wb, cuenta_id, titular, periodo, saldo_inicial, saldo_final, transactions):
    """Genera una hoja Excel con el formato estándar para una cuenta"""
    sheet_name = clean_for_excel(cuenta_id)[:31]
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_view.showGridLines = False

    color_bg_main = "006341"
    color_txt_main = "FFFFFF"
    thin_border = Border(
        left=Side(style='thin', color="A6A6A6"),
        right=Side(style='thin', color="A6A6A6"),
        top=Side(style='thin', color="A6A6A6"),
        bottom=Side(style='thin', color="A6A6A6")
    )
    fill_head_deb = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    fill_col_deb = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
    fill_row_deb = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")
    fill_head_cred = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    fill_col_cred = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")
    fill_row_cred = PatternFill(start_color="F2F9F1", end_color="F2F9F1", fill_type="solid")

    df = pd.DataFrame(transactions)
    creditos = df[df["Importe"] > 0].copy()
    debitos = df[df["Importe"] < 0].copy()
    if not debitos.empty:
        debitos["Importe"] = debitos["Importe"].abs()

    # Header
    ws.merge_cells("A1:G1")
    tit = ws["A1"]
    tit.value = f"REPORTE BANCO PATAGONIA - CTA {clean_for_excel(cuenta_id)}"
    tit.font = Font(size=14, bold=True, color=color_txt_main)
    tit.fill = PatternFill(start_color=color_bg_main, end_color=color_bg_main, fill_type="solid")
    tit.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25

    # Metadata
    ws["A3"] = "SALDO INICIAL"
    ws["A3"].font = Font(bold=True, size=10, color="666666")
    ws["B3"] = saldo_inicial
    ws["B3"].number_format = '"$ "#,##0.00'
    ws["B3"].font = Font(bold=True, size=11)
    ws["B3"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

    ws["A4"] = "SALDO FINAL"
    ws["A4"].font = Font(bold=True, size=10, color="666666")
    ws["B4"] = saldo_final
    ws["B4"].number_format = '"$ "#,##0.00'
    ws["B4"].font = Font(bold=True, size=11)
    ws["B4"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

    ws["D3"] = "TITULAR"
    ws["D3"].alignment = Alignment(horizontal='right')
    ws["D3"].font = Font(bold=True, color="666666", size=10)
    ws["E3"] = clean_for_excel(titular)
    ws["E3"].font = Font(bold=True, size=11)
    ws["E3"].alignment = Alignment(horizontal='center')
    ws.merge_cells("E3:G3")
    for c in ["E","F","G"]: ws[f"{c}3"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

    ws["D4"] = "PERÍODO"
    ws["D4"].alignment = Alignment(horizontal='right')
    ws["D4"].font = Font(bold=True, color="666666", size=10)
    ws["E4"] = clean_for_excel(periodo)
    ws["E4"].font = Font(bold=True, size=11)
    ws["E4"].alignment = Alignment(horizontal='center')
    ws.merge_cells("E4:G4")
    for c in ["E","F","G"]: ws[f"{c}4"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

    # Tablas
    fila_inicio = 10
    f_header = fila_inicio

    ws.merge_cells(f"A{f_header}:C{f_header}")
    ws[f"A{f_header}"] = "CRÉDITOS"
    ws[f"A{f_header}"].fill = fill_head_cred
    ws[f"A{f_header}"].font = Font(bold=True, color="FFFFFF")
    ws[f"A{f_header}"].alignment = Alignment(horizontal='center')
    ws[f"A{f_header}"].border = thin_border

    ws.merge_cells(f"E{f_header}:G{f_header}")
    ws[f"E{f_header}"] = "DÉBITOS"
    ws[f"E{f_header}"].fill = fill_head_deb
    ws[f"E{f_header}"].font = Font(bold=True, color="FFFFFF")
    ws[f"E{f_header}"].alignment = Alignment(horizontal='center')
    ws[f"E{f_header}"].border = thin_border

    headers = ["Fecha", "Descripción", "Importe"]
    cols_cred = ["A", "B", "C"]
    cols_deb = ["E", "F", "G"]
    f_sub = f_header + 1

    for i, h in enumerate(headers):
        c = ws[f"{cols_cred[i]}{f_sub}"]
        c.value = h; c.fill = fill_col_cred; c.font = Font(bold=True)
        c.alignment = Alignment(horizontal='center'); c.border = thin_border
        d = ws[f"{cols_deb[i]}{f_sub}"]
        d.value = h; d.fill = fill_col_deb; d.font = Font(bold=True)
        d.alignment = Alignment(horizontal='center'); d.border = thin_border

    fila_a_llenar = f_sub + 1

    # Créditos
    f_c = fila_a_llenar
    if creditos.empty:
        ws.merge_cells(f"A{f_c}:C{f_c}")
        ws[f"A{f_c}"] = "SIN MOVIMIENTOS"; ws[f"A{f_c}"].border = thin_border
        f_c += 1
    else:
        start_c = f_c
        for _, r in creditos.iterrows():
            ws[f"A{f_c}"] = r["Fecha"]; ws[f"A{f_c}"].fill = fill_row_cred
            ws[f"A{f_c}"].border = thin_border; ws[f"A{f_c}"].alignment = Alignment(horizontal='center')
            ws[f"B{f_c}"] = r["Descripcion"]; ws[f"B{f_c}"].fill = fill_row_cred; ws[f"B{f_c}"].border = thin_border
            ws[f"C{f_c}"] = r["Importe"]; ws[f"C{f_c}"].number_format = '"$ "#,##0.00'
            ws[f"C{f_c}"].fill = fill_row_cred; ws[f"C{f_c}"].border = thin_border
            f_c += 1
        ws.merge_cells(f"A{f_c}:B{f_c}")
        ws[f"A{f_c}"] = "TOTAL CRÉDITOS"; ws[f"A{f_c}"].font = Font(bold=True)
        ws[f"A{f_c}"].alignment = Alignment(horizontal='right')
        ws[f"C{f_c}"] = f"=SUM(C{start_c}:C{f_c-1})"
        ws[f"C{f_c}"].font = Font(bold=True); ws[f"C{f_c}"].number_format = '"$ "#,##0.00'
        f_c += 1

    # Débitos
    f_d = fila_a_llenar
    if debitos.empty:
        ws.merge_cells(f"E{f_d}:G{f_d}")
        ws[f"E{f_d}"] = "SIN MOVIMIENTOS"; ws[f"E{f_d}"].border = thin_border
        f_d += 1
    else:
        start_d = f_d
        for _, r in debitos.iterrows():
            ws[f"E{f_d}"] = r["Fecha"]; ws[f"E{f_d}"].fill = fill_row_deb
            ws[f"E{f_d}"].border = thin_border; ws[f"E{f_d}"].alignment = Alignment(horizontal='center')
            ws[f"F{f_d}"] = r["Descripcion"]; ws[f"F{f_d}"].fill = fill_row_deb; ws[f"F{f_d}"].border = thin_border
            ws[f"G{f_d}"] = r["Importe"]; ws[f"G{f_d}"].number_format = '"$ "#,##0.00'
            ws[f"G{f_d}"].fill = fill_row_deb; ws[f"G{f_d}"].border = thin_border
            f_d += 1
        ws.merge_cells(f"E{f_d}:F{f_d}")
        ws[f"E{f_d}"] = "TOTAL DÉBITOS"; ws[f"E{f_d}"].font = Font(bold=True)
        ws[f"E{f_d}"].alignment = Alignment(horizontal='right')
        ws[f"G{f_d}"] = f"=SUM(G{start_d}:G{f_d-1})"
        ws[f"G{f_d}"].font = Font(bold=True); ws[f"G{f_d}"].number_format = '"$ "#,##0.00'
        f_d += 1

    # Control de Saldos
    ws["D6"] = "CONTROL DE SALDOS"
    ws["D6"].font = Font(bold=True, size=10, color="666666")
    ws["D6"].alignment = Alignment(horizontal='center')
    ref_tot_c = f"C{f_c-1}" if not creditos.empty else "0"
    ref_tot_d = f"G{f_d-1}" if not debitos.empty else "0"
    ws["D7"] = f"=ROUND(B3+{ref_tot_c}-{ref_tot_d}-B4, 2)"
    ws["D7"].number_format = '"$ "#,##0.00'
    ws["D7"].font = Font(bold=True)
    ws["D7"].alignment = Alignment(horizontal='center')
    ws["D7"].border = thin_border
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    red_font = Font(color='9C0006', bold=True)
    ws.conditional_formatting.add('D7', CellIsRule(operator='notEqual', formula=['0'], stopIfTrue=True, fill=red_fill, font=red_font))

    # Anchos
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 45
    ws.column_dimensions["G"].width = 18


def procesar_patagonia_formato_2(archivo_pdf):
    """Procesa extractos de Banco Patagonia - Formato 2"""
    st.info("Procesando archivo del Banco Patagonia (Formato 2)...")
    try:
        archivo_pdf.seek(0)
        with pdfplumber.open(io.BytesIO(archivo_pdf.read())) as pdf:
            texto_completo = ""
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    texto_completo += t + "\n"

        lineas = texto_completo.splitlines()

        # 1. METADATOS
        titular = "Sin Especificar"
        for line in lineas:
            m_cuit = re.match(r'C\.U\.I\.T\.\s+\d+\s+(.+)', line.strip())
            if m_cuit:
                titular = m_cuit.group(1).strip()
                break

        # 2. DETECTAR SECCIONES DE CUENTAS
        patron_cuenta = re.compile(
            r'(?:CUENTA CORRIENTE EN PESOS|CCTE ESP PJ[\w\s/]*)\s+(\d+)\s+SUBCTA\s+(\d+)\s+SUC\s+(\d+)\s+CBU:\s*(\S+)'
        )
        secciones_raw = []
        for i, line in enumerate(lineas):
            m = patron_cuenta.search(line.strip())
            if m:
                secciones_raw.append({
                    "cuenta": m.group(1), "subcta": m.group(2),
                    "suc": m.group(3), "cbu": m.group(4), "start_idx": i
                })

        if not secciones_raw:
            st.warning("No se encontraron cuentas en el PDF.")
            return None

        # Agrupar secciones por número de cuenta
        from collections import OrderedDict
        cuentas_agrupadas = OrderedDict()
        for sec in secciones_raw:
            cta = sec["cuenta"]
            if cta not in cuentas_agrupadas:
                cuentas_agrupadas[cta] = []
            cuentas_agrupadas[cta].append(sec)

        # 3. REGEX MOVIMIENTOS
        patron_mov = re.compile(
            r'^(\d{1,2}/\d{2}/\d{2})\s+'
            r'(.+?)\s+'
            r'([\d.]+,\d{2})'
            r'(?:\s+([\d.]+,\d{2}))?$'
        )
        patron_saldo_ant = re.compile(r'^(?:\d{1,2}/\d{2}/\d{2}|0/00/00)\s+SALDO ANTERIOR\s+([\d.]+,\d{2})')
        patron_saldo_act = re.compile(r'^(\d{1,2}/\d{2}/\d{2})\s+SALDO ACTUAL\s+([\d.]+,\d{2})')

        # 4. PROCESAR CADA CUENTA (agrupada por número)
        output = io.BytesIO()
        wb = Workbook()
        wb.remove(wb.active)
        total_movimientos = 0

        all_secciones = list(secciones_raw)

        for cuenta_id, secciones in cuentas_agrupadas.items():
            # Parsear movimientos de TODAS las secciones de esta cuenta
            all_movimientos_raw = []
            primer_saldo_inicial = None
            ultimo_saldo_final = 0.0

            for sec in secciones:
                sec_start = sec["start_idx"]
                # Encontrar el fin de esta sección
                sec_idx_in_all = all_secciones.index(sec)
                sec_end = all_secciones[sec_idx_in_all + 1]["start_idx"] if sec_idx_in_all + 1 < len(all_secciones) else len(lineas)
                seccion_lineas = lineas[sec_start:sec_end]

                in_movements = False
                for line in seccion_lineas:
                    line_s = line.strip()
                    if not line_s:
                        continue

                    m_sa = patron_saldo_ant.match(line_s)
                    if m_sa:
                        saldo_val = parse_numero_ar(m_sa.group(1))
                        if primer_saldo_inicial is None:
                            primer_saldo_inicial = saldo_val
                        # Insertar marcador de saldo anterior como checkpoint
                        all_movimientos_raw.append({
                            "fecha": "", "descripcion": "__SALDO_ANT__",
                            "importe_str": "0", "saldo": saldo_val,
                            "tiene_saldo": True, "_es_marcador": True
                        })
                        in_movements = True
                        continue

                    m_sf = patron_saldo_act.match(line_s)
                    if m_sf:
                        ultimo_saldo_final = parse_numero_ar(m_sf.group(2))
                        in_movements = False
                        continue

                    if not in_movements:
                        continue

                    m_mov = patron_mov.match(line_s)
                    if m_mov:
                        all_movimientos_raw.append({
                            "fecha": convertir_fecha_corta(m_mov.group(1)),
                            "descripcion": m_mov.group(2).strip(),
                            "importe_str": m_mov.group(3),
                            "saldo": parse_numero_ar(m_mov.group(4)) if m_mov.group(4) else None,
                            "tiene_saldo": m_mov.group(4) is not None,
                            "_es_marcador": False
                        })

            if primer_saldo_inicial is None:
                primer_saldo_inicial = 0.0

            # Filtrar marcadores y procesar
            movimientos_raw = [m for m in all_movimientos_raw if not m.get("_es_marcador", False)]

            # DETERMINAR DÉBITO/CRÉDITO con verificación matemática
            transactions = []
            prev_saldo = primer_saldo_inicial
            i = 0

            while i < len(movimientos_raw):
                mov = movimientos_raw[i]

                if mov["tiene_saldo"]:
                    importe = round(mov["saldo"] - prev_saldo, 2)
                    prev_saldo = mov["saldo"]
                    transactions.append({
                        "Fecha": mov["fecha"],
                        "Descripcion": clean_for_excel(mov["descripcion"]),
                        "Importe": importe
                    })
                    i += 1
                else:
                    # Agrupar entradas consecutivas sin saldo
                    group = []
                    while i < len(movimientos_raw) and not movimientos_raw[i]["tiene_saldo"]:
                        group.append(movimientos_raw[i])
                        i += 1

                    amounts = [parse_numero_ar(m["importe_str"]) for m in group]

                    # Determinar el saldo objetivo después de este grupo
                    if i < len(movimientos_raw) and movimientos_raw[i]["tiene_saldo"]:
                        next_entry_saldo = movimientos_raw[i]["saldo"]
                        next_entry_amount = parse_numero_ar(movimientos_raw[i]["importe_str"])
                        # El grupo + next_entry van de prev_saldo a next_entry_saldo
                        # next_entry puede ser +/- next_entry_amount
                        # Probar ambas opciones para next_entry
                        target_total = round(next_entry_saldo - prev_saldo, 2)
                        # target_total = group_net + next_signed
                        # Opción A: next es débito → group_net = target_total + next_amount
                        # Opción B: next es crédito → group_net = target_total - next_amount
                        target_a = round(target_total + next_entry_amount, 2)
                        target_b = round(target_total - next_entry_amount, 2)
                    else:
                        # Sin next entry, verificar contra saldo_final
                        target_a = round(ultimo_saldo_final - prev_saldo, 2)
                        target_b = None

                    # Buscar combinación de signos con subset-sum
                    signs = find_signs(amounts, target_a)
                    if signs is None and target_b is not None:
                        signs = find_signs(amounts, target_b)

                    if signs is not None:
                        for j, m in enumerate(group):
                            imp = round(signs[j] * amounts[j], 2)
                            transactions.append({
                                "Fecha": m["fecha"],
                                "Descripcion": clean_for_excel(m["descripcion"]),
                                "Importe": imp
                            })
                            prev_saldo += imp
                    else:
                        # Fallback: asumir todo débito (caso más común)
                        for j, m in enumerate(group):
                            imp = round(-amounts[j], 2)
                            transactions.append({
                                "Fecha": m["fecha"],
                                "Descripcion": clean_for_excel(m["descripcion"]),
                                "Importe": imp
                            })
                            prev_saldo += imp

            if not transactions:
                periodo = "Sin movimientos"
                generar_hoja(wb, cuenta_id, titular, periodo, primer_saldo_inicial, ultimo_saldo_final, [])
                continue

            total_movimientos += len(transactions)

            # Deducir periodo
            fechas = sorted(set(t["Fecha"] for t in transactions),
                          key=lambda x: (int(x[6:10]), int(x[3:5]), int(x[0:2])))
            periodo = f"Del {fechas[0]} al {fechas[-1]}" if fechas else "Sin Especificar"

            generar_hoja(wb, cuenta_id, titular, periodo, primer_saldo_inicial, ultimo_saldo_final, transactions)

        st.success(f"Se procesaron {len(cuentas_agrupadas)} cuenta(s) con {total_movimientos} movimientos totales.")

        wb.save(output)
        output.seek(0)
        return output.getvalue()

    except Exception as e:
        import traceback
        st.error(f"Error al procesar: {e}")
        print(traceback.format_exc())
        return None
