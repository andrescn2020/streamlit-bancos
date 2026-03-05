import streamlit as st
import pdfplumber
import re
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule

# Regex para limpiar caracteres ilegales de Excel
ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

def clean_for_excel(text):
    """Elimina caracteres ilegales para Excel y espacios extra"""
    if not text: return ""
    text = str(text)
    text = ILLEGAL_CHARACTERS_RE.sub("", text)
    return text.strip()

def parse_numero_ar(s):
    """Parsea número en formato argentino (1.234,56) a float. Soporta sufijo - para negativos."""
    if not s: return 0.0
    s = s.strip()
    neg = False
    if s.endswith("-"):
        neg = True
        s = s[:-1]
    elif s.startswith("-"):
        neg = True
        s = s[1:]
    s = s.replace(".", "").replace(",", ".")
    try:
        val = float(s)
        return -val if neg else val
    except:
        return 0.0

# Mapeo de meses abreviados en español a número
MESES = {
    "ENE": "01", "FEB": "02", "MAR": "03", "ABR": "04",
    "MAY": "05", "JUN": "06", "JUL": "07", "AGO": "08",
    "SEP": "09", "OCT": "10", "NOV": "11", "DIC": "12"
}

def convertir_fecha(fecha_str):
    """Convierte '11-JUN-2025' a '11/06/2025'"""
    try:
        partes = fecha_str.split("-")
        dia = partes[0]
        mes = MESES.get(partes[1].upper(), partes[1])
        anio = partes[2]
        return f"{dia}/{mes}/{anio}"
    except:
        return fecha_str

def procesar_ciudad(archivo_pdf):
    st.info("Procesando archivo del Banco Ciudad...")
    try:
        archivo_pdf.seek(0)


        archivo_pdf.seek(0)
        with pdfplumber.open(io.BytesIO(archivo_pdf.read())) as pdf:
            texto_completo = ""
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    texto_completo += t + "\n"

        lineas = texto_completo.splitlines()

        # ============================================================
        # 1. METADATOS
        # ============================================================
        titular = "Sin Especificar"
        periodo = "Sin Especificar"
        cuenta = "Sin Especificar"
        saldo_inicial = 0.0
        saldo_final = 0.0

        # Titular: línea después de "CUIL/CUIT/CDI" que tiene nombre + CUIT
        for i, line in enumerate(lineas):
            if "CUIL/CUIT/CDI" in line:
                if i + 1 < len(lineas):
                    # "LW LOGISTICA Y CARGAS SA 30-71003586-1"
                    next_line = lineas[i + 1].strip()
                    # Sacar el CUIT del final
                    m = re.match(r'(.+?)\s+(\d{2}-\d{8}-\d)', next_line)
                    if m:
                        titular = m.group(1).strip()
                    else:
                        titular = next_line
                break

        # Cuenta: buscar en texto con espacios "C U E N T A N Ú M ER O 5029821/7"
        m_cuenta = re.search(r'C\s*U\s*E\s*N\s*T\s*A\s*N.*?(\d[\d/]+)', texto_completo)
        if m_cuenta:
            cuenta = m_cuenta.group(1).strip()

        # Saldo Anterior: en la línea con "S A L D O  A N T E R I O R" seguido del monto
        m_sa = re.search(r'S\s*A\s*L\s*D\s*O\s*A\s*N\s*T\s*E\s*R\s*I?\s*O?\s*R\s+([\d.,]+)-?', texto_completo)
        if m_sa:
            # Buscar el monto completo incluyendo posible signo negativo al final
            m_sa_full = re.search(r'S\s*A\s*L\s*D\s*O\s*A\s*N\s*T\s*E\s*R\s*I?\s*O?\s*R\s+([\d.,]+-?)', texto_completo)
            if m_sa_full:
                saldo_inicial = parse_numero_ar(m_sa_full.group(1))

        # Saldo Final: "SALDO AL dd/mm/yyyy monto" o "SALDO AL dd/mm/yyyy monto-"
        m_sf = re.search(r'SALDO\s+AL\s+\d{2}/\d{2}/\d{4}\s+([\d.,]+-?)', texto_completo)
        if m_sf:
            saldo_final = parse_numero_ar(m_sf.group(1))

        # Periodo: buscar si está explícito en el PDF
        m_periodo = re.search(r'(?:Per[ií]odo|Desde)\s*:?\s*(\d{2}/\d{2}/\d{4})\s*(?:al|hasta|a)\s*(\d{2}/\d{2}/\d{4})', texto_completo, re.IGNORECASE)
        if m_periodo:
            periodo = f"Del {m_periodo.group(1)} al {m_periodo.group(2)}"

        # ============================================================
        # 2. MOVIMIENTOS - Estrategia: calcular importe desde diferencia de saldos
        # ============================================================
        transactions = []

        # Patrón: dd-MMM-yyyy seguido de descripción y montos
        # Los montos están en formato argentino: 38.059,16 o 8.324,31-
        patron_fecha = re.compile(r'^(\d{2}-[A-Z]{3}-\d{4})\s+(.+)')
        patron_monto = re.compile(r'([\d.]+,\d{2}-?)')

        movimientos_raw = []

        for line in lineas:
            line = line.strip()
            if not line:
                continue

            match = patron_fecha.match(line)
            if match:
                fecha_raw = match.group(1)
                resto = match.group(2)

                # Ignorar líneas de saldo diario o encabezados
                if any(skip in resto.upper() for skip in ["SALDO FINAL DEL", "SALDO INICIAL", "SALDO AL"]):
                    continue

                # Extraer todos los montos de la línea
                montos = patron_monto.findall(resto)

                if montos:
                    # El último monto es siempre el SALDO
                    saldo_str = montos[-1]
                    saldo = parse_numero_ar(saldo_str)

                    # CONCEPTO: todo antes del primer monto
                    pos_primer_monto = resto.find(montos[0])
                    concepto = resto[:pos_primer_monto].strip().rstrip(":")

                    # DESCRIPCIÓN DE MOVIMIENTO: todo después del último monto
                    pos_ultimo_monto = resto.rfind(montos[-1])
                    texto_despues = resto[pos_ultimo_monto + len(montos[-1]):].strip()

                    # Concatenar concepto + descripción
                    if texto_despues:
                        descripcion = f"{concepto} - {texto_despues}" if concepto else texto_despues
                    else:
                        descripcion = concepto

                    movimientos_raw.append({
                        "fecha_raw": fecha_raw,
                        "fecha": convertir_fecha(fecha_raw),
                        "descripcion": descripcion,
                        "saldo": saldo,
                        "montos": montos
                    })

        # Calcular importe como diferencia de saldos
        saldo_previo = saldo_inicial
        for mov in movimientos_raw:
            importe = mov["saldo"] - saldo_previo
            saldo_previo = mov["saldo"]

            transactions.append({
                "Fecha": mov["fecha"],
                "Descripcion": clean_for_excel(mov["descripcion"]),
                "Importe": round(importe, 2)
            })

        if not transactions:
            st.info("No se encontraron movimientos. Se generará el Excel solo con los saldos.")

        if transactions:
            st.success(f"Se encontraron {len(transactions)} movimientos.")

        # ============================================================
        # 3. GENERAR EXCEL
        # ============================================================
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Ciudad"
        ws.sheet_view.showGridLines = False

        # Paleta Banco Ciudad (azul oscuro)
        color_bg_main = "003366"
        color_txt_main = "FFFFFF"

        thin_border = Border(
            left=Side(style='thin', color="A6A6A6"),
            right=Side(style='thin', color="A6A6A6"),
            top=Side(style='thin', color="A6A6A6"),
            bottom=Side(style='thin', color="A6A6A6")
        )

        fill_head_deb = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        fill_col_deb = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
        fill_row_deb = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")

        fill_head_cred = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        fill_col_cred = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")
        fill_row_cred = PatternFill(start_color="F2F9F1", end_color="F2F9F1", fill_type="solid")

        df = pd.DataFrame(transactions) if transactions else pd.DataFrame(columns=["Fecha", "Descripcion", "Importe"])
        creditos = df[df["Importe"] > 0].copy() if not df.empty else pd.DataFrame(columns=["Fecha", "Descripcion", "Importe"])
        debitos = df[df["Importe"] < 0].copy() if not df.empty else pd.DataFrame(columns=["Fecha", "Descripcion", "Importe"])
        if not debitos.empty:
            debitos["Importe"] = debitos["Importe"].abs()

        # --- Header ---
        ws.merge_cells("A1:G1")
        tit = ws["A1"]
        tit.value = f"REPORTE BANCO CIUDAD - CTA {clean_for_excel(cuenta)}"
        tit.font = Font(size=14, bold=True, color=color_txt_main)
        tit.fill = PatternFill(start_color=color_bg_main, end_color=color_bg_main, fill_type="solid")
        tit.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25

        # --- Metadata ---
        ws["A3"] = "SALDO INICIAL"
        ws["A3"].font = Font(bold=True, size=10, color="666666")
        ws["B3"] = saldo_inicial
        ws["B3"].number_format = '"$ "#,##0.00'
        ws["B3"].font = Font(bold=True, size=11)
        ws["B3"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

        ws["A4"] = "SALDO FINAL"
        ws["A4"].font = Font(bold=True, size=10, color="666666")
        ws["B4"] = saldo_final
        ws["B4"].number_format = '"$ "#,##0.00'
        ws["B4"].font = Font(bold=True, size=11)
        ws["B4"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

        ws["D3"] = "TITULAR"
        ws["D3"].alignment = Alignment(horizontal='right')
        ws["D3"].font = Font(bold=True, color="666666", size=10)
        ws["E3"] = clean_for_excel(titular)
        ws["E3"].font = Font(bold=True, size=11)
        ws["E3"].alignment = Alignment(horizontal='center')
        ws.merge_cells("E3:G3")
        for c in ["E","F","G"]: ws[f"{c}3"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

        ws["D4"] = "PERÍODO"
        ws["D4"].alignment = Alignment(horizontal='right')
        ws["D4"].font = Font(bold=True, color="666666", size=10)
        ws["E4"] = clean_for_excel(periodo)
        ws["E4"].font = Font(bold=True, size=11)
        ws["E4"].alignment = Alignment(horizontal='center')
        ws.merge_cells("E4:G4")
        for c in ["E","F","G"]: ws[f"{c}4"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

        # --- Tablas Créditos / Débitos ---
        fila_inicio = 10
        f_header = fila_inicio

        ws.merge_cells(f"A{f_header}:C{f_header}")
        ws[f"A{f_header}"] = "CRÉDITOS"
        ws[f"A{f_header}"].fill = fill_head_cred
        ws[f"A{f_header}"].font = Font(bold=True, color="FFFFFF")
        ws[f"A{f_header}"].alignment = Alignment(horizontal='center')
        ws[f"A{f_header}"].border = thin_border

        ws.merge_cells(f"E{f_header}:G{f_header}")
        ws[f"E{f_header}"] = "DÉBITOS"
        ws[f"E{f_header}"].fill = fill_head_deb
        ws[f"E{f_header}"].font = Font(bold=True, color="FFFFFF")
        ws[f"E{f_header}"].alignment = Alignment(horizontal='center')
        ws[f"E{f_header}"].border = thin_border

        headers = ["Fecha", "Descripción", "Importe"]
        cols_cred = ["A", "B", "C"]
        cols_deb = ["E", "F", "G"]
        f_sub = f_header + 1

        for i, h in enumerate(headers):
            c = ws[f"{cols_cred[i]}{f_sub}"]
            c.value = h
            c.fill = fill_col_cred
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal='center')
            c.border = thin_border

            d = ws[f"{cols_deb[i]}{f_sub}"]
            d.value = h
            d.fill = fill_col_deb
            d.font = Font(bold=True)
            d.alignment = Alignment(horizontal='center')
            d.border = thin_border

        fila_a_llenar = f_sub + 1

        # Créditos
        f_c = fila_a_llenar
        if creditos.empty:
            ws.merge_cells(f"A{f_c}:C{f_c}")
            ws[f"A{f_c}"] = "SIN MOVIMIENTOS"
            ws[f"A{f_c}"].border = thin_border
            f_c += 1
        else:
            start_c = f_c
            for _, r in creditos.iterrows():
                ws[f"A{f_c}"] = r["Fecha"]
                ws[f"A{f_c}"].fill = fill_row_cred
                ws[f"A{f_c}"].border = thin_border
                ws[f"A{f_c}"].alignment = Alignment(horizontal='center')
                ws[f"B{f_c}"] = r["Descripcion"]
                ws[f"B{f_c}"].fill = fill_row_cred
                ws[f"B{f_c}"].border = thin_border
                ws[f"C{f_c}"] = r["Importe"]
                ws[f"C{f_c}"].number_format = '"$ "#,##0.00'
                ws[f"C{f_c}"].fill = fill_row_cred
                ws[f"C{f_c}"].border = thin_border
                f_c += 1
            ws.merge_cells(f"A{f_c}:B{f_c}")
            ws[f"A{f_c}"] = "TOTAL CRÉDITOS"
            ws[f"A{f_c}"].font = Font(bold=True)
            ws[f"A{f_c}"].alignment = Alignment(horizontal='right')
            ws[f"C{f_c}"] = f"=SUM(C{start_c}:C{f_c-1})"
            ws[f"C{f_c}"].font = Font(bold=True)
            ws[f"C{f_c}"].number_format = '"$ "#,##0.00'
            f_c += 1

        # Débitos
        f_d = fila_a_llenar
        if debitos.empty:
            ws.merge_cells(f"E{f_d}:G{f_d}")
            ws[f"E{f_d}"] = "SIN MOVIMIENTOS"
            ws[f"E{f_d}"].border = thin_border
            f_d += 1
        else:
            start_d = f_d
            for _, r in debitos.iterrows():
                ws[f"E{f_d}"] = r["Fecha"]
                ws[f"E{f_d}"].fill = fill_row_deb
                ws[f"E{f_d}"].border = thin_border
                ws[f"E{f_d}"].alignment = Alignment(horizontal='center')
                ws[f"F{f_d}"] = r["Descripcion"]
                ws[f"F{f_d}"].fill = fill_row_deb
                ws[f"F{f_d}"].border = thin_border
                ws[f"G{f_d}"] = r["Importe"]
                ws[f"G{f_d}"].number_format = '"$ "#,##0.00'
                ws[f"G{f_d}"].fill = fill_row_deb
                ws[f"G{f_d}"].border = thin_border
                f_d += 1
            ws.merge_cells(f"E{f_d}:F{f_d}")
            ws[f"E{f_d}"] = "TOTAL DÉBITOS"
            ws[f"E{f_d}"].font = Font(bold=True)
            ws[f"E{f_d}"].alignment = Alignment(horizontal='right')
            ws[f"G{f_d}"] = f"=SUM(G{start_d}:G{f_d-1})"
            ws[f"G{f_d}"].font = Font(bold=True)
            ws[f"G{f_d}"].number_format = '"$ "#,##0.00'
            f_d += 1

        # --- Control de Saldos ---
        ws["D6"] = "CONTROL DE SALDOS"
        ws["D6"].font = Font(bold=True, size=10, color="666666")
        ws["D6"].alignment = Alignment(horizontal='center')

        ref_tot_c = f"C{f_c-1}" if not creditos.empty else "0"
        ref_tot_d = f"G{f_d-1}" if not debitos.empty else "0"
        ws["D7"] = f"=ROUND(B3+{ref_tot_c}-{ref_tot_d}-B4, 2)"
        ws["D7"].number_format = '"$ "#,##0.00'
        ws["D7"].font = Font(bold=True)
        ws["D7"].alignment = Alignment(horizontal='center')
        ws["D7"].border = thin_border

        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        red_font = Font(color='9C0006', bold=True)
        ws.conditional_formatting.add('D7', CellIsRule(operator='notEqual', formula=['0'], stopIfTrue=True, fill=red_fill, font=red_font))

        # --- Anchos de columnas ---
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 25
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 40
        ws.column_dimensions["G"].width = 18

        wb.save(output)
        output.seek(0)
        return output.getvalue()

    except Exception as e:
        import traceback
        st.error(f"Error al procesar: {e}")
        print(traceback.format_exc())
        return None
