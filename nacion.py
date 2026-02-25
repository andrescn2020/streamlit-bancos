import streamlit as st
import pdfplumber
import re
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule

# Regex para caracteres ilegales en Excel
ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

def clean_for_excel(text):
    """Elimina caracteres ilegales para Excel y espacios extra"""
    if not text: return ""
    text = str(text)
    text = ILLEGAL_CHARACTERS_RE.sub("", text)
    return text.strip()

def procesar_nacion(archivo_pdf):
    """Procesa archivos PDF del banco Nación con Estilo Dashboard"""
    st.info("Procesando archivo del banco Nación...")

    try:
        # Reinicializar el archivo para lectura
        archivo_pdf.seek(0)

        # Expresión regular para buscar una fecha en formato dd/mm/yyyy
        patron_fecha = r"\d{2}/\d{2}/\d{4}"

        with pdfplumber.open(io.BytesIO(archivo_pdf.read())) as pdf:
            texto_completo = ""
            for page in pdf.pages:
                texto_completo += page.extract_text() + "\n"
            
            lineas = texto_completo.splitlines()


        # 1. Metadatos (Titular, Periodo)
        titular_global = "Sin Especificar"
        periodo_global = "Sin Especificar"
        
        # Titular: "MENDEZ CLAUDIO OSCAR CUIT: ..."
        for l in lineas[:20]:
            match_tit = re.search(r"^(.*?)\s+CUIT:", l)
            if match_tit:
                titular_global = match_tit.group(1).strip()
                break
        
        # Periodo: "PERIODO: 29/12/2023 AL 31/01/2024"
        match_per = re.search(r"PERIODO:\s+(\d{2}/\d{2}/\d{4})\s+AL\s+(\d{2}/\d{2}/\d{4})", texto_completo, re.IGNORECASE)
        if match_per:
            periodo_global = f"Del {match_per.group(1)} al {match_per.group(2)}"

        inicio = next(
            (i for i, line in enumerate(lineas) if "SALDO ANTERIOR" in line), None
        )
        fin = next((i for i, line in enumerate(lineas) if "SALDO FINAL" in line), None)

        if inicio is None or fin is None:
            st.error(
                "No se encontraron las secciones 'SALDO ANTERIOR' o 'SALDO FINAL' en el PDF"
            )
            return None

        movimientos_extraidos = lineas[inicio - 1 : fin + 1]

        transactions = []
        previous_balance = None
        saldo_inicial = 0.0
        saldo_final = 0.0

        for i, line in enumerate(movimientos_extraidos):
            # Limpieza: eliminar caracteres basura al inicio (e.g. "____ 03/01/25")
            line = line.strip()
            line = re.sub(r"^[_.]+", "", line).strip()

            parts = line.split()
            if "SALDO ANTERIOR" in line:
                try:
                    # parts[-1] es el saldo.
                    val_str = parts[-1].replace(".", "").replace(",", ".")
                    # Handle trailing sign if present (Format: 1.000,00-)
                    if val_str.endswith("-"):
                        val_str = "-" + val_str[:-1]
                    
                    previous_balance = float(val_str)
                    saldo_inicial = previous_balance
                except ValueError:
                    st.warning(f"Error procesando la línea de saldo anterior: {line}")
                continue
            if "SALDO FINAL" in line:
                match = re.search(r"(\d{1,3}(?:\.\d{3})*,\d{2}-?)", line)
                if match:
                    saldo_str = match.group(0)
                    sign = 1
                    if saldo_str.endswith("-"):
                        sign = -1
                        saldo_str = saldo_str[:-1]
                    saldo_final = float(saldo_str.replace(".", "").replace(",", ".")) * sign

            if "FECHA MOVIMIENTOS" in line:
                continue

            if len(parts) < 3:
                # Líneas vacías o cortes de página
                continue

            date = parts[0]
            # Validación simple de fecha
            if not re.match(r"\d{2}/\d{2}/\d{2,4}", date):
                continue

            # Amount is usually parts[-2], Balance parts[-1]. Description in middle.
            # But we need to be careful about parts length.
            # Logic preserved from original script: description = parts[1:-3]
            
            description = " ".join(parts[1:-3])
            amount_str = parts[-2]
            balance_str = parts[-1]

            try:
                # Helper for sign handling
                def parse_amount(s):
                    sign = 1
                    # Sufijo "A" = Anulación (Banco Nación)
                    if s.upper().endswith("A"):
                        s = s[:-1]
                    if s.endswith("-"):
                        sign = -1
                        s = s[:-1]
                    elif s.startswith("-"):
                        sign = -1
                        s = s[1:]
                    return float(s.replace(".", "").replace(",", ".")) * sign

                amount = parse_amount(amount_str)
                balance = parse_amount(balance_str)
            except ValueError:
                # st.warning(f"No se pudo procesar montos en línea: {line}")
                continue

            # Invertir lógica de Débito/Crédito: 
            # Original script logic:
            # if balance < previous_balance: Debit
            # else: Credit
            # BUT wait, amount sign logic in original script was:
            # amount = float(...) * (-1 if "-" in amount_str else 1)
            # If amount_str has "-", it's negative.
            # Usually in bank statements, negative amount = debit.
            
            # Let's trust the sign of 'amount'.
            # If amount > 0 => Credit. If amount < 0 => Debit.
            
            # Original logic re-check:
            # if balance < previous_balance: transaction_type = "Débito", amount = -abs(amount)
            # This implies amount sign might be ambiguous in text, so it relies on balance diff.
            # Let's keep that robustness.
            
            if previous_balance is not None:
                # Recalculamos el signo basado en la variación de saldo para máxima seguridad
                # Diferencia teórica
                diff = balance - previous_balance
                # Si diff es negativa, es un débito.
                # Ajustamos el signo de 'amount' para que coincida.
                if diff < -0.01: # Tolerancia float
                     if amount > 0: amount = -amount
                elif diff > 0.01:
                     if amount < 0: amount = -amount
            
            transactions.append({
                "Fecha": date,
                "Descripcion": clean_for_excel(description),
                "Importe": amount
            })

            previous_balance = balance

        if not transactions:
            st.warning("No se encontraron movimientos en el PDF")
            return None

        # --- GENERACIÓN EXCEL (DASHBOARD) ---
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Nacion"
        ws.sheet_view.showGridLines = False
        
        # Paleta Nacion (Celeste/Azul)
        color_bg_main = "0066CC" 
        color_txt_main = "FFFFFF"
        
        thin_border = Border(left=Side(style='thin', color="A6A6A6"), 
                             right=Side(style='thin', color="A6A6A6"), 
                             top=Side(style='thin', color="A6A6A6"), 
                             bottom=Side(style='thin', color="A6A6A6"))
                             
        fill_head_deb = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        fill_col_deb = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
        fill_row_deb = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")

        fill_head_cred = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        fill_col_cred = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")
        fill_row_cred = PatternFill(start_color="F2F9F1", end_color="F2F9F1", fill_type="solid")

        df = pd.DataFrame(transactions)
        if not df.empty:
            creditos = df[df["Importe"] > 0].copy()
            debitos = df[df["Importe"] < 0].copy()
            debitos["Importe"] = debitos["Importe"].abs()
        else:
             creditos = pd.DataFrame(columns=["Fecha", "Descripcion", "Importe"])
             debitos = pd.DataFrame(columns=["Fecha", "Descripcion", "Importe"])

        # 1. Header
        ws.merge_cells("A1:G1")
        tit = ws["A1"]
        tit.value = f"REPORTE NACIÓN - {clean_for_excel(titular_global)}"
        tit.font = Font(size=14, bold=True, color=color_txt_main)
        tit.fill = PatternFill(start_color=color_bg_main, end_color=color_bg_main, fill_type="solid")
        tit.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25

        # 2. Metadata y Saldos
        ws["A3"] = "SALDO INICIAL"
        ws["A3"].font = Font(bold=True, size=10, color="666666")
        ws["B3"] = saldo_inicial
        ws["B3"].number_format = '"$ "#,##0.00'
        ws["B3"].font = Font(bold=True, size=11)
        ws["B3"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

        ws["A4"] = "SALDO FINAL"
        ws["A4"].font = Font(bold=True, size=10, color="666666")
        ws["B4"] = saldo_final
        ws["B4"].number_format = '"$ "#,##0.00'
        ws["B4"].font = Font(bold=True, size=11)
        ws["B4"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

        ws["D3"] = "TITULAR"
        ws["D3"].alignment = Alignment(horizontal='right')
        ws["D3"].font = Font(bold=True, color="666666", size=10)
        ws["E3"] = clean_for_excel(titular_global)
        ws["E3"].font = Font(bold=True, size=11)
        ws["E3"].alignment = Alignment(horizontal='center')
        ws.merge_cells("E3:G3")
        for c in ["E","F","G"]: ws[f"{c}3"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

        ws["D4"] = "PERÍODO"
        ws["D4"].alignment = Alignment(horizontal='right')
        ws["D4"].font = Font(bold=True, color="666666", size=10)
        ws["E4"] = clean_for_excel(periodo_global)
        ws["E4"].font = Font(bold=True, size=11)
        ws["E4"].alignment = Alignment(horizontal='center')
        ws.merge_cells("E4:G4")
        for c in ["E","F","G"]: ws[f"{c}4"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

        ws["D6"] = "CONTROL DE SALDOS"
        ws["D6"].font = Font(bold=True, size=10, color="666666")
        ws["D6"].alignment = Alignment(horizontal='center')
        
        cell_ctl = ws["D7"]
        cell_ctl.font = Font(bold=True, size=12)
        cell_ctl.alignment = Alignment(horizontal='center')
        cell_ctl.border = thin_border

        # 3. Tablas Paralelas
        fila_inicio = 10
        
        # Headers
        f_header = fila_inicio
        ws.merge_cells(f"A{f_header}:C{f_header}")
        ws[f"A{f_header}"] = "CRÉDITOS" 
        ws[f"A{f_header}"].fill = fill_head_cred
        ws[f"A{f_header}"].font = Font(bold=True, color="FFFFFF")
        ws[f"A{f_header}"].alignment = Alignment(horizontal='center')
        ws[f"A{f_header}"].border = thin_border
        
        headers = ["Fecha", "Descripción", "Importe"]
        cols_cred = ["A", "B", "C"]
        f_sub = f_header + 1
        for i, h in enumerate(headers):
            c = ws[f"{cols_cred[i]}{f_sub}"]
            c.value = h
            c.fill = fill_col_cred
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal='center')
            c.border = thin_border

        ws.merge_cells(f"E{f_header}:G{f_header}")
        ws[f"E{f_header}"] = "DÉBITOS" 
        ws[f"E{f_header}"].fill = fill_head_deb
        ws[f"E{f_header}"].font = Font(bold=True, color="FFFFFF")
        ws[f"E{f_header}"].alignment = Alignment(horizontal='center')
        ws[f"E{f_header}"].border = thin_border
        
        cols_deb = ["E", "F", "G"]
        for i, h in enumerate(headers):
            c = ws[f"{cols_deb[i]}{f_sub}"]
            c.value = h
            c.fill = fill_col_deb
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal='center')
            c.border = thin_border

        # Datos
        fila_dato_start = f_sub + 1
        
        # Créditos
        f_cred = fila_dato_start
        if creditos.empty:
            ws.merge_cells(f"A{f_cred}:C{f_cred}")
            ws[f"A{f_cred}"] = "SIN MOVIMIENTOS"
            ws[f"A{f_cred}"].font = Font(italic=True, color="666666")
            ws[f"A{f_cred}"].alignment = Alignment(horizontal='center')
            ws[f"A{f_cred}"].border = thin_border
            f_cred += 1
        else:
            start_c = f_cred
            for _, r in creditos.iterrows():
                ws[f"A{f_cred}"] = clean_for_excel(r["Fecha"])
                ws[f"A{f_cred}"].fill = fill_row_cred
                ws[f"A{f_cred}"].alignment = Alignment(horizontal='center')
                ws[f"A{f_cred}"].border = thin_border
                ws[f"B{f_cred}"] = clean_for_excel(r["Descripcion"])
                ws[f"B{f_cred}"].fill = fill_row_cred
                ws[f"B{f_cred}"].border = thin_border
                ws[f"C{f_cred}"] = r["Importe"]
                ws[f"C{f_cred}"].number_format = '"$ "#,##0.00'
                ws[f"C{f_cred}"].fill = fill_row_cred
                ws[f"C{f_cred}"].border = thin_border
                f_cred += 1
            ws.merge_cells(f"A{f_cred}:B{f_cred}")
            ws[f"A{f_cred}"] = "TOTAL CRÉDITOS"
            ws[f"A{f_cred}"].font = Font(bold=True)
            ws[f"A{f_cred}"].alignment = Alignment(horizontal='right')
            ws[f"A{f_cred}"].border = thin_border
            ws[f"C{f_cred}"] = f"=SUM(C{start_c}:C{f_cred-1})"
            ws[f"C{f_cred}"].number_format = '"$ "#,##0.00'
            ws[f"C{f_cred}"].font = Font(bold=True)
            ws[f"C{f_cred}"].border = thin_border
            f_cred += 1

        # Débitos
        f_deb = fila_dato_start
        if debitos.empty:
            ws.merge_cells(f"E{f_deb}:G{f_deb}")
            ws[f"E{f_deb}"] = "SIN MOVIMIENTOS"
            ws[f"E{f_deb}"].font = Font(italic=True, color="666666")
            ws[f"E{f_deb}"].alignment = Alignment(horizontal='center')
            ws[f"E{f_deb}"].border = thin_border
            f_deb += 1
        else:
            start_d = f_deb
            for _, r in debitos.iterrows():
                ws[f"E{f_deb}"] = clean_for_excel(r["Fecha"])
                ws[f"E{f_deb}"].fill = fill_row_deb
                ws[f"E{f_deb}"].alignment = Alignment(horizontal='center')
                ws[f"E{f_deb}"].border = thin_border
                ws[f"F{f_deb}"] = clean_for_excel(r["Descripcion"])
                ws[f"F{f_deb}"].fill = fill_row_deb
                ws[f"F{f_deb}"].border = thin_border
                ws[f"G{f_deb}"] = r["Importe"]
                ws[f"G{f_deb}"].number_format = '"$ "#,##0.00'
                ws[f"G{f_deb}"].fill = fill_row_deb
                ws[f"G{f_deb}"].border = thin_border
                f_deb += 1
            ws.merge_cells(f"E{f_deb}:F{f_deb}")
            ws[f"E{f_deb}"] = "TOTAL DÉBITOS"
            ws[f"E{f_deb}"].font = Font(bold=True)
            ws[f"E{f_deb}"].alignment = Alignment(horizontal='right')
            ws[f"E{f_deb}"].border = thin_border
            ws[f"G{f_deb}"] = f"=SUM(G{start_d}:G{f_deb-1})"
            ws[f"G{f_deb}"].number_format = '"$ "#,##0.00'
            ws[f"G{f_deb}"].font = Font(bold=True)
            ws[f"G{f_deb}"].border = thin_border
            f_deb += 1

        # Formula
        f_ini = "B3"
        f_tot_cred = f"C{f_cred-1}" if not creditos.empty else "0"
        f_tot_deb = f"G{f_deb-1}" if not debitos.empty else "0"
        f_fin = "B4"
        ws["D7"] = f"=ROUND({f_ini}+{f_tot_cred}-{f_tot_deb}-{f_fin}, 2)"
        ws["D7"].number_format = '"$ "#,##0.00'
        
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        red_font = Font(color='9C0006', bold=True)
        ws.conditional_formatting.add('D7', CellIsRule(operator='notEqual', formula=['0'], stopIfTrue=True, fill=red_fill, font=red_font))

        # Anchos
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 25
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 40
        ws.column_dimensions["G"].width = 18

        wb.save(output)
        output.seek(0)
        return output.getvalue()

    except Exception as e:
        import traceback
        st.error(f"Error al procesar el archivo: {str(e)}")
        print(traceback.format_exc())
        return None
