import streamlit as st
import pdfplumber
import pandas as pd
import re
import io
import PyPDF2
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def clean_for_excel(text):
    """Elimina caracteres ilegales para Excel."""
    if not text: return ""
    return re.sub(r'[\000-\010]|[\013-\014]|[\016-\037]', '', str(text)).strip()

def procesar_galicia_mas(archivo_pdf, debug=False):
    """
    Procesador Galicia MÃ¡s V1.0 - Motor Layout + Regex
    -------------------------------------------------------
    Clon del motor HSBC V8.1.
    Detecta movimientos por el patron " - " (guion separador).
    Soporta movimientos sin fecha (heredan fecha anterior).
    Transacciones se detectan ANTES del filtro de basura.
    """
    st.info("Procesando Galicia MÃ¡s V1.0 (Motor Layout + Regex)...")
    
    try:
        archivo_pdf.seek(0)
        
        # Estructuras de datos
        saldos_iniciales = {}
        saldos_finales = {}
        cuentas_data = {}
        info_cuentas = {}
        
        titular_str = "S/D"
        year = "2024"
        periodo_str = ""
        
        current_account = None
        last_date = None
        
        # Regex
        re_cuenta = re.compile(r"(\d{3,4}-?\d?-?\d{5}-\d)")
        re_trx = re.compile(r"^\s*(?:(\d{2}-[A-Z]{3})\s+)?-\s+(.+)")
        # Montos: soporta 100,000.00 | 583.05 | .03 (centavos)
        # Word boundaries para no matchear parciales como 25.41 de 25.413
        re_monto = re.compile(r"(?<!\d)(\d{1,3}(?:,\d{3})*\.\d{2}|\.\d{2})(?!\d)")
        
        debug_lines_raw = []      # texto crudo por pagina
        debug_movs = []            # movimientos parseados
        debug_discarded = []       # lineas descartadas por filtro basura
        
        with pdfplumber.open(io.BytesIO(archivo_pdf.read())) as pdf:
            # ============================================================
            # FASE 1: METADATA GLOBAL (pagina 1, sin layout)
            # ============================================================
            text_p1 = pdf.pages[0].extract_text() or ""
            
            # 1.1 AÃ‘O y PERIODO
            match_periodo = re.search(r"EXTRACTO DEL (\d{2}/\d{2}/\d{4}) AL (\d{2}/\d{2}/\d{4})", text_p1)
            if match_periodo:
                periodo_str = f"{match_periodo.group(1)} al {match_periodo.group(2)}"
                year = match_periodo.group(2)[-4:]
            else:
                match_anio = re.search(r"EXTRACTO DEL \d{2}/\d{2}/(\d{4})", text_p1)
                if match_anio: year = match_anio.group(1)
            
            last_date = f"01-ENE-{year}"
            
            # 1.2 TITULAR
            for l in text_p1.splitlines():
                ls = l.strip().upper()
                if "ESTIMADO" in ls:
                    # Intentar formato persona: "ESTIMADO APELLIDO, NOMBRE"
                    m = re.search(r"([A-Z][A-Z]+,\s+[A-Z][A-Z ]+)\s*$", l.strip())
                    if m:
                        titular_str = m.group(1).strip()
                    else:
                        # Formato empresa: "ESTIMADOS SEÃ‘ORES EMPRESA S.A."
                        m2 = re.search(r"ESTIMADOS?\s+(?:SE.ORES?\s+)?(.+)$", l.strip(), re.IGNORECASE)
                        if m2:
                            titular_str = m2.group(1).strip()
                    break
            
            # 1.3 TABLA RESUMEN DE CUENTAS
            lines_p1 = text_p1.splitlines()
            in_summary = False
            for l in lines_p1:
                ls = l.strip()
                if "PRODUCTO" in ls and "SALDO" in ls:
                    in_summary = True
                    continue
                
                if in_summary:
                    if "NRO." in ls and "EN" in ls:
                        in_summary = False
                        continue
                    if not ls or "DETALLE" in ls:
                        in_summary = False
                        continue
                    
                    match_cta = re_cuenta.search(l)
                    if match_cta:
                        cta_num = match_cta.group(1)
                        pre = l[:l.find(cta_num)].strip()
                        parts = pre.split()
                        if parts and len(parts[-1]) >= 3 and parts[-1].isalpha():
                            prod_name = " ".join(parts[:-1])
                        else:
                            prod_name = pre
                        prod_name = prod_name.strip()
                        
                        info_cuentas[cta_num] = prod_name
                        
                        montos = re_monto.findall(l)
                        if len(montos) >= 2:
                            try:
                                saldos_iniciales[cta_num] = float(montos[-2].replace(",",""))
                                saldos_finales[cta_num] = float(montos[-1].replace(",",""))
                            except: pass
                        elif len(montos) == 1:
                            try:
                                saldos_iniciales[cta_num] = 0.0
                                saldos_finales[cta_num] = float(montos[0].replace(",",""))
                            except: pass


            for cta in info_cuentas:
                cuentas_data[cta] = []
            
            # ============================================================
            # FASE 2: PARSEO DE MOVIMIENTOS (layout=True)
            # ============================================================
            # IMPORTANTE: Transacciones se detectan ANTES del filtro de basura.
            # El filtro de basura SOLO aplica a lineas de continuacion.
            # ============================================================
            current_account = None
            
            # Keywords de basura SOLO para filtrar lineas de CONTINUACION
            BASURA_CONT = [
                "SALDO ANTERIOR", "SALDO FINAL", "HOJA ", "PAGINA",
                "DETALLE DE OPERACIONES", "TITULARIDAD", "CUIT", "INGRESOS BRUTOS",
                "COMUNICACION", "B.C.R.A", "GALICIA",
                "SEGURIDAD", "EXTRACTO NRO", "EXTRACTO DEL",
                "PRODUCTO", "ESTIMADO", "IMPORTANTE",
                "FECHA      REFERENCIA", "NO HUBO NINGUNA", "CALCULO DE",
                "MOVIMIENTOS INFORMADOS", "DEPOSITOS DE AHORRO",
                "REGIMEN DE TRANSPARENCIA", "PUEDE SOLICITAR", "PUEDE CONSULTAR",
                "CREDITO FISCAL", "PC BANKING", "TODAVIA NO", "WWW.GALICIA",
                "REGISTRACION DE CHEQUE", "CUSTODIA DE", "CHEQUE RECHAZADO",
                "SI DESEA MAYOR", "CENTRO DE ATENCION", "COMISION POR",
                "EL MONTO DEL IVA"
            ]
            
            for page_idx, page in enumerate(pdf.pages):
                text = page.extract_text(layout=True)
                if not text: continue
                
                if debug:
                    debug_lines_raw.append((page_idx + 1, text))
                
                for line in text.splitlines():
                    line_clean = line.strip()
                    if not line_clean: continue
                    
                    # --- 1. Detectar cambio de cuenta ---
                    if "NRO." in line_clean.upper() and re_cuenta.search(line_clean):
                        found = re_cuenta.search(line_clean).group(1)
                        if found in info_cuentas:
                            current_account = found
                        continue
                    
                    if not current_account:
                        continue
                    
                    # --- 2. PRIMERO: Detectar transaccion (ANTES de filtro basura) ---
                    match_trx = re_trx.match(line_clean)
                    
                    # Excepciones: SALDO ANTERIOR/FINAL empiezan con "- " pero no son movimientos
                    is_balance = "SALDO ANTERIOR" in line_clean or "SALDO FINAL" in line_clean
                    
                    if match_trx and not is_balance:
                        raw_date = match_trx.group(1)
                        desc_part = match_trx.group(2)
                        
                        if raw_date:
                            full_date = f"{raw_date}-{year}"
                            last_date = full_date
                        else:
                            full_date = last_date
                        
                        montos = re_monto.findall(desc_part)
                        
                        if len(montos) >= 2:
                            try:
                                importe = float(montos[-2].replace(",", ""))
                                saldo = float(montos[-1].replace(",", ""))
                            except:
                                continue
                            
                            # Detectar saldo negativo con guion al final (ej: 11,254.73-)
                            last_monto_str = montos[-1]
                            pos_last = desc_part.rfind(last_monto_str)
                            if pos_last != -1:
                                after_saldo = desc_part[pos_last + len(last_monto_str):].strip()
                                if after_saldo.startswith("-"):
                                    saldo = -saldo
                            
                            if current_account not in cuentas_data:
                                cuentas_data[current_account] = []
                            
                            if cuentas_data[current_account]:
                                saldo_prev = cuentas_data[current_account][-1]["Saldo"]
                            else:
                                saldo_prev = saldos_iniciales.get(current_account, 0.0)
                            
                            diff_deb = abs((saldo_prev - importe) - saldo)
                            diff_cred = abs((saldo_prev + importe) - saldo)
                            
                            # Elegir la MEJOR coincidencia (menor diferencia)
                            if diff_deb <= diff_cred and diff_deb < 1.0:
                                es_credito = False
                            elif diff_cred < diff_deb and diff_cred < 1.0:
                                es_credito = True
                            else:
                                es_credito = any(kw in desc_part.upper() for kw in ["DEP.", "DEPOSITO", "CRED", "CREDI"])
                            
                            desc_clean = desc_part
                            for m in reversed(montos):
                                idx = desc_clean.rfind(m)
                                if idx != -1:
                                    desc_clean = desc_clean[:idx]
                            desc_clean = desc_clean.strip("- ").strip()
                            
                            mov_entry = {
                                "Fecha": full_date,
                                "Descripcion": desc_clean,
                                "Debito": importe if not es_credito else 0.0,
                                "Credito": importe if es_credito else 0.0,
                                "Saldo": saldo
                            }
                            cuentas_data[current_account].append(mov_entry)
                            
                            if debug:
                                saldo_calc = saldo_prev + (importe if es_credito else -importe)
                                diff_ctrl = round(saldo_calc - saldo, 2)
                                debug_movs.append({
                                    "Cuenta": current_account,
                                    "Pag": page_idx + 1,
                                    "Fecha": full_date,
                                    "Desc": desc_clean[:50],
                                    "Tipo": "CRED" if es_credito else "DEB",
                                    "Importe": importe,
                                    "Saldo PDF": saldo,
                                    "Saldo Calc": round(saldo_calc, 2),
                                    "Diff": diff_ctrl,
                                    "Montos raw": montos,
                                    "Linea": line_clean[:80]
                                })
                    else:
                        # --- 3. DESPUES: Linea de continuacion con filtro basura ---
                        if current_account and cuentas_data.get(current_account):
                            line_upper = line_clean.upper()
                            is_junk = any(kw in line_upper for kw in BASURA_CONT)
                            if re.search(r"\d{6,}-[A-Z]", line_clean): is_junk = True
                            if line_clean.startswith("_"): is_junk = True
                            if len(line_clean) > 100: is_junk = True
                            
                            if not is_junk and not is_balance:
                                cuentas_data[current_account][-1]["Descripcion"] += " " + line_clean
                            elif debug and is_junk:
                                debug_discarded.append({
                                    "Cuenta": current_account,
                                    "Pag": page_idx + 1,
                                    "Linea": line_clean[:100]
                                })
        
        # ============================================================
        # DEBUG OUTPUT - todo en un solo bloque copiable
        # ============================================================
        if debug:
            debug_output = []
            debug_output.append("=" * 60)
            debug_output.append("GALICIA MAS - DEBUG COMPLETO")
            debug_output.append("=" * 60)
            
            # Control por cuenta
            debug_output.append("\n--- CONTROL POR CUENTA ---")
            for cta in info_cuentas:
                movs = cuentas_data.get(cta, [])
                s_ini = saldos_iniciales.get(cta, 0.0)
                s_fin = saldos_finales.get(cta, 0.0)
                total_cred = sum(m["Credito"] for m in movs)
                total_deb = sum(m["Debito"] for m in movs)
                control = round(s_ini + total_cred - total_deb - s_fin, 2)
                marca = "OK" if control == 0 else f"DIFF={control}"
                debug_output.append(f"[{marca}] {info_cuentas[cta]} ({cta}): SI={s_ini:,.2f} + CRED={total_cred:,.2f} - DEB={total_deb:,.2f} = {round(s_ini + total_cred - total_deb, 2):,.2f} vs SF={s_fin:,.2f}")
            
            # Movimientos parseados
            debug_output.append("\n--- MOVIMIENTOS PARSEADOS ---")
            if debug_movs:
                debug_output.append(f"Total: {len(debug_movs)}")
                for i, m in enumerate(debug_movs, 1):
                    flag = " *** DIFF ***" if m["Diff"] != 0 else ""
                    debug_output.append(
                        f"{i:3d}| Cta={m['Cuenta']} Pag={m['Pag']} {m['Fecha']} "
                        f"{m['Tipo']} {m['Importe']:>12,.2f} "
                        f"SaldoPDF={m['Saldo PDF']:>12,.2f} SaldoCalc={m['Saldo Calc']:>12,.2f} "
                        f"Diff={m['Diff']:>8,.2f}{flag} | {m['Desc']}"
                    )
            else:
                debug_output.append("Sin movimientos.")
            
            # Lineas descartadas
            debug_output.append(f"\n--- LINEAS DESCARTADAS ({len(debug_discarded)}) ---")
            for d in debug_discarded:
                debug_output.append(f"  Cta={d['Cuenta']} Pag={d['Pag']} | {d['Linea']}")
            
            # Texto crudo
            debug_output.append("\n--- TEXTO CRUDO PDF ---")
            for pg_num, pg_text in debug_lines_raw:
                debug_output.append(f"\n>> PAGINA {pg_num} <<")
                debug_output.append(pg_text)
            
            debug_output.append("\n" + "=" * 60)
            
            st.markdown("---")
            st.subheader("ðŸ” DEBUG OUTPUT (copiÃ¡ todo)")
            st.code("\n".join(debug_output), language=None)
        
        # ============================================================
        # FASE 3: GENERAR EXCEL
        # ============================================================
        if not any(cuentas_data.values()):
            st.warning("No se extrajeron movimientos de ninguna cuenta.")
            return None

        output = io.BytesIO()
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
            
        # Estilos - Galicia MÃ¡s (naranja corporativo)
        color_galicia_mas = "FF6600"
        fill_header = PatternFill(start_color=color_galicia_mas, end_color=color_galicia_mas, fill_type="solid")
        font_header = Font(color="FFFFFF", bold=True, size=12)
        font_bold = Font(bold=True)
        thin_border = Border(
            left=Side(style='thin', color="A6A6A6"), right=Side(style='thin', color="A6A6A6"),
            top=Side(style='thin', color="A6A6A6"), bottom=Side(style='thin', color="A6A6A6")
        )
        fill_deb_h = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        fill_deb_c = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
        fill_cred_h = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        fill_cred_c = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")

        for cta in info_cuentas:
            movs = cuentas_data.get(cta, [])
            prod = info_cuentas.get(cta, "")
            
            if "u$s" in prod.lower() or "dol" in prod.lower():
                safe_name = f"USD {cta}"
            else:
                safe_name = f"ARS {cta}"
            
            ws = wb.create_sheet(title=clean_for_excel(safe_name)[:30])
            ws.sheet_view.showGridLines = False
            
            ws.column_dimensions['A'].width = 14
            ws.column_dimensions['B'].width = 55
            ws.column_dimensions['C'].width = 16
            ws.column_dimensions['D'].width = 4
            ws.column_dimensions['E'].width = 14
            ws.column_dimensions['F'].width = 55
            ws.column_dimensions['G'].width = 16
            ws.column_dimensions['I'].width = 20
            ws.column_dimensions['J'].width = 25
            
            ws.merge_cells("A1:G1")
            ws["A1"] = f"REPORTE GALICIA MÃS - {prod} - {cta}"
            ws["A1"].fill = fill_header
            ws["A1"].font = font_header
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 28
            
            s_ini = saldos_iniciales.get(cta, 0.0)
            s_fin = saldos_finales.get(cta, 0.0)
            
            ws["A3"] = "SALDO INICIAL"
            ws["A3"].font = Font(bold=True, color="666666")
            ws["B3"] = s_ini
            ws["B3"].number_format = '"$ "#,##0.00'
            ws["B3"].font = font_bold
            
            ws["A4"] = "SALDO FINAL"
            ws["A4"].font = Font(bold=True, color="666666")
            ws["B4"] = s_fin
            ws["B4"].number_format = '"$ "#,##0.00'
            ws["B4"].font = font_bold
            
            ws["I3"] = "TITULAR"
            ws["I3"].font = Font(bold=True, color="666666")
            ws["J3"] = titular_str
            ws["J3"].font = font_bold
            
            ws["I4"] = "PERIODO"
            ws["I4"].font = Font(bold=True, color="666666")
            ws["J4"] = periodo_str if periodo_str else year
            ws["J4"].font = font_bold
            
            if not movs:
                ws.merge_cells("A7:G7")
                ws["A7"] = "NO HUBO MOVIMIENTOS EN ESTE PERIODO"
                ws["A7"].font = Font(italic=True, color="666666", size=11)
                ws["A7"].alignment = Alignment(horizontal="center")
                continue
            
            df = pd.DataFrame(movs)
            creditos = df[df["Credito"] > 0]
            debitos = df[df["Debito"] > 0]
            
            fila = 7
            headers = ["Fecha", "DescripciÃ³n", "Importe"]
            
            # CREDITOS (A-C)
            ws.merge_cells(f"A{fila}:C{fila}")
            ws[f"A{fila}"] = "CRÃ‰DITOS"
            ws[f"A{fila}"].fill = fill_cred_h
            ws[f"A{fila}"].font = Font(color="FFFFFF", bold=True)
            ws[f"A{fila}"].alignment = Alignment(horizontal="center")
            
            for i, h in enumerate(headers):
                c = ws.cell(row=fila+1, column=1+i, value=h)
                c.fill = fill_cred_c; c.font = font_bold
                c.border = thin_border; c.alignment = Alignment(horizontal="center")
            
            r_cred = fila + 2
            start_cred = r_cred
            if creditos.empty:
                ws[f"A{r_cred}"] = "SIN MOVIMIENTOS"
                ws[f"A{r_cred}"].font = Font(italic=True, color="999999")
                r_cred += 1
            else:
                for _, row in creditos.iterrows():
                    ws[f"A{r_cred}"] = row["Fecha"]
                    ws[f"A{r_cred}"].alignment = Alignment(horizontal="center")
                    ws[f"B{r_cred}"] = clean_for_excel(row["Descripcion"])
                    ws[f"C{r_cred}"] = row["Credito"]
                    ws[f"C{r_cred}"].number_format = '"$ "#,##0.00'
                    for col in ["A","B","C"]:
                        ws[f"{col}{r_cred}"].border = thin_border
                        if r_cred % 2 == 0: ws[f"{col}{r_cred}"].fill = fill_cred_c
                    r_cred += 1
            
            ws[f"B{r_cred}"] = "TOTAL CRÃ‰DITOS"
            ws[f"B{r_cred}"].font = font_bold
            ws[f"B{r_cred}"].alignment = Alignment(horizontal="right")
            ref_cred = f"C{r_cred}"
            if not creditos.empty:
                ws[f"C{r_cred}"] = f"=SUM(C{start_cred}:C{r_cred-1})"
            else:
                ws[f"C{r_cred}"] = 0
            ws[f"C{r_cred}"].number_format = '"$ "#,##0.00'
            ws[f"C{r_cred}"].font = font_bold
            
            # DEBITOS (E-G)
            ws.merge_cells(f"E{fila}:G{fila}")
            ws[f"E{fila}"] = "DÃ‰BITOS"
            ws[f"E{fila}"].fill = fill_deb_h
            ws[f"E{fila}"].font = Font(color="FFFFFF", bold=True)
            ws[f"E{fila}"].alignment = Alignment(horizontal="center")
            
            for i, h in enumerate(headers):
                c = ws.cell(row=fila+1, column=5+i, value=h)
                c.fill = fill_deb_c; c.font = font_bold
                c.border = thin_border; c.alignment = Alignment(horizontal="center")
            
            r_deb = fila + 2
            start_deb = r_deb
            if debitos.empty:
                ws[f"E{r_deb}"] = "SIN MOVIMIENTOS"
                ws[f"E{r_deb}"].font = Font(italic=True, color="999999")
                r_deb += 1
            else:
                for _, row in debitos.iterrows():
                    ws[f"E{r_deb}"] = row["Fecha"]
                    ws[f"E{r_deb}"].alignment = Alignment(horizontal="center")
                    ws[f"F{r_deb}"] = clean_for_excel(row["Descripcion"])
                    ws[f"G{r_deb}"] = row["Debito"]
                    ws[f"G{r_deb}"].number_format = '"$ "#,##0.00'
                    for col in ["E","F","G"]:
                        ws[f"{col}{r_deb}"].border = thin_border
                        if r_deb % 2 == 0: ws[f"{col}{r_deb}"].fill = fill_deb_c
                    r_deb += 1
            
            ws[f"F{r_deb}"] = "TOTAL DÃ‰BITOS"
            ws[f"F{r_deb}"].font = font_bold
            ws[f"F{r_deb}"].alignment = Alignment(horizontal="right")
            ref_deb = f"G{r_deb}"
            if not debitos.empty:
                ws[f"G{r_deb}"] = f"=SUM(G{start_deb}:G{r_deb-1})"
            else:
                ws[f"G{r_deb}"] = 0
            ws[f"G{r_deb}"].number_format = '"$ "#,##0.00'
            ws[f"G{r_deb}"].font = font_bold
            
            # CONTROL FORMULA
            ws["I6"] = "CONTROL (debe ser 0)"
            ws["I6"].font = Font(bold=True, color="666666")
            ws["I6"].alignment = Alignment(horizontal="center")
            
            formula = f"=ROUND(B3 + {ref_cred} - {ref_deb} - B4, 2)"
            ws["I7"] = formula
            ws["I7"].font = Font(bold=True, size=14)
            ws["I7"].alignment = Alignment(horizontal="center")
            ws["I7"].number_format = '"$ "#,##0.00'
            ws["I7"].border = thin_border
        
        wb.save(output)
        output.seek(0)
        st.success("âœ… Procesamiento completado (Galicia MÃ¡s V1.0)")
        return output.getvalue()
        
    except Exception as e:
        st.error(f"Error CrÃ­tico: {str(e)}")
        import traceback
        st.code(traceback.format_exc())
        return None
