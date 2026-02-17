import streamlit as st
import re
import pandas as pd
import PyPDF2
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- UTILIDADES DE LIMPIEZA ---
ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
def clean_for_excel(text):
    if text is None: return ""
    return ILLEGAL_CHARACTERS_RE.sub("", str(text)).strip()

def procesar_supervielle(archivo_pdf):
    """Procesa archivos PDF del banco Supervielle (Formato Original)"""
    st.info("Procesando archivo del banco Supervielle (Formato Original)...")

    try:
        # Reinicializar el archivo para lectura
        archivo_pdf.seek(0)
        
        # --- LÓGICA DE EXTRACCIÓN ORIGINAL (Preservada) ---
        def procesar_pdf(file_bytes):
            reader = PyPDF2.PdfReader(io.BytesIO(file_bytes))
            texto = "".join(page.extract_text() + "\n" for page in reader.pages)
            lineas = texto.splitlines()

            capturar = False
            numero_de_cuenta_temporal = ""
            movimientos = []
            cuentas = []
            periodo_global = "S/D"
            titular_global = "S/D"

            for linea in lineas:
                # Extracción de HEADER GLOBAL (Periodo y Titular)
                # L10: 'RESUMEN DE CUENTA DESDE 01/03/23 HASTA 31/03/23'
                if "RESUMEN DE CUENTA DESDE" in linea:
                    periodo_global = linea.replace("RESUMEN DE CUENTA ", "").strip()
                
                # L19: 'BURSTEIN NORBERTO              C.U.I.T. 020-12290006-2...'
                if "C.U.I.T." in linea and titular_global == "S/D":
                    parts = linea.split("C.U.I.T.")
                    if len(parts) > 0:
                        titular_global = parts[0].strip()

                if capturar:
                    linea = linea.strip()
                    if re.match(r"^\d{2}/\d{2}/\d{2}", linea):
                        movimientos.append(linea)

                if "NUMERO DE CUENTA" in linea or "Nro.:" in linea:
                    capturar = True
                    # Intento 1: Formato "NUMERO DE CUENTA 00-00000000/0"
                    match = re.search(r"NUMERO DE CUENTA\s+(\d{2}-\d{8}/\d)", linea)
                    if not match:
                        # Intento 2: Formato "Nro.: 00053031-001"
                        match = re.search(r"Nro.:\s*([\d-]+)", linea)
                    
                    if match:
                        cuenta = {}
                        numero_cuenta = match.group(1)
                        # Verificar que no hayamos procesado esta cuenta ya (a veces se repite en header de paginas)
                        if not any(c['cuenta'] == numero_cuenta for c in cuentas):
                           cuenta["cuenta"] = numero_cuenta
                           cuentas.append(cuenta)
                        numero_de_cuenta_temporal = numero_cuenta

                if "Saldo del per" in linea and "anterior" in linea:
                     # Modificado para permitir espacios o signo negativo al final.
                    match = re.search(r"([\d\.]+,\d{2}[\-]?)", linea.strip())
                    if match:
                        importe_raw = match.group(1)
                        es_negativo = importe_raw.endswith("-")
                        importe_str = importe_raw.replace("-", "").replace(".", "").replace(",", ".")
                        importe = float(importe_str)
                        if es_negativo: importe *= -1

                        resultado = next(
                            (d for d in cuentas if d["cuenta"] == numero_de_cuenta_temporal),
                            None,
                        )
                        if resultado:
                            resultado["saldo_inicial"] = importe

                if "SALDO PERIODO ACTUAL" in linea:
                    resultado_movimientos = next(
                        (d for d in cuentas if d["cuenta"] == numero_de_cuenta_temporal),
                        None,
                    )
                    if resultado_movimientos:
                        resultado_movimientos["movimientos"] = movimientos.copy()
                        movimientos = []
                    capturar = False

                    match = re.search(r"([\d\.]+,\d{2}[\-]?)", linea.strip())
                    if match:
                        importe_raw = match.group(1)
                        es_negativo = importe_raw.endswith("-")
                        importe_str = importe_raw.replace("-", "").replace(".", "").replace(",", ".")
                        importe = float(importe_str)
                        if es_negativo: importe *= -1

                        resultado = next(
                            (d for d in cuentas if d["cuenta"] == numero_de_cuenta_temporal),
                            None,
                        )
                        if resultado:
                            resultado["saldo_final"] = importe

            def procesar_movimientos(movimientos_cuenta, saldo_inicial):
                movimientos_limpios = []
                # El saldo inicial viene del header "Saldo del período anterior"
                saldo_actual_calculado = saldo_inicial 
                
                # Regex para montos: numeros con puntos y coma decimal, OPCIONALMENTE signo menos al final
                # Agregamos ?: al grupo externo para no capturarlo si usamos findall
                pattern_monto = re.compile(r"((?:\d{1,3}(?:\.\d{3})*)?,\d{2}-?)")

                for movimiento in movimientos_cuenta:
                    # Formato esperado: "03/02/25  Descripcion....   Importe   Saldo"
                    
                    matches = pattern_monto.findall(movimiento)
                    
                    if len(matches) >= 2:
                        # LOGICA ESTANDAR: Tiene Importe y Saldo (al menos 2 montos)
                        # Asumimos que el ULTIMO es el Saldo Resultante
                        saldo_str_raw = matches[-1]
                        
                        es_negativo_saldo = saldo_str_raw.endswith("-")
                        saldo_limpio = saldo_str_raw.replace("-", "").replace(".", "").replace(",", ".")
                        saldo_linea = float(saldo_limpio)
                        if es_negativo_saldo: saldo_linea *= -1
                        
                        # Calculamos el importe por diferencia de saldos
                        # Importe = Saldo_Linea - Saldo_Anterior
                        importe_calculado = saldo_linea - saldo_actual_calculado
                        
                        # Limpiar descripcion
                        fecha = movimiento[:8]
                        resto = movimiento[9:].strip()
                        
                        # Intentar limpiar tokens finales (saldos/importes) de la descripcion
                        # Si la descripcion termina con el saldo encontrado, lo quitamos
                        if resto.endswith(saldo_str_raw):
                             resto = resto[:len(resto)-len(saldo_str_raw)].strip()
                        
                        mov_obj = {
                            "Fecha": fecha,
                            "Descripcion": resto.split("   ")[0], 
                            "Importe": importe_calculado
                        }
                        
                        movimientos_limpios.append(mov_obj)
                        saldo_actual_calculado = saldo_linea

                    elif len(matches) == 1:
                        # LOGICA ESTRICTA: Solo tiene un monto (Importe), NO tiene saldo.
                        # Ej: "30/12/25  Impuesto a las Ganancias       0206580294       3.567,97"
                        # En este caso, NO podemos calcular por diferencia. Usamos el monto directo.
                        # Y NO actualizamos saldo_actual_calculado porque la cadena de saldos parece saltar estos movimientos.

                        monto_str_raw = matches[0]
                        es_negativo_monto = monto_str_raw.endswith("-")
                        monto_limpio = monto_str_raw.replace("-", "").replace(".", "").replace(",", ".")
                        importe_directo = float(monto_limpio)
                        if es_negativo_monto: importe_directo *= -1
                        
                        fecha = movimiento[:8]
                        resto = movimiento[9:].strip()
                        
                        # Limpiar descripcion del monto al final
                        if resto.endswith(monto_str_raw):
                             resto = resto[:len(resto)-len(monto_str_raw)].strip()
                        
                        descripcion = resto.split("   ")[0]

                        # HEURISTICA DE SIGNO: Si no vino con signo negativo explicito,
                        # intentamos deducir si es DEBITO por palabras clave en la descripcion.
                        # (Si ya es negativo, lo dejamos asi)
                        if not es_negativo_monto:
                            # Palabras clave que indican SALIDA de dinero (Debito)
                            keywords_debito = [
                                "Impuesto", "IVA", "Comision", "Gasto", "Débito", 
                                "Retencion", "Percep", "IIBB", "Sellos", "Mantenimiento",
                                "Cheque Rechazado", "Debito", "DEB", "Credito DEBIN" 
                            ]
                            # Agregue mas keywords
                            if any(kw.lower() in descripcion.lower() for kw in keywords_debito):
                                importe_directo *= -1
                        
                        mov_obj = {
                            "Fecha": fecha,
                            "Descripcion": descripcion,
                            "Importe": importe_directo
                        }
                        movimientos_limpios.append(mov_obj)
                    else:
                        # No encontramos montos, quiza wrappeo de texto? Ignoramos por ahora
                        continue

                return movimientos_limpios

            return cuentas, procesar_movimientos, periodo_global, titular_global, texto
        
        # --- FIN LÓGICA ORIGINAL ---

        # Ejecutar extracción
        cuentas, procesar_movimientos_func, periodo, nombre_titular, texto_raw = procesar_pdf(archivo_pdf.read())



        if not cuentas:
            st.warning("No se encontraron cuentas en el PDF (Formato Original)")
            return None

        st.success(f"Se encontraron {len(cuentas)} cuenta(s)")

        # --- GENERACIÓN DE EXCEL DASHBOARD (ESTILO FORMATO 2 EXACTO) ---
        output = io.BytesIO()
        wb = Workbook()
        # Eliminar hoja default
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        for cuenta in cuentas:
            saldo_inicial = cuenta.get("saldo_inicial", 0.0)
            saldo_final = cuenta.get("saldo_final", 0.0)
            numero_cuenta = cuenta["cuenta"]
            movimientos_raw = cuenta.get("movimientos", [])

            # --- CALIBRACION AUTOMATICA CON FILA DE AJUSTE ---
            # Algunos movimientos (ej: 30/12) están listados pero YA incluidos en el Saldo Inicial (31/12).
            # Para respetar el "Saldo Inicial" exacto del PDF y a la vez que cierre el control:
            # Agregamos una fila de "Ajuste" que compense la diferencia.
            
            datos = procesar_movimientos_func(movimientos_raw, saldo_inicial)
            
            total_movimientos = sum(d["Importe"] for d in datos)
            saldo_final_teorico = saldo_inicial + total_movimientos
            diferencia = saldo_final_teorico - saldo_final
            
            # Si hay diferencia significativa (> 0.01), agregamos movimiento de ajuste
            if abs(diferencia) > 0.01:
                # --- VALIDACION INTELIGENTE ---
                # Chequeamos si la diferencia coincide EXACTAMENTE con la suma de los primeros movimientos.
                
                suma_acumulada = 0.0
                es_error_conocido = False
                indices_coincidentes = 0
                
                # Probamos sumando los primeros 5 movimientos a ver si alguno calza
                for i in range(min(5, len(datos))):
                    suma_acumulada += datos[i]["Importe"]
                    # Chequeamos si la diferencia es igual a esta suma acumulada (con tolerancia)
                    if abs(diferencia - suma_acumulada) < 0.01:
                        es_error_conocido = True
                        indices_coincidentes = i + 1
                        break
                
                if es_error_conocido:
                    # CASO 1: Movimientos pre-periodo detectados.
                    # El usuario pidio NO extraerlos si generan diferencia.
                    # Los eliminamos de la lista.
                    
                    st.warning(f"⚠️ **Ajuste Automático en Cuenta {numero_cuenta}**")
                    st.info(f"Se detectó que los primeros {indices_coincidentes} movimientos (Suma: ${suma_acumulada:,.2f}) sobran en el cálculo del saldo.")
                    st.success("✅ **Acción:** Se han eliminado estos movimientos del reporte para que el saldo cuadre perfecto.")
                    
                    # Eliminamos los N primeros
                    datos = datos[indices_coincidentes:]
                    
                else:
                    # CASO 2: Error desconocido.
                    # Mantenemos la lógica de fila de AJUSTE para alertar que algo esta mal.
                    st.error(f"❌ **Diferencia de Saldos NO explicada (${diferencia:,.2f}) en Cuenta {numero_cuenta}**")
                    st.warning("Podría haber un error de extracción (movimiento faltante o mal leido). Revisar el Excel.")
                    
                    ajuste_row = {
                        "Fecha": datos[0]["Fecha"] if datos else "",
                        "Descripcion": "AJUSTE POR DIFERENCIA DE SALDOS (REVISAR)",
                        "Importe": -diferencia
                    }
                    datos.insert(0, ajuste_row)

            df = pd.DataFrame(datos, columns=["Fecha", "Descripcion", "Importe"])

            # Separar Creditos y Debitos
            if not df.empty:
                creditos = df[df["Importe"] > 0].copy()
                debitos = df[df["Importe"] < 0].copy()
                debitos["Importe"] = debitos["Importe"].abs()
            else:
                creditos = pd.DataFrame(columns=["Fecha", "Descripcion", "Importe"])
                debitos = pd.DataFrame(columns=["Fecha", "Descripcion", "Importe"])
            nombre_hoja = numero_cuenta.replace("/", "-")[:30]
            ws = wb.create_sheet(title=nombre_hoja)
            
            # --- SETUP ESTILOS GLOBAL ---
            ws.sheet_view.showGridLines = False
            # Columnas: A,B,C (Creditos) - D (Control) - E,F,G (Debitos)
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 50
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 28 # Centro
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 50
            ws.column_dimensions['G'].width = 15

            thin_border = Border(left=Side(style='thin', color="A6A6A6"), 
                                 right=Side(style='thin', color="A6A6A6"), 
                                 top=Side(style='thin', color="A6A6A6"), 
                                 bottom=Side(style='thin', color="A6A6A6"))
            
            color_bg_main = "2C3E50"
            color_txt_main = "FFFFFF"
            
            # Colores Débitos (Rojos)
            fill_head_deb = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
            fill_col_deb = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
            fill_row_deb = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")

            # Colores Créditos (Verdes)
            fill_head_cred = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
            fill_col_cred = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")
            fill_row_cred = PatternFill(start_color="F2F9F1", end_color="F2F9F1", fill_type="solid")

            # Procesar datos (YA REALIZADO ARRIBA para calibracion)
            # datos = procesar_movimientos_func(movimientos_raw, saldo_inicial)
            # df = pd.DataFrame(datos, columns=["Fecha", "Descripcion", "Importe"])

            # --- HEADER ---
            ws.merge_cells("A1:G1")
            tit = ws["A1"]
            tit.value = f"REPORTE SUPERVIELLE - CTA {numero_cuenta}"
            tit.font = Font(size=14, bold=True, color=color_txt_main)
            tit.fill = PatternFill(start_color=color_bg_main, end_color=color_bg_main, fill_type="solid")
            tit.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 25

            # Saldo Inicial (AJUSTADO)
            ws["A3"] = "SALDO INICIAL"
            ws["A3"].font = Font(bold=True, size=10, color="666666")
            ws["B3"] = saldo_inicial 
            ws["B3"].number_format = '"$ "#,##0.00'
            ws["B3"].font = Font(bold=True, size=11)
            ws["B3"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

            # Saldo Final
            ws["A4"] = "SALDO FINAL"
            ws["A4"].font = Font(bold=True, size=10, color="666666")
            ws["B4"] = saldo_final
            ws["B4"].number_format = '"$ "#,##0.00'
            ws["B4"].font = Font(bold=True, size=11)
            ws["B4"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

            # Titular / Cuenta (Derecha)
            ws["D3"] = "TITULAR"
            ws["D3"].alignment = Alignment(horizontal='right')
            ws["D3"].font = Font(bold=True, color="666666", size=10)
            
            # Usamos el nombre extraído si existe, sino el número de cuenta
            val_titular = nombre_titular if nombre_titular != "S/D" else numero_cuenta
            ws["E3"] = clean_for_excel(val_titular)
            ws["E3"].font = Font(bold=True, size=11)
            ws["E3"].alignment = Alignment(horizontal='center')
            ws.merge_cells("E3:G3")
            for c in ["E","F","G"]: ws[f"{c}3"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

            # Período
            ws["D4"] = "PERÍODO"
            ws["D4"].alignment = Alignment(horizontal='right')
            ws["D4"].font = Font(bold=True, color="666666", size=10)
            ws["E4"] = clean_for_excel(periodo)
            ws["E4"].font = Font(bold=True, size=11)
            ws["E4"].alignment = Alignment(horizontal='center')
            ws.merge_cells("E4:G4")
            for c in ["E","F","G"]: ws[f"{c}4"].border = Border(bottom=Side(style='thin', color="DDDDDD"))


            # Control Central
            ws["D6"] = "CONTROL DE SALDOS"
            ws["D6"].font = Font(bold=True, size=10, color="666666")
            ws["D6"].alignment = Alignment(horizontal='center')
            
            cell_ctl = ws["D7"]
            cell_ctl.font = Font(bold=True, size=12)
            cell_ctl.alignment = Alignment(horizontal='center')
            cell_ctl.border = thin_border
            
            # --- TABLAS (Inicio Fila 10) ---
            fila_inicio = 10
            
            # CRÉDITOS (A-C)
            f_header = fila_inicio
            ws.merge_cells(f"A{f_header}:C{f_header}")
            ws[f"A{f_header}"] = "CRÉDITOS" 
            ws[f"A{f_header}"].fill = fill_head_cred
            ws[f"A{f_header}"].font = Font(bold=True, color="FFFFFF")
            ws[f"A{f_header}"].alignment = Alignment(horizontal='center')
            ws[f"A{f_header}"].border = thin_border
            
            headers = ["Fecha", "Descripción", "Importe"]
            cols_cred = ["A", "B", "C"]
            f_sub = f_header + 1
            for i, h in enumerate(headers):
                c = ws[f"{cols_cred[i]}{f_sub}"]
                c.value = h
                c.fill = fill_col_cred
                c.font = Font(bold=True)
                c.alignment = Alignment(horizontal='center')
                c.border = thin_border
            
            # DÉBITOS (E-G)
            ws.merge_cells(f"E{f_header}:G{f_header}")
            ws[f"E{f_header}"] = "DÉBITOS" 
            ws[f"E{f_header}"].fill = fill_head_deb
            ws[f"E{f_header}"].font = Font(bold=True, color="FFFFFF")
            ws[f"E{f_header}"].alignment = Alignment(horizontal='center')
            ws[f"E{f_header}"].border = thin_border
            
            cols_deb = ["E", "F", "G"]
            for i, h in enumerate(headers):
                c = ws[f"{cols_deb[i]}{f_sub}"]
                c.value = h
                c.fill = fill_col_deb
                c.font = Font(bold=True)
                c.alignment = Alignment(horizontal='center')
                c.border = thin_border
            
            # --- LLENADO DE DATOS (PARALELO) ---
            fila_dato_start = f_sub + 1
            
            # 1. CRÉDITOS
            f_cred = fila_dato_start
            if creditos.empty:
                ws.merge_cells(f"A{f_cred}:C{f_cred}")
                ws[f"A{f_cred}"] = "SIN MOVIMIENTOS"
                ws[f"A{f_cred}"].font = Font(italic=True, color="666666")
                ws[f"A{f_cred}"].alignment = Alignment(horizontal='center')
                ws[f"A{f_cred}"].border = thin_border
                f_cred += 1 # Ocupa 1 fila
            else:
                start_c = f_cred
                for _, r in creditos.iterrows():
                    ws[f"A{f_cred}"] = clean_for_excel(r["Fecha"])
                    ws[f"A{f_cred}"].fill = fill_row_cred
                    ws[f"A{f_cred}"].alignment = Alignment(horizontal='center')
                    ws[f"A{f_cred}"].border = thin_border

                    ws[f"B{f_cred}"] = clean_for_excel(r["Descripcion"])
                    ws[f"B{f_cred}"].fill = fill_row_cred
                    ws[f"B{f_cred}"].border = thin_border

                    ws[f"C{f_cred}"] = r["Importe"]
                    ws[f"C{f_cred}"].number_format = '"$ "#,##0.00'
                    ws[f"C{f_cred}"].fill = fill_row_cred
                    ws[f"C{f_cred}"].border = thin_border
                    f_cred += 1
                
                # Total Créditos
                ws[f"B{f_cred}"] = "TOTAL CRÉDITOS"
                ws[f"B{f_cred}"].font = Font(bold=True)
                ws[f"B{f_cred}"].alignment = Alignment(horizontal='right')
                ws[f"C{f_cred}"] = f"=SUM(C{start_c}:C{f_cred-1})"
                ws[f"C{f_cred}"].number_format = '"$ "#,##0.00'
                ws[f"C{f_cred}"].font = Font(bold=True)
                f_cred += 1 # Dejar espacio tras total

            # 2. DÉBITOS
            f_deb = fila_dato_start
            if debitos.empty:
                ws.merge_cells(f"E{f_deb}:G{f_deb}")
                ws[f"E{f_deb}"] = "SIN MOVIMIENTOS"
                ws[f"E{f_deb}"].font = Font(italic=True, color="666666")
                ws[f"E{f_deb}"].alignment = Alignment(horizontal='center')
                ws[f"E{f_deb}"].border = thin_border
                f_deb += 1
            else:
                start_d = f_deb
                for _, r in debitos.iterrows():
                    ws[f"E{f_deb}"] = clean_for_excel(r["Fecha"])
                    ws[f"E{f_deb}"].fill = fill_row_deb
                    ws[f"E{f_deb}"].alignment = Alignment(horizontal='center')
                    ws[f"E{f_deb}"].border = thin_border

                    ws[f"F{f_deb}"] = clean_for_excel(r["Descripcion"])
                    ws[f"F{f_deb}"].fill = fill_row_deb
                    ws[f"F{f_deb}"].border = thin_border

                    ws[f"G{f_deb}"] = r["Importe"]
                    ws[f"G{f_deb}"].number_format = '"$ "#,##0.00'
                    ws[f"G{f_deb}"].fill = fill_row_deb
                    ws[f"G{f_deb}"].border = thin_border
                    f_deb += 1
                
                # Total Débitos
                ws[f"F{f_deb}"] = "TOTAL DÉBITOS"
                ws[f"F{f_deb}"].font = Font(bold=True)
                ws[f"F{f_deb}"].alignment = Alignment(horizontal='right')
                ws[f"G{f_deb}"] = f"=SUM(G{start_d}:G{f_deb-1})"
                ws[f"G{f_deb}"].number_format = '"$ "#,##0.00'
                ws[f"G{f_deb}"].font = Font(bold=True)
                f_deb += 1

            # --- FORMULA KPI CONTROL ---
            # Debe abarcar el máximo de filas usadas
            # Si NO hubo movimientos, la suma será 0 (rangos válidos ficticios o controlados)
            # Usamos fila_dato_start hasta max(f_cred, f_deb) - 1 (pero cuidado si sólo hay 1 fila)
            # Mejor usar SUMIF o rango fijo grande
            
            # Referencias a sumas totales
            # Si creditos.empty, suma = 0. Si no, suma = C{f_cred-1}
            # PERO f_cred ya incrementó. La celda total está en f_cred-1
            
            ref_tot_cred = f"C{f_cred-1}" if not creditos.empty else "0"
            ref_tot_deb = f"G{f_deb-1}" if not debitos.empty else "0"
            
            formula_ctl = f"=ROUND(B3 + {ref_tot_cred} - {ref_tot_deb} - B4, 2)"
            ws["D7"].value = formula_ctl
            ws["D7"].number_format = '"$ "#,##0.00'


        wb.save(output) 
        output.seek(0)
        return output.getvalue()

    except Exception as e:
        import traceback
        st.error(f"Error CRÍTICO al procesar el archivo: {str(e)}")
        print(traceback.format_exc()) # Debug en consola
        return None
