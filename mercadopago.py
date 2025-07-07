import streamlit as st
import PyPDF2
import re
import pandas as pd
import io


def limpiar_nombre_hoja(nombre):
    """Limpia el nombre para que sea válido como nombre de hoja de Excel"""
    # Caracteres no permitidos en nombres de hojas de Excel
    caracteres_invalidos = ["\\", "/", "*", "[", "]", ":", "?"]
    nombre_limpio = nombre
    for char in caracteres_invalidos:
        nombre_limpio = nombre_limpio.replace(char, "_")

    # Limitar longitud a 31 caracteres (límite de Excel)
    if len(nombre_limpio) > 31:
        nombre_limpio = nombre_limpio[:31]

    return nombre_limpio


def procesar_mercadopago(archivo_pdf):
    """Procesa archivos PDF de MercadoPago"""

    try:
        # Reinicializar el archivo para lectura
        archivo_pdf.seek(0)

        # Abrir y leer el archivo PDF
        reader = PyPDF2.PdfReader(io.BytesIO(archivo_pdf.read()))
        texto = "".join(page.extract_text() + "\n" for page in reader.pages)
        lineas = texto.splitlines()

        # Variables para almacenar la información
        saldo_inicial = None
        saldo_final = None
        movimientos = []
        nombre_titular = None
        cvu = None

        # Procesar líneas
        i = 0
        while i < len(lineas):
            linea = lineas[i].strip()

            # Extraer nombre del titular (primera línea después de "RESUMEN DE CUENTA")
            if i > 0 and lineas[i - 1].strip() == "RESUMEN DE CUENTA" and linea:
                nombre_titular = linea

            # Extraer CVU
            if linea.startswith("CVU:"):
                cvu_match = re.search(r"CVU:\s*(\d+)", linea)
                if cvu_match:
                    cvu = cvu_match.group(1)

            # Extraer saldo inicial
            if "Saldo inicial:" in linea:
                saldo_match = re.search(r"Saldo inicial:\s*\$\s*([\d,.]+)", linea)
                if saldo_match:
                    saldo_inicial = saldo_match.group(1)

            # Extraer saldo final
            if "Saldo final:" in linea:
                saldo_match = re.search(r"Saldo final:\s*\$\s*([\d,.]+)", linea)
                if saldo_match:
                    saldo_final = saldo_match.group(1)

            # Extraer movimientos (líneas que empiezan con fecha)
            if re.match(r"^\d{2}-\d{2}-\d{4}", linea):
                # Línea actual puede ser el inicio de un movimiento
                linea_movimiento = linea

                # Verificar si la línea actual contiene los montos
                if not re.search(r"\$\s*-?[\d,]+\.?\d*", linea_movimiento):
                    # Si no tiene montos, combinar con la siguiente línea
                    if i + 1 < len(lineas):
                        linea_siguiente = lineas[i + 1].strip()
                        linea_movimiento = linea + " " + linea_siguiente
                        i += 1  # Saltar la siguiente línea ya que la procesamos

                # Limpiar la línea: quitar saltos de línea internos y espacios extra
                linea_movimiento = " ".join(linea_movimiento.split())

                # Extraer fecha usando regex (primeros 10 caracteres en formato DD-MM-YYYY)
                fecha_match = re.match(r"^(\d{2}-\d{2}-\d{4})", linea_movimiento)
                if fecha_match:
                    fecha = fecha_match.group(1)

                    # Buscar montos con regex mejorado - incluir decimales opcionales
                    # Primero buscar todos los fragmentos de números
                    fragmentos_numericos = re.findall(r"[\d.,]+", linea_movimiento)

                    # Reconstruir montos válidos
                    montos_validos = []
                    i_frag = 0
                    while i_frag < len(fragmentos_numericos):
                        fragmento = fragmentos_numericos[i_frag]

                        # Si el fragmento termina en coma, buscar el siguiente fragmento como decimales
                        if fragmento.endswith(",") and i_frag + 1 < len(
                            fragmentos_numericos
                        ):
                            siguiente = fragmentos_numericos[i_frag + 1]
                            # Si el siguiente fragmento son solo 2 dígitos, es parte decimal
                            if re.match(r"^\d{2}$", siguiente):
                                monto_completo = fragmento + siguiente
                                montos_validos.append(monto_completo)
                                i_frag += 2  # Saltar el siguiente fragmento
                                continue

                        # Verificar si es un monto válido (formato argentino)
                        if re.match(r"^\d{1,3}(?:\.\d{3})*(?:,\d{2})?$", fragmento):
                            montos_validos.append(fragmento)

                        i_frag += 1

                    if len(montos_validos) >= 2:
                        importe = montos_validos[-2]  # Penúltimo monto válido

                        # Detectar si el importe es negativo buscando el signo - antes del monto
                        # Buscar la posición del importe en la línea
                        posicion_importe = linea_movimiento.find(importe)
                        if posicion_importe > 0:
                            # Revisar los caracteres antes del importe para buscar el signo -
                            texto_antes = linea_movimiento[:posicion_importe]
                            # Buscar el último $ seguido opcionalmente de espacios y -
                            if re.search(r"\$\s*-\s*$", texto_antes):
                                importe = "-" + importe

                        # Extraer descripción (todo después de la fecha hasta antes del ID y montos)
                        # Remover la fecha del inicio
                        resto_linea = linea_movimiento[
                            10:
                        ].strip()  # Quitar los primeros 10 caracteres (fecha)

                        # Buscar el ID (número largo) para separar descripción de montos
                        descripcion_match = re.search(
                            r"^(.+?)\s*\d{10,}\s*", resto_linea
                        )
                        if descripcion_match:
                            descripcion = descripcion_match.group(1).strip()
                        else:
                            # Fallback: tomar todo hasta antes de los montos
                            descripcion_match = re.search(
                                r"^(.+?)\s*[\d.,]+", resto_linea
                            )
                            if descripcion_match:
                                descripcion = descripcion_match.group(1).strip()
                            else:
                                # Último fallback: tomar las primeras palabras
                                partes = resto_linea.split()
                                descripcion = (
                                    " ".join(partes[:-3])
                                    if len(partes) > 3
                                    else resto_linea
                                )

                        movimiento = {
                            "Fecha": fecha,
                            "Descripcion": descripcion,
                            "Importe": importe,
                        }
                        movimientos.append(movimiento)

                    elif len(montos_validos) == 1:
                        # Si solo hay un monto válido
                        importe = montos_validos[0]

                        # Detectar si el importe es negativo
                        posicion_importe = linea_movimiento.find(importe)
                        if posicion_importe > 0:
                            texto_antes = linea_movimiento[:posicion_importe]
                            if re.search(r"\$\s*-\s*$", texto_antes):
                                importe = "-" + importe

                        # Extraer descripción
                        resto_linea = linea_movimiento[
                            10:
                        ].strip()  # Quitar los primeros 10 caracteres (fecha)

                        descripcion_match = re.search(
                            r"^(.+?)\s*\d{10,}\s*", resto_linea
                        )
                        if descripcion_match:
                            descripcion = descripcion_match.group(1).strip()
                        else:
                            descripcion_match = re.search(
                                r"^(.+?)\s*[\d.,]+", resto_linea
                            )
                            if descripcion_match:
                                descripcion = descripcion_match.group(1).strip()
                            else:
                                partes = resto_linea.split()
                                descripcion = (
                                    " ".join(partes[:-2])
                                    if len(partes) > 2
                                    else resto_linea
                                )

                        movimiento = {
                            "Fecha": fecha,
                            "Descripcion": descripcion,
                            "Importe": importe,
                        }
                        movimientos.append(movimiento)

            i += 1

        # Crear el archivo Excel
        if saldo_inicial and saldo_final:
            try:
                output = io.BytesIO()

                # Crear DataFrame con los movimientos
                df = pd.DataFrame(movimientos)

                # Función simple para convertir importes a numérico
                def convertir_a_numerico(importe_str):
                    """Convierte importe a numérico tratando punto como separador de miles y coma como decimal"""
                    if not importe_str:
                        return 0.0

                    # Limpiar espacios y detectar signo
                    importe_str = str(importe_str).strip()
                    signo = -1 if importe_str.startswith("-") else 1
                    importe_str = importe_str.lstrip("-").strip()

                    # Formato argentino: punto = separador de miles, coma = decimal
                    # Ejemplos: -1.400 = -1400, 33.688,50 = 33688.50, 1.234,56 = 1234.56

                    if "," in importe_str:
                        # Tiene decimales: 33.688,50
                        partes = importe_str.split(",")
                        parte_entera = partes[0].replace(
                            ".", ""
                        )  # Quitar puntos de miles: 33.688 -> 33688
                        parte_decimal = partes[1]  # Mantener decimales: 50
                        numero_str = f"{parte_entera}.{parte_decimal}"  # 33688.50
                    else:
                        # Solo enteros con separador de miles: 1.400 -> 1400
                        numero_str = importe_str.replace(
                            ".", ""
                        )  # Quitar puntos: 1.400 -> 1400

                    try:
                        return signo * float(numero_str)
                    except ValueError:
                        # Si no se puede convertir, devolver 0
                        return 0.0

                # Convertir la columna Importe a numérico
                if not df.empty:
                    df["Importe"] = df["Importe"].apply(convertir_a_numerico)

                # Separar movimientos en créditos y débitos
                creditos = (
                    df[df["Importe"] > 0].copy()
                    if not df.empty
                    else pd.DataFrame(columns=["Fecha", "Descripcion", "Importe"])
                )
                debitos = (
                    df[df["Importe"] < 0].copy()
                    if not df.empty
                    else pd.DataFrame(columns=["Fecha", "Descripcion", "Importe"])
                )

                # Convertir débitos a valores absolutos para mejor visualización
                if not debitos.empty:
                    debitos["Importe"] = debitos["Importe"].abs()

                # Nombre de la hoja
                nombre_hoja = (
                    f"MercadoPago {nombre_titular[:15] if nombre_titular else 'Cuenta'}"
                )
                nombre_limpio = limpiar_nombre_hoja(nombre_hoja)

                # Crear el workbook y worksheet manualmente
                from openpyxl import Workbook

                wb = Workbook()
                ws = wb.active
                ws.title = nombre_limpio

                # Agregar información del CVU y saldos (sin titular)
                ws["A1"] = "CVU:"
                ws["B1"] = cvu if cvu else "No disponible"

                # Agregar saldos con formato numérico
                ws["A2"] = "Saldo Inicial:"
                if saldo_inicial:
                    ws["B2"] = convertir_a_numerico(saldo_inicial)
                    ws["B2"].number_format = "#,##0.00"
                else:
                    ws["B2"] = "No disponible"

                ws["A3"] = "Saldo Final:"
                if saldo_final:
                    ws["B3"] = convertir_a_numerico(saldo_final)
                    ws["B3"].number_format = "#,##0.00"
                else:
                    ws["B3"] = "No disponible"

                # Tabla de DÉBITOS (columnas A-C, empezar en fila 5)
                fila_actual = 5
                ws[f"A{fila_actual}"] = "DÉBITOS"
                ws[f"A{fila_actual}"].font = ws[f"A{fila_actual}"].font.copy(bold=True)
                fila_actual += 1

                # Headers para débitos
                ws[f"A{fila_actual}"] = "Fecha"
                ws[f"B{fila_actual}"] = "Descripción"
                ws[f"C{fila_actual}"] = "Importe"

                # Hacer headers en negrita
                for col in ["A", "B", "C"]:
                    ws[f"{col}{fila_actual}"].font = ws[
                        f"{col}{fila_actual}"
                    ].font.copy(bold=True)

                fila_actual += 1
                inicio_debitos = fila_actual

                # Agregar datos de débitos
                for _, row in debitos.iterrows():
                    ws[f"A{fila_actual}"] = row["Fecha"]
                    ws[f"B{fila_actual}"] = row["Descripcion"]
                    ws[f"C{fila_actual}"] = row["Importe"]
                    ws[f"C{fila_actual}"].number_format = "#,##0.00"
                    fila_actual += 1

                # Agregar total de débitos
                if not debitos.empty:
                    ws[f"B{fila_actual}"] = "TOTAL DÉBITOS:"
                    ws[f"B{fila_actual}"].font = ws[f"B{fila_actual}"].font.copy(
                        bold=True
                    )
                    ws[f"C{fila_actual}"] = debitos["Importe"].sum()
                    ws[f"C{fila_actual}"].number_format = "#,##0.00"
                    ws[f"C{fila_actual}"].font = ws[f"C{fila_actual}"].font.copy(
                        bold=True
                    )
                    fila_total_debitos = fila_actual
                else:
                    fila_total_debitos = None

                # Tabla de CRÉDITOS (columnas E-G, empezar en fila 5)
                fila_creditos = 5
                ws[f"E{fila_creditos}"] = "CRÉDITOS"
                ws[f"E{fila_creditos}"].font = ws[f"E{fila_creditos}"].font.copy(
                    bold=True
                )
                fila_creditos += 1

                # Headers para créditos
                ws[f"E{fila_creditos}"] = "Fecha"
                ws[f"F{fila_creditos}"] = "Descripción"
                ws[f"G{fila_creditos}"] = "Importe"

                # Hacer headers en negrita
                for col in ["E", "F", "G"]:
                    ws[f"{col}{fila_creditos}"].font = ws[
                        f"{col}{fila_creditos}"
                    ].font.copy(bold=True)

                fila_creditos += 1
                inicio_creditos = fila_creditos

                # Agregar datos de créditos
                for _, row in creditos.iterrows():
                    ws[f"E{fila_creditos}"] = row["Fecha"]
                    ws[f"F{fila_creditos}"] = row["Descripcion"]
                    ws[f"G{fila_creditos}"] = row["Importe"]
                    ws[f"G{fila_creditos}"].number_format = "#,##0.00"
                    fila_creditos += 1

                # Agregar total de créditos
                if not creditos.empty:
                    ws[f"F{fila_creditos}"] = "TOTAL CRÉDITOS:"
                    ws[f"F{fila_creditos}"].font = ws[f"F{fila_creditos}"].font.copy(
                        bold=True
                    )
                    ws[f"G{fila_creditos}"] = creditos["Importe"].sum()
                    ws[f"G{fila_creditos}"].number_format = "#,##0.00"
                    ws[f"G{fila_creditos}"].font = ws[f"G{fila_creditos}"].font.copy(
                        bold=True
                    )
                    fila_total_creditos = fila_creditos
                else:
                    fila_total_creditos = None

                # Agregar fórmula de control en J7
                # Fórmula: Saldo Inicial + Total Créditos - Total Débitos - Saldo Final
                ws["I7"] = "Control:"
                ws["I7"].font = ws["I7"].font.copy(bold=True)

                # Construir la fórmula dinámicamente
                formula_parts = ["B2"]  # Saldo inicial

                if fila_total_creditos:
                    formula_parts.append(f"G{fila_total_creditos}")  # Total créditos
                else:
                    formula_parts.append("0")  # Si no hay créditos

                if fila_total_debitos:
                    formula_parts.append(
                        f"C{fila_total_debitos}"
                    )  # Total débitos (se resta)
                else:
                    formula_parts.append("0")  # Si no hay débitos

                formula_parts.append("B3")  # Saldo final (se resta)

                # Crear la fórmula: =B2+G[fila_creditos]-C[fila_debitos]-B3
                formula = f"=B2+{formula_parts[1]}-{formula_parts[2]}-B3"
                ws["J7"] = formula
                ws["J7"].number_format = "#,##0.00"
                ws["J7"].font = ws["J7"].font.copy(bold=True)

                # Ajustar ancho de columnas
                ws.column_dimensions["A"].width = 12  # Fecha débitos
                ws.column_dimensions["B"].width = 50  # Descripción débitos
                ws.column_dimensions["C"].width = 15  # Importe débitos
                ws.column_dimensions["E"].width = 12  # Fecha créditos
                ws.column_dimensions["F"].width = 50  # Descripción créditos
                ws.column_dimensions["G"].width = 15  # Importe créditos
                ws.column_dimensions["I"].width = 12  # Etiqueta control
                ws.column_dimensions["J"].width = 15  # Fórmula control

                # Guardar en BytesIO
                wb.save(output)

                # Preparar el archivo para descarga
                output.seek(0)

                st.success(f"Archivo Excel creado con {len(movimientos)} movimientos")
                return output.getvalue()

            except Exception as e:
                st.error(f"Error creando archivo Excel: {str(e)}")
                return None
        else:
            st.warning("No se encontraron saldos inicial y final en el PDF")
            return None

    except Exception as e:
        st.error(f"Error al procesar el archivo: {str(e)}")
        import traceback

        st.error(f"Detalles del error: {traceback.format_exc()}")
        return None
