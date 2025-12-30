import streamlit as st
import io
import pdfplumber
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule

# Regex para caracteres ilegales en Excel
ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

def clean_for_excel(text):
    """Elimina caracteres ilegales para Excel y espacios extra"""
    if not text: return ""
    text = str(text)
    text = ILLEGAL_CHARACTERS_RE.sub("", text)
    return text.strip()

def parse_importe(importe_str):
    """Convierte string de importe ($ -1.234,56) a float"""
    try:
        # Eliminar símbolo moneda y espacios
        clean = importe_str.replace("$", "").strip()
        # Eliminar puntos de miles y reemplazar coma decimal
        clean = clean.replace(".", "").replace(",", ".")
        return float(clean)
    except:
        return 0.0

def procesar_icbc_formato_2(archivo_pdf):
    """Procesa archivos PDF de ICBC Formato 2 con Estilo Dashboard"""
    st.info("Procesando archivo ICBC (Formato 2)...")

    try:
        # Leer PDF completo
        with pdfplumber.open(io.BytesIO(archivo_pdf.read())) as pdf:
            texto_completo = ""
            for page in pdf.pages:
                texto_completo += page.extract_text() + "\n"
            
        lineas = texto_completo.splitlines()
        
        # --- 1. Extracción de Metadata ---
        titular_global = "Sin Especificar"
        periodo_global = "Sin Especificar"
        
        # Buscar Titular y Periodo en las primeras líneas
        for i, l in enumerate(lineas[:20]):
            # Titular: suele estar solo en una linea o cerca de CUIT
            # En el ejemplo: "MOVE ART GROUP S.A. |"
            # Estrategia: Buscar lineas con texto en mayusculas que no sean encabezados conocidos
            if "Cuentas CC" in l:
                # A veces el titular esta en la linea SIGUIENTE a "Cuentas CC"
                if i + 1 < len(lineas):
                    cand = lineas[i+1].strip()
                    if "|" in cand: cand = cand.split("|")[0]
                    titular_global = cand.strip()
            
            # Periodo: "FILTROS Fecha desde:01-may.-2025 Fecha hasta:31-may.-2025"
            if "FILTROS" in l and "Fecha desde" in l:
                busqueda = re.search(r"Fecha desde:(.*?) Fecha hasta:(.*)", l)
                if busqueda:
                    desde = busqueda.group(1).strip()
                    hasta = busqueda.group(2).strip()
                    periodo_global = f"Del {desde} al {hasta}"
        
        # --- 2. Extracción de Movimientos ---
        # Regex: Fecha (DD-mmm.-YYYY) + Descripcion + Importe + Saldo
        # Ej: 30-may-2025 TRANSF CONNBKG $ -250.271,00 $ -735.333,38
        # Nota: El mes puede ser 'may' o 'may.' o '05' dependiendo variante, ajustamos regex
        # Pattern:
        # 1. Fecha: (\d{2}-[\w\.]+-\d{4})
        # 2. Descripcion: (.+?) (non-greedy)
        # 3. Importe: (\$\s*[-]?(?:\d{1,3}(?:[\.,]\d{3})*)?[\.,]\d{2})
        #    Permite importes como "$-,63" o "$ -0,63"
        # 4. Saldo:   Igual al importe
        
        regex_mov = r"(\d{2}-[\w\.]+-\d{4})\s+(.+?)\s+(\$\s*[-]?(?:\d{1,3}(?:[\.,]\d{3})*)?[\.,]\d{2})\s+(\$\s*[-]?(?:\d{1,3}(?:[\.,]\d{3})*)?[\.,]\d{2})"
        
        movimientos = []
        
        for l in lineas:
            match = re.search(regex_mov, l)
            if match:
                fecha = match.group(1)
                desc = match.group(2).strip()
                imp_str = match.group(3)
                saldo_str = match.group(4)
                
                importe = parse_importe(imp_str)
                saldo = parse_importe(saldo_str)
                
                movimientos.append({
                    "Fecha": fecha,
                    "Descripcion": desc,
                    "Importe": importe,
                    "Saldo": saldo
                })
        
        if not movimientos:
            st.error("No se encontraron movimientos. Verifique el formato.")
            return None

        # Convertimos a DataFrame
        df = pd.DataFrame(movimientos)
        
        # DEBUG: Mostrar primer y ultimo movimiento capturado antes de ordenar
        print("DEBUG - Primer movimiento capturado (Top PDF):", df.iloc[0].to_dict())
        print("DEBUG - Ultimo movimiento capturado (Bottom PDF):", df.iloc[-1].to_dict())

        # Ordenar cronológicamente (Invertir porque el PDF viene Descendente)
        df = df.iloc[::-1].reset_index(drop=True)
        
        # Calcular Saldos (Ahora df está ordenado: [0] = Más antiguo, [-1] = Más reciente)
        
        # Calcular Saldos (Cronológico Correcto para PDF Descendente)
        # El PDF viene ordenado de Reciente (Top) a Antiguo (Bottom).
        # DF ya está invertido cronológicamente: [0] = Antiguo (Bottom PDF), [-1] = Reciente (Top PDF).
        
        # Saldo Inicial: Se basa en el movimiento MÁS ANTIGUO (Bottom del PDF, fila [0] del df ordenado)
        mov_mas_antiguo = df.iloc[0]
        saldo_inicial = mov_mas_antiguo["Saldo"] - mov_mas_antiguo["Importe"]
        
        # Saldo Final: Es el saldo resultante del movimiento MÁS RECIENTE (Top del PDF, fila [-1] del df)
        mov_mas_reciente = df.iloc[-1]
        saldo_final = mov_mas_reciente["Saldo"]
        
        print(f"DEBUG - Lógica Cronológica Estándar")
        print(f"DEBUG - Movimiento Mas Antiguo ([0]): {mov_mas_antiguo['Fecha']} | Saldo: {mov_mas_antiguo['Saldo']} | Imp: {mov_mas_antiguo['Importe']}")
        print(f"DEBUG - Saldo Inicial Calc (Antiguo - Importe): {saldo_inicial}")
        print(f"DEBUG - Movimiento Mas Reciente ([-1]): {mov_mas_reciente['Fecha']} | Saldo: {mov_mas_reciente['Saldo']}")
        print(f"DEBUG - Saldo Final (Saldo Reciente): {saldo_final}")

        # Formatear Fecha a dd/mm/aaaa
        # Diccionario de meses
        meses = {
            "ene": "01", "feb": "02", "mar": "03", "abr": "04", "may": "05", "jun": "06",
            "jul": "07", "ago": "08", "sep": "09", "oct": "10", "nov": "11", "dic": "12"
        }
        
        def format_fecha(f):
            # Entrada esperada: 30-may-2025 o 30-may.-2025 o 05-06-2025
            try:
                # Caso 1: dd-mmm-yyyy
                match = re.search(r"(\d{2})-([a-z]{3})\.?-(20\d{2})", f.lower())
                if match:
                    d, m_str, y = match.groups()
                    if m_str in meses:
                        return f"{d}/{meses[m_str]}/{y}"
                
                # Caso 2: dd-mm-yyyy (Ej: 05-06-2025 pagina 1)
                match_num = re.search(r"(\d{2})-(\d{2})-(20\d{2})", f)
                if match_num:
                    return f"{match_num.group(1)}/{match_num.group(2)}/{match_num.group(3)}"
                
                return f
            except: 
                return f

        df["Fecha"] = df["Fecha"].apply(format_fecha)

        # --- GENERACIÓN EXCEL (ESTILO DASHBOARD) ---
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte ICBC"
        ws.sheet_view.showGridLines = False
        
        # Estilos
        color_bg_main = "C5001A" # Rojo ICBC
        color_txt_main = "FFFFFF"
        
        thin_border = Border(left=Side(style='thin', color="A6A6A6"), 
                             right=Side(style='thin', color="A6A6A6"), 
                             top=Side(style='thin', color="A6A6A6"), 
                             bottom=Side(style='thin', color="A6A6A6"))
        
        fill_head_deb = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        fill_col_deb = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
        fill_row_deb = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")

        fill_head_cred = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        fill_col_cred = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")
        fill_row_cred = PatternFill(start_color="F2F9F1", end_color="F2F9F1", fill_type="solid")

        # Separar Creditos y Debitos
        creditos = df[df["Importe"] > 0].copy()
        debitos = df[df["Importe"] < 0].copy()
        debitos["Importe"] = debitos["Importe"].abs()

        # 1. Header Global
        ws.merge_cells("A1:G1")
        tit = ws["A1"]
        tit.value = f"REPORTE ICBC - {clean_for_excel(titular_global)}"
        tit.font = Font(size=14, bold=True, color=color_txt_main)
        tit.fill = PatternFill(start_color=color_bg_main, end_color=color_bg_main, fill_type="solid")
        tit.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25

        # 2. Metadata y Saldos
        ws["A3"] = "SALDO INICIAL"
        ws["A3"].font = Font(bold=True, size=10, color="666666")
        ws["B3"] = saldo_inicial
        ws["B3"].number_format = '"$ "#,##0.00'
        ws["B3"].font = Font(bold=True, size=11)
        ws["B3"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

        ws["A4"] = "SALDO FINAL"
        ws["A4"].font = Font(bold=True, size=10, color="666666")
        ws["B4"] = saldo_final
        ws["B4"].number_format = '"$ "#,##0.00'
        ws["B4"].font = Font(bold=True, size=11)
        ws["B4"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

        ws["D3"] = "TITULAR"
        ws["D3"].alignment = Alignment(horizontal='right')
        ws["D3"].font = Font(bold=True, color="666666", size=10)
        ws["E3"] = clean_for_excel(titular_global)
        ws["E3"].font = Font(bold=True, size=11)
        ws["E3"].alignment = Alignment(horizontal='center')
        ws.merge_cells("E3:G3")
        for c in ["E","F","G"]: ws[f"{c}3"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

        ws["D4"] = "PERÍODO"
        ws["D4"].alignment = Alignment(horizontal='right')
        ws["D4"].font = Font(bold=True, color="666666", size=10)
        ws["E4"] = clean_for_excel(periodo_global)
        ws["E4"].font = Font(bold=True, size=11)
        ws["E4"].alignment = Alignment(horizontal='center')
        ws.merge_cells("E4:G4")
        for c in ["E","F","G"]: ws[f"{c}4"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

        ws["D6"] = "CONTROL DE SALDOS"
        ws["D6"].font = Font(bold=True, size=10, color="666666")
        ws["D6"].alignment = Alignment(horizontal='center')
        
        cell_ctl = ws["D7"]
        cell_ctl.font = Font(bold=True, size=12)
        cell_ctl.alignment = Alignment(horizontal='center')
        cell_ctl.border = thin_border

        # 3. Tablas Paralelas
        fila_inicio = 10
        
        # Headers Créditos
        f_header = fila_inicio
        ws.merge_cells(f"A{f_header}:C{f_header}")
        ws[f"A{f_header}"] = "CRÉDITOS" 
        ws[f"A{f_header}"].fill = fill_head_cred
        ws[f"A{f_header}"].font = Font(bold=True, color="FFFFFF")
        ws[f"A{f_header}"].alignment = Alignment(horizontal='center')
        ws[f"A{f_header}"].border = thin_border
        
        headers = ["Fecha", "Descripción", "Importe"]
        cols_cred = ["A", "B", "C"]
        f_sub = f_header + 1
        for i, h in enumerate(headers):
            c = ws[f"{cols_cred[i]}{f_sub}"]
            c.value = h
            c.fill = fill_col_cred
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal='center')
            c.border = thin_border

        # Headers Débitos
        ws.merge_cells(f"E{f_header}:G{f_header}")
        ws[f"E{f_header}"] = "DÉBITOS" 
        ws[f"E{f_header}"].fill = fill_head_deb
        ws[f"E{f_header}"].font = Font(bold=True, color="FFFFFF")
        ws[f"E{f_header}"].alignment = Alignment(horizontal='center')
        ws[f"E{f_header}"].border = thin_border
        
        cols_deb = ["E", "F", "G"]
        for i, h in enumerate(headers):
            c = ws[f"{cols_deb[i]}{f_sub}"]
            c.value = h
            c.fill = fill_col_deb
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal='center')
            c.border = thin_border

        # Llenado de Datos
        fila_dato_start = f_sub + 1
        
        # Créditos
        f_cred = fila_dato_start
        if creditos.empty:
            ws.merge_cells(f"A{f_cred}:C{f_cred}")
            ws[f"A{f_cred}"] = "SIN MOVIMIENTOS"
            ws[f"A{f_cred}"].font = Font(italic=True, color="666666")
            ws[f"A{f_cred}"].alignment = Alignment(horizontal='center')
            ws[f"A{f_cred}"].border = thin_border
            f_cred += 1
        else:
            start_c = f_cred
            for _, r in creditos.iterrows():
                ws[f"A{f_cred}"] = clean_for_excel(r["Fecha"])
                ws[f"A{f_cred}"].fill = fill_row_cred
                ws[f"A{f_cred}"].alignment = Alignment(horizontal='center')
                ws[f"A{f_cred}"].border = thin_border

                ws[f"B{f_cred}"] = clean_for_excel(r["Descripcion"])
                ws[f"B{f_cred}"].fill = fill_row_cred
                ws[f"B{f_cred}"].border = thin_border

                ws[f"C{f_cred}"] = r["Importe"]
                ws[f"C{f_cred}"].number_format = '"$ "#,##0.00'
                ws[f"C{f_cred}"].fill = fill_row_cred
                ws[f"C{f_cred}"].border = thin_border
                f_cred += 1
            
            # Total
            ws.merge_cells(f"A{f_cred}:B{f_cred}")
            ws[f"A{f_cred}"] = "TOTAL CRÉDITOS"
            ws[f"A{f_cred}"].font = Font(bold=True)
            ws[f"A{f_cred}"].alignment = Alignment(horizontal='right')
            ws[f"A{f_cred}"].border = thin_border
            
            ws[f"C{f_cred}"] = f"=SUM(C{start_c}:C{f_cred-1})"
            ws[f"C{f_cred}"].number_format = '"$ "#,##0.00'
            ws[f"C{f_cred}"].font = Font(bold=True)
            ws[f"C{f_cred}"].border = thin_border
            f_cred += 1

        # Débitos
        f_deb = fila_dato_start
        if debitos.empty:
            ws.merge_cells(f"E{f_deb}:G{f_deb}")
            ws[f"E{f_deb}"] = "SIN MOVIMIENTOS"
            ws[f"E{f_deb}"].font = Font(italic=True, color="666666")
            ws[f"E{f_deb}"].alignment = Alignment(horizontal='center')
            ws[f"E{f_deb}"].border = thin_border
            f_deb += 1
        else:
            start_d = f_deb
            for _, r in debitos.iterrows():
                ws[f"E{f_deb}"] = clean_for_excel(r["Fecha"])
                ws[f"E{f_deb}"].fill = fill_row_deb
                ws[f"E{f_deb}"].alignment = Alignment(horizontal='center')
                ws[f"E{f_deb}"].border = thin_border

                ws[f"F{f_deb}"] = clean_for_excel(r["Descripcion"])
                ws[f"F{f_deb}"].fill = fill_row_deb
                ws[f"F{f_deb}"].border = thin_border

                ws[f"G{f_deb}"] = r["Importe"]
                ws[f"G{f_deb}"].number_format = '"$ "#,##0.00'
                ws[f"G{f_deb}"].fill = fill_row_deb
                ws[f"G{f_deb}"].border = thin_border
                f_deb += 1
            
            # Total
            ws.merge_cells(f"E{f_deb}:F{f_deb}")
            ws[f"E{f_deb}"] = "TOTAL DÉBITOS"
            ws[f"E{f_deb}"].font = Font(bold=True)
            ws[f"E{f_deb}"].alignment = Alignment(horizontal='right')
            ws[f"E{f_deb}"].border = thin_border
            
            ws[f"G{f_deb}"] = f"=SUM(G{start_d}:G{f_deb-1})"
            ws[f"G{f_deb}"].number_format = '"$ "#,##0.00'
            ws[f"G{f_deb}"].font = Font(bold=True)
            ws[f"G{f_deb}"].border = thin_border
            f_deb += 1

        # Formula Control
        f_ini = "B3"
        f_tot_cred = f"C{f_cred-1}" if not creditos.empty else "0"
        f_tot_deb = f"G{f_deb-1}" if not debitos.empty else "0"
        f_fin = "B4"
        
        ws["D7"] = f"=ROUND({f_ini}+{f_tot_cred}-{f_tot_deb}-{f_fin}, 2)"
        ws["D7"].number_format = '"$ "#,##0.00'
        
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        red_font = Font(color='9C0006', bold=True)
        ws.conditional_formatting.add('D7', CellIsRule(operator='notEqual', formula=['0'], stopIfTrue=True, fill=red_fill, font=red_font))

        # Anchos
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 25
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 40
        ws.column_dimensions["G"].width = 18

        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    except Exception as e:
        import traceback
        st.error(f"Error al procesar ICBC (Formato 2): {str(e)}")
        print(traceback.format_exc())
        return None
