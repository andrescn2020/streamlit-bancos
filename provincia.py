import streamlit as st
import PyPDF2
import re
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule

# Regex para caracteres ilegales en Excel
ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

def clean_for_excel(text):
    """Elimina caracteres ilegales para Excel y espacios extra"""
    if not text: return ""
    text = str(text)
    text = ILLEGAL_CHARACTERS_RE.sub("", text)
    return text.strip()

def procesar_provincia(archivo_pdf):
    """Procesa archivos PDF del banco Provincia (Formato 1) con Estilo Dashboard"""
    st.info("Procesando archivo del banco Provincia...")

    try:
        # Reinicializar el archivo para lectura
        archivo_pdf.seek(0)

        # Abrir y leer el archivo PDF
        with io.BytesIO(archivo_pdf.read()) as pdf_file:
            reader = PyPDF2.PdfReader(pdf_file)
            texto_completo = "".join(page.extract_text() + "\n" for page in reader.pages)
            lineas = texto_completo.splitlines()

        # 1. Metadatos (Titular, Periodo)
        titular_global = "Sin Especificar"
        periodo_global = "Sin Especificar"
        
        # Titular: "CAJA DE AHORROS EN PESOSSra. ANALIA GISELLE VOUMARD"
        for l in lineas[:15]:
            # Busca "EN PESOS" o "EN DOLARES" y toma lo que sigue
            match_tit = re.search(r"EN (?:PESOS|DOLARES)(.*)$", l, re.IGNORECASE)
            if match_tit:
                titular_global = match_tit.group(1).strip()
                break

        # Delimitar las líneas de movimientos
        inicio = next(
            (i for i, line in enumerate(lineas) if "SALDO ANTERIOR" in line), None
        )
        fin = next(
            (i for i, line in enumerate(lineas) if "Todas las comisiones" in line), None
        )

        if inicio is None or fin is None:
            st.error(
                "No se encontraron las secciones 'SALDO ANTERIOR' o 'Todas las comisiones' en el PDF"
            )
            return None

        movimientos_extraidos = lineas[inicio:fin]

        # Variables para acumular movimientos y procesarlos
        movimientos = []
        # saldo_inicial se capturará del "SALDO ANTERIOR"
        saldo_inicial = 0.0
        saldo_final = 0.0
        
        saldo_anterior = None # Para el loop
        linea_actual = ""
        advertencias = []

        # Patrón que busca los movimientos
        # Fecha ... Descripcion ... Importe ... FechaCorta ... Saldo
        # El Importe se lee literal del PDF (no se deriva restando saldos:
        # eso es fragil si alguna linea no matchea, ya que contamina el
        # calculo de la siguiente). El espacio antes del importe es \s*
        # (no \s+) porque PyPDF2 a veces no deja espacio entre el texto
        # y el signo del importe (ej. "...AUTOMATI-381.29 01-06 41333.70").
        patron_movimiento = re.compile(
            r"^(\d{2}/\d{2}/\d{4})\s+(.*?)\s*([-+]?\d+\.\d{2})\s+(\d{2}-\d{2})\s+([-+]?\d+\.\d{2})$"
        )

        def procesar_bloque(texto, saldo_previo):
            """Interpreta un bloque de texto acumulado como movimiento.
            Devuelve (dict_movimiento_o_None, saldo_resultante).
            Como el Importe y el Saldo se leen de forma independiente,
            se valida que saldo_previo + importe coincida con el saldo
            impreso: si no coincide, es la firma de un movimiento perdido
            o mal fusionado en el medio (ej. boilerplate de salto de
            pagina no filtrado)."""
            m = patron_movimiento.match(texto)
            if not m:
                advertencias.append(f"No se pudo interpretar como movimiento: \"{texto[:100]}\"")
                return None, saldo_previo

            fecha = m.group(1)
            descripcion = m.group(2).strip()
            importe = float(m.group(3))
            saldo_impreso = float(m.group(5))

            saldo_esperado = round(saldo_previo + importe, 2)
            if abs(saldo_esperado - saldo_impreso) > 0.01:
                advertencias.append(
                    f"Salto de saldo inesperado despues de '{fecha} {descripcion[:60]}': "
                    f"esperado $ {saldo_esperado:,.2f}, impreso en el PDF $ {saldo_impreso:,.2f} "
                    "(posible movimiento no reconocido)."
                )

            return {
                "Fecha": fecha,
                "Descripcion": clean_for_excel(descripcion),
                "Importe": importe
            }, saldo_impreso

        for linea in movimientos_extraidos:
            linea_s = linea.strip()
            if not linea_s:
                continue

            if "SALDO ANTERIOR" in linea_s:
                match = re.search(r"SALDO ANTERIOR\s+([-+]?\d+\.\d{2})", linea_s)
                if match:
                    saldo_anterior = float(match.group(1))
                    saldo_inicial = saldo_anterior
                # No es un movimiento a interpretar, solo el checkpoint inicial
                linea_actual = ""
                continue

            # Si empieza con fecha, procesamos la linea acomulada anterior (si existe) o preparamos nueva
            if re.match(r"^\d{2}/\d{2}/\d{4}", linea_s):
                if linea_actual and saldo_anterior is not None:
                    mov, saldo_anterior = procesar_bloque(linea_actual.strip(), saldo_anterior)
                    if mov:
                        movimientos.append(mov)

                linea_actual = linea_s
            else:
                # Puede ser una continuación legítima (wrap de descripción) o
                # boilerplate de salto de página (titular, CUIT, CBU, encabezado
                # de columnas repetido, etc). Si el movimiento acumulado ya es
                # válido y completo, cualquier línea extra es boilerplate y se
                # descarta; si no, es continuación real y se concatena.
                if linea_actual and not patron_movimiento.match(linea_actual):
                    linea_actual += " " + linea_s

        # Procesar el último movimiento remanente en linea_actual
        if linea_actual.strip() and saldo_anterior is not None:
            mov, saldo_anterior = procesar_bloque(linea_actual.strip(), saldo_anterior)
            if mov:
                movimientos.append(mov)

        if saldo_anterior is not None:
            saldo_final = saldo_anterior

        if not movimientos:
            st.warning("No se encontraron movimientos en el PDF")
            return None

        if advertencias:
            st.warning(
                "⚠️ Se detectaron posibles inconsistencias al leer el extracto. "
                "Revisar manualmente antes de usar el reporte:\n\n"
                + "\n".join(f"- {a}" for a in advertencias)
            )

        # --- GENERACIÓN EXCEL (DASHBOARD) ---
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Provincia"
        ws.sheet_view.showGridLines = False
        
        # Paleta Provincia (Verde)
        color_bg_main = "00703C" # Verde Provincia aprox.
        color_txt_main = "FFFFFF"
        
        thin_border = Border(left=Side(style='thin', color="A6A6A6"), 
                             right=Side(style='thin', color="A6A6A6"), 
                             top=Side(style='thin', color="A6A6A6"), 
                             bottom=Side(style='thin', color="A6A6A6"))
                             
        fill_head_deb = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        fill_col_deb = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
        fill_row_deb = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")

        fill_head_cred = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        fill_col_cred = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")
        fill_row_cred = PatternFill(start_color="F2F9F1", end_color="F2F9F1", fill_type="solid")

        df = pd.DataFrame(movimientos)
        if not df.empty:
            creditos = df[df["Importe"] > 0].copy()
            debitos = df[df["Importe"] < 0].copy()
            debitos["Importe"] = debitos["Importe"].abs()
        else:
             creditos = pd.DataFrame(columns=["Fecha", "Descripcion", "Importe"])
             debitos = pd.DataFrame(columns=["Fecha", "Descripcion", "Importe"])

        # 1. Header
        ws.merge_cells("A1:G1")
        tit = ws["A1"]
        tit.value = f"REPORTE PROVINCIA - {clean_for_excel(titular_global)}"
        tit.font = Font(size=14, bold=True, color=color_txt_main)
        tit.fill = PatternFill(start_color=color_bg_main, end_color=color_bg_main, fill_type="solid")
        tit.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25

        # 2. Metadata y Saldos
        ws["A3"] = "SALDO INICIAL"
        ws["A3"].font = Font(bold=True, size=10, color="666666")
        ws["B3"] = saldo_inicial
        ws["B3"].number_format = '"$ "#,##0.00'
        ws["B3"].font = Font(bold=True, size=11)
        ws["B3"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

        ws["A4"] = "SALDO FINAL"
        ws["A4"].font = Font(bold=True, size=10, color="666666")
        ws["B4"] = saldo_final
        ws["B4"].number_format = '"$ "#,##0.00'
        ws["B4"].font = Font(bold=True, size=11)
        ws["B4"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

        ws["D3"] = "TITULAR"
        ws["D3"].alignment = Alignment(horizontal='right')
        ws["D3"].font = Font(bold=True, color="666666", size=10)
        ws["E3"] = clean_for_excel(titular_global)
        ws["E3"].font = Font(bold=True, size=11)
        ws["E3"].alignment = Alignment(horizontal='center')
        ws.merge_cells("E3:G3")
        for c in ["E","F","G"]: ws[f"{c}3"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

        ws["D4"] = "PERÍODO"
        ws["D4"].alignment = Alignment(horizontal='right')
        ws["D4"].font = Font(bold=True, color="666666", size=10)
        ws["E4"] = clean_for_excel(periodo_global)
        ws["E4"].font = Font(bold=True, size=11)
        ws["E4"].alignment = Alignment(horizontal='center')
        ws.merge_cells("E4:G4")
        for c in ["E","F","G"]: ws[f"{c}4"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

        ws["D6"] = "CONTROL DE SALDOS"
        ws["D6"].font = Font(bold=True, size=10, color="666666")
        ws["D6"].alignment = Alignment(horizontal='center')
        
        cell_ctl = ws["D7"]
        cell_ctl.font = Font(bold=True, size=12)
        cell_ctl.alignment = Alignment(horizontal='center')
        cell_ctl.border = thin_border

        # 3. Tablas Paralelas
        fila_inicio = 10
        
        # Headers
        f_header = fila_inicio
        ws.merge_cells(f"A{f_header}:C{f_header}")
        ws[f"A{f_header}"] = "CRÉDITOS" 
        ws[f"A{f_header}"].fill = fill_head_cred
        ws[f"A{f_header}"].font = Font(bold=True, color="FFFFFF")
        ws[f"A{f_header}"].alignment = Alignment(horizontal='center')
        ws[f"A{f_header}"].border = thin_border
        
        headers = ["Fecha", "Descripción", "Importe"]
        cols_cred = ["A", "B", "C"]
        f_sub = f_header + 1
        for i, h in enumerate(headers):
            c = ws[f"{cols_cred[i]}{f_sub}"]
            c.value = h
            c.fill = fill_col_cred
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal='center')
            c.border = thin_border

        ws.merge_cells(f"E{f_header}:G{f_header}")
        ws[f"E{f_header}"] = "DÉBITOS" 
        ws[f"E{f_header}"].fill = fill_head_deb
        ws[f"E{f_header}"].font = Font(bold=True, color="FFFFFF")
        ws[f"E{f_header}"].alignment = Alignment(horizontal='center')
        ws[f"E{f_header}"].border = thin_border
        
        cols_deb = ["E", "F", "G"]
        for i, h in enumerate(headers):
            c = ws[f"{cols_deb[i]}{f_sub}"]
            c.value = h
            c.fill = fill_col_deb
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal='center')
            c.border = thin_border

        # Datos
        fila_dato_start = f_sub + 1
        
        # Créditos
        f_cred = fila_dato_start
        if creditos.empty:
            ws.merge_cells(f"A{f_cred}:C{f_cred}")
            ws[f"A{f_cred}"] = "SIN MOVIMIENTOS"
            ws[f"A{f_cred}"].font = Font(italic=True, color="666666")
            ws[f"A{f_cred}"].alignment = Alignment(horizontal='center')
            ws[f"A{f_cred}"].border = thin_border
            f_cred += 1
        else:
            start_c = f_cred
            for _, r in creditos.iterrows():
                ws[f"A{f_cred}"] = clean_for_excel(r["Fecha"])
                ws[f"A{f_cred}"].fill = fill_row_cred
                ws[f"A{f_cred}"].alignment = Alignment(horizontal='center')
                ws[f"A{f_cred}"].border = thin_border
                ws[f"B{f_cred}"] = clean_for_excel(r["Descripcion"])
                ws[f"B{f_cred}"].fill = fill_row_cred
                ws[f"B{f_cred}"].border = thin_border
                ws[f"C{f_cred}"] = r["Importe"]
                ws[f"C{f_cred}"].number_format = '"$ "#,##0.00'
                ws[f"C{f_cred}"].fill = fill_row_cred
                ws[f"C{f_cred}"].border = thin_border
                f_cred += 1
            ws.merge_cells(f"A{f_cred}:B{f_cred}")
            ws[f"A{f_cred}"] = "TOTAL CRÉDITOS"
            ws[f"A{f_cred}"].font = Font(bold=True)
            ws[f"A{f_cred}"].alignment = Alignment(horizontal='right')
            ws[f"A{f_cred}"].border = thin_border
            ws[f"C{f_cred}"] = f"=SUM(C{start_c}:C{f_cred-1})"
            ws[f"C{f_cred}"].number_format = '"$ "#,##0.00'
            ws[f"C{f_cred}"].font = Font(bold=True)
            ws[f"C{f_cred}"].border = thin_border
            f_cred += 1

        # Débitos
        f_deb = fila_dato_start
        if debitos.empty:
            ws.merge_cells(f"E{f_deb}:G{f_deb}")
            ws[f"E{f_deb}"] = "SIN MOVIMIENTOS"
            ws[f"E{f_deb}"].font = Font(italic=True, color="666666")
            ws[f"E{f_deb}"].alignment = Alignment(horizontal='center')
            ws[f"E{f_deb}"].border = thin_border
            f_deb += 1
        else:
            start_d = f_deb
            for _, r in debitos.iterrows():
                ws[f"E{f_deb}"] = clean_for_excel(r["Fecha"])
                ws[f"E{f_deb}"].fill = fill_row_deb
                ws[f"E{f_deb}"].alignment = Alignment(horizontal='center')
                ws[f"E{f_deb}"].border = thin_border
                ws[f"F{f_deb}"] = clean_for_excel(r["Descripcion"])
                ws[f"F{f_deb}"].fill = fill_row_deb
                ws[f"F{f_deb}"].border = thin_border
                ws[f"G{f_deb}"] = r["Importe"]
                ws[f"G{f_deb}"].number_format = '"$ "#,##0.00'
                ws[f"G{f_deb}"].fill = fill_row_deb
                ws[f"G{f_deb}"].border = thin_border
                f_deb += 1
            ws.merge_cells(f"E{f_deb}:F{f_deb}")
            ws[f"E{f_deb}"] = "TOTAL DÉBITOS"
            ws[f"E{f_deb}"].font = Font(bold=True)
            ws[f"E{f_deb}"].alignment = Alignment(horizontal='right')
            ws[f"E{f_deb}"].border = thin_border
            ws[f"G{f_deb}"] = f"=SUM(G{start_d}:G{f_deb-1})"
            ws[f"G{f_deb}"].number_format = '"$ "#,##0.00'
            ws[f"G{f_deb}"].font = Font(bold=True)
            ws[f"G{f_deb}"].border = thin_border
            f_deb += 1

        # Formula
        f_ini = "B3"
        f_tot_cred = f"C{f_cred-1}" if not creditos.empty else "0"
        f_tot_deb = f"G{f_deb-1}" if not debitos.empty else "0"
        f_fin = "B4"
        ws["D7"] = f"=ROUND({f_ini}+{f_tot_cred}-{f_tot_deb}-{f_fin}, 2)"
        ws["D7"].number_format = '"$ "#,##0.00'
        
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        red_font = Font(color='9C0006', bold=True)
        ws.conditional_formatting.add('D7', CellIsRule(operator='notEqual', formula=['0'], stopIfTrue=True, fill=red_fill, font=red_font))

        # Anchos
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 25
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 40
        ws.column_dimensions["G"].width = 18

        wb.save(output)
        output.seek(0)
        return output.getvalue()

    except Exception as e:
        import traceback
        st.error(f"Error al procesar el archivo: {str(e)}")
        print(traceback.format_exc())
        return None
