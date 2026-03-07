import streamlit as st
import pdfplumber
import re
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule

# Regex para limpiar caracteres ilegales de Excel
ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

def clean_for_excel(text):
    """Elimina caracteres ilegales para Excel y espacios extra"""
    if not text: return ""
    text = str(text)
    text = ILLEGAL_CHARACTERS_RE.sub("", text)
    return text.strip()

def parse_numero_ar(s):
    """Parsea número en formato argentino (1.234,56) a float."""
    if not s: return 0.0
    s = s.strip()
    neg = False
    if s.endswith("-"):
        neg = True
        s = s[:-1]
    elif s.startswith("-"):
        neg = True
        s = s[1:]
    s = s.replace(".", "").replace(",", ".")
    try:
        val = float(s)
        return -val if neg else val
    except:
        return 0.0


def procesar_patagonia(archivo_pdf):
    """Procesa extractos de Banco Patagonia"""
    st.info("Procesando archivo del Banco Patagonia...")
    try:
        archivo_pdf.seek(0)
        with pdfplumber.open(io.BytesIO(archivo_pdf.read())) as pdf:
            texto_completo = ""
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    texto_completo += t + "\n"

        lineas = texto_completo.splitlines()

        # ============================================================
        # 1. METADATOS
        # ============================================================
        titular = "Sin Especificar"
        cuenta = "Sin Especificar"
        periodo = "Sin Especificar"

        for line in lineas:
            line_s = line.strip()
            # Cuenta: "Cuenta: CC$ 106-106018243-000"
            m_cuenta = re.match(r'Cuenta:\s*(.+)', line_s)
            if m_cuenta:
                cuenta = m_cuenta.group(1).strip()

            # Titularidad: "Titularidad: GMI TRASLADOS SA"
            m_titular = re.match(r'Titularidad:\s*(.+)', line_s)
            if m_titular:
                titular = m_titular.group(1).strip()

        # ============================================================
        # 2. PARSEO DE MOVIMIENTOS
        # ============================================================
        # Patrón: línea que empieza con fecha DD/MM/YYYY, tiene descripción,
        # y termina con dos montos (importe y saldo) en formato argentino
        patron_mov = re.compile(
            r'^(\d{2}/\d{2}/\d{4})\s+'   # Fecha
            r'(.+?)\s+'                    # Descripción + referencia
            r'([\d.]+,\d{2})\s+'           # Importe (débito o crédito)
            r'([\d.]+,\d{2})$'             # Saldo
        )

        # Líneas que NO son movimientos y NO son texto extra (a ignorar)
        patron_skip = re.compile(
            r'^(Página\s+\d+|---\s*FIN|Movimientos\s+de\s+Cuenta|'
            r'Cuenta:|Titularidad:|Fecha\s+Descripción)'
        )

        movimientos_raw = []
        pending_prefix = []

        def es_movimiento_con_extra(desc):
            """Determina si un movimiento puede tener líneas extra (sufijo como nombre de empresa)"""
            desc_upper = desc.upper()
            return any(kw in desc_upper for kw in [
                "TRANSF. A TERCEROS", "TRANSFERENCIA E-BANK",
                "TRANSFERENCIA D/C", "DEBITO AUTOMATICO",
                "CREDITO POR TRANSFERENCIA"
            ])

        expecting_suffix = False

        for line in lineas:
            line_s = line.strip()
            if not line_s:
                continue

            # Saltar líneas de encabezado/pie
            if patron_skip.match(line_s):
                continue

            match = patron_mov.match(line_s)
            if match:
                fecha = match.group(1)
                descripcion = match.group(2).strip()
                importe_str = match.group(3)
                saldo_str = match.group(4)

                # Incorporar prefijo pendiente (CUIT, nombre parcial, etc.)
                if pending_prefix:
                    descripcion = descripcion + " " + " ".join(pending_prefix)
                    pending_prefix = []

                movimientos_raw.append({
                    "fecha": fecha,
                    "descripcion": descripcion,
                    "importe_str": importe_str,
                    "saldo": parse_numero_ar(saldo_str),
                })

                # Determinar si el siguiente non-date line es sufijo de este movimiento
                expecting_suffix = es_movimiento_con_extra(descripcion)
            else:
                # Línea extra (CUIT, nombre empresa, etc.)
                if expecting_suffix and movimientos_raw:
                    # Es sufijo del último movimiento (nombre de empresa, "SRL", etc.)
                    movimientos_raw[-1]["descripcion"] += f" {line_s}"
                    expecting_suffix = False
                else:
                    # Es prefijo del próximo movimiento
                    pending_prefix.append(line_s)
                    expecting_suffix = False

        # Si quedan prefijos pendientes, agregarlos al último movimiento
        if pending_prefix and movimientos_raw:
            movimientos_raw[-1]["descripcion"] += " " + " ".join(pending_prefix)

        # ============================================================
        # 3. DETERMINAR DÉBITO/CRÉDITO POR DIFERENCIA DE SALDOS
        # ============================================================
        # Los movimientos están en orden DESCENDENTE (más reciente primero)
        # Invertimos para procesar cronológicamente
        movimientos_raw.reverse()

        transactions = []
        for i, mov in enumerate(movimientos_raw):
            if i == 0:
                # Para el primer movimiento (más antiguo), no tenemos saldo anterior
                # Usamos el importe_str directamente; determinamos signo con heurística
                importe_val = parse_numero_ar(mov["importe_str"])
                # Si hay un segundo movimiento, deducimos el saldo anterior
                if len(movimientos_raw) > 1:
                    saldo_anterior = mov["saldo"] - importe_val  # asumir crédito
                    # Verificar con el siguiente: si saldo_anterior es consistente
                    next_saldo = movimientos_raw[1]["saldo"]
                    diff_next = next_saldo - mov["saldo"]
                    next_importe = parse_numero_ar(movimientos_raw[1]["importe_str"])
                    # Si la diferencia del siguiente coincide con su importe (positivo o negativo)
                    # entonces nuestro saldo_anterior está bien
                    if abs(abs(diff_next) - next_importe) > 0.01:
                        # Probablemente el primer mov es débito
                        importe_val = -importe_val
                else:
                    # Solo un movimiento, usar descripción como heurística
                    desc_upper = mov["descripcion"].upper()
                    if any(kw in desc_upper for kw in ["TRANSF. A TERCEROS", "TRANSF. TERCEROS O/BCO",
                                                        "IMP.DB/CR", "DEBITO", "COMISION", "IIBB"]):
                        importe_val = -importe_val
            else:
                saldo_anterior = movimientos_raw[i - 1]["saldo"]
                importe_val = round(mov["saldo"] - saldo_anterior, 2)

            transactions.append({
                "Fecha": mov["fecha"],
                "Descripcion": clean_for_excel(mov["descripcion"]),
                "Importe": round(importe_val, 2)
            })

        if not transactions:
            st.warning("No se encontraron movimientos en el PDF.")
            return None

        st.success(f"Se encontraron {len(transactions)} movimientos.")

        # Saldos: el más antiguo (primer movimiento invertido) y el más reciente (último)
        saldo_final = movimientos_raw[-1]["saldo"]
        # Saldo inicial = saldo del primer mov - importe del primer mov
        saldo_inicial = round(movimientos_raw[0]["saldo"] - transactions[0]["Importe"], 2)

        # Deducir periodo de las fechas
        fechas_unicas = sorted(set(mov["fecha"] for mov in movimientos_raw),
                               key=lambda x: (int(x[6:10]), int(x[3:5]), int(x[0:2])))
        if fechas_unicas:
            periodo = f"Del {fechas_unicas[0]} al {fechas_unicas[-1]}"

        # ============================================================
        # 4. GENERAR EXCEL (mismo estilo que otros bancos)
        # ============================================================
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        nombre_hoja = clean_for_excel(cuenta)[:31]
        ws.title = nombre_hoja
        ws.sheet_view.showGridLines = False

        # Paleta Banco Patagonia (verde oscuro)
        color_bg_main = "006341"
        color_txt_main = "FFFFFF"

        thin_border = Border(
            left=Side(style='thin', color="A6A6A6"),
            right=Side(style='thin', color="A6A6A6"),
            top=Side(style='thin', color="A6A6A6"),
            bottom=Side(style='thin', color="A6A6A6")
        )

        fill_head_deb = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        fill_col_deb = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
        fill_row_deb = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")

        fill_head_cred = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        fill_col_cred = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")
        fill_row_cred = PatternFill(start_color="F2F9F1", end_color="F2F9F1", fill_type="solid")

        df = pd.DataFrame(transactions)
        creditos = df[df["Importe"] > 0].copy()
        debitos = df[df["Importe"] < 0].copy()
        if not debitos.empty:
            debitos["Importe"] = debitos["Importe"].abs()

        # --- Header ---
        ws.merge_cells("A1:G1")
        tit = ws["A1"]
        tit.value = f"REPORTE BANCO PATAGONIA - {clean_for_excel(cuenta)}"
        tit.font = Font(size=14, bold=True, color=color_txt_main)
        tit.fill = PatternFill(start_color=color_bg_main, end_color=color_bg_main, fill_type="solid")
        tit.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25

        # --- Metadata ---
        ws["A3"] = "SALDO INICIAL"
        ws["A3"].font = Font(bold=True, size=10, color="666666")
        ws["B3"] = saldo_inicial
        ws["B3"].number_format = '"$ "#,##0.00'
        ws["B3"].font = Font(bold=True, size=11)
        ws["B3"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

        ws["A4"] = "SALDO FINAL"
        ws["A4"].font = Font(bold=True, size=10, color="666666")
        ws["B4"] = saldo_final
        ws["B4"].number_format = '"$ "#,##0.00'
        ws["B4"].font = Font(bold=True, size=11)
        ws["B4"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

        ws["D3"] = "TITULAR"
        ws["D3"].alignment = Alignment(horizontal='right')
        ws["D3"].font = Font(bold=True, color="666666", size=10)
        ws["E3"] = clean_for_excel(titular)
        ws["E3"].font = Font(bold=True, size=11)
        ws["E3"].alignment = Alignment(horizontal='center')
        ws.merge_cells("E3:G3")
        for c in ["E","F","G"]: ws[f"{c}3"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

        ws["D4"] = "PERÍODO"
        ws["D4"].alignment = Alignment(horizontal='right')
        ws["D4"].font = Font(bold=True, color="666666", size=10)
        ws["E4"] = clean_for_excel(periodo)
        ws["E4"].font = Font(bold=True, size=11)
        ws["E4"].alignment = Alignment(horizontal='center')
        ws.merge_cells("E4:G4")
        for c in ["E","F","G"]: ws[f"{c}4"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

        # --- Tablas Créditos / Débitos ---
        fila_inicio = 10
        f_header = fila_inicio

        ws.merge_cells(f"A{f_header}:C{f_header}")
        ws[f"A{f_header}"] = "CRÉDITOS"
        ws[f"A{f_header}"].fill = fill_head_cred
        ws[f"A{f_header}"].font = Font(bold=True, color="FFFFFF")
        ws[f"A{f_header}"].alignment = Alignment(horizontal='center')
        ws[f"A{f_header}"].border = thin_border

        ws.merge_cells(f"E{f_header}:G{f_header}")
        ws[f"E{f_header}"] = "DÉBITOS"
        ws[f"E{f_header}"].fill = fill_head_deb
        ws[f"E{f_header}"].font = Font(bold=True, color="FFFFFF")
        ws[f"E{f_header}"].alignment = Alignment(horizontal='center')
        ws[f"E{f_header}"].border = thin_border

        headers = ["Fecha", "Descripción", "Importe"]
        cols_cred = ["A", "B", "C"]
        cols_deb = ["E", "F", "G"]
        f_sub = f_header + 1

        for i, h in enumerate(headers):
            c = ws[f"{cols_cred[i]}{f_sub}"]
            c.value = h
            c.fill = fill_col_cred
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal='center')
            c.border = thin_border

            d = ws[f"{cols_deb[i]}{f_sub}"]
            d.value = h
            d.fill = fill_col_deb
            d.font = Font(bold=True)
            d.alignment = Alignment(horizontal='center')
            d.border = thin_border

        fila_a_llenar = f_sub + 1

        # Créditos
        f_c = fila_a_llenar
        if creditos.empty:
            ws.merge_cells(f"A{f_c}:C{f_c}")
            ws[f"A{f_c}"] = "SIN MOVIMIENTOS"
            ws[f"A{f_c}"].border = thin_border
            f_c += 1
        else:
            start_c = f_c
            for _, r in creditos.iterrows():
                ws[f"A{f_c}"] = r["Fecha"]
                ws[f"A{f_c}"].fill = fill_row_cred
                ws[f"A{f_c}"].border = thin_border
                ws[f"A{f_c}"].alignment = Alignment(horizontal='center')
                ws[f"B{f_c}"] = r["Descripcion"]
                ws[f"B{f_c}"].fill = fill_row_cred
                ws[f"B{f_c}"].border = thin_border
                ws[f"C{f_c}"] = r["Importe"]
                ws[f"C{f_c}"].number_format = '"$ "#,##0.00'
                ws[f"C{f_c}"].fill = fill_row_cred
                ws[f"C{f_c}"].border = thin_border
                f_c += 1
            ws.merge_cells(f"A{f_c}:B{f_c}")
            ws[f"A{f_c}"] = "TOTAL CRÉDITOS"
            ws[f"A{f_c}"].font = Font(bold=True)
            ws[f"A{f_c}"].alignment = Alignment(horizontal='right')
            ws[f"C{f_c}"] = f"=SUM(C{start_c}:C{f_c-1})"
            ws[f"C{f_c}"].font = Font(bold=True)
            ws[f"C{f_c}"].number_format = '"$ "#,##0.00'
            f_c += 1

        # Débitos
        f_d = fila_a_llenar
        if debitos.empty:
            ws.merge_cells(f"E{f_d}:G{f_d}")
            ws[f"E{f_d}"] = "SIN MOVIMIENTOS"
            ws[f"E{f_d}"].border = thin_border
            f_d += 1
        else:
            start_d = f_d
            for _, r in debitos.iterrows():
                ws[f"E{f_d}"] = r["Fecha"]
                ws[f"E{f_d}"].fill = fill_row_deb
                ws[f"E{f_d}"].border = thin_border
                ws[f"E{f_d}"].alignment = Alignment(horizontal='center')
                ws[f"F{f_d}"] = r["Descripcion"]
                ws[f"F{f_d}"].fill = fill_row_deb
                ws[f"F{f_d}"].border = thin_border
                ws[f"G{f_d}"] = r["Importe"]
                ws[f"G{f_d}"].number_format = '"$ "#,##0.00'
                ws[f"G{f_d}"].fill = fill_row_deb
                ws[f"G{f_d}"].border = thin_border
                f_d += 1
            ws.merge_cells(f"E{f_d}:F{f_d}")
            ws[f"E{f_d}"] = "TOTAL DÉBITOS"
            ws[f"E{f_d}"].font = Font(bold=True)
            ws[f"E{f_d}"].alignment = Alignment(horizontal='right')
            ws[f"G{f_d}"] = f"=SUM(G{start_d}:G{f_d-1})"
            ws[f"G{f_d}"].font = Font(bold=True)
            ws[f"G{f_d}"].number_format = '"$ "#,##0.00'
            f_d += 1

        # --- Control de Saldos ---
        ws["D6"] = "CONTROL DE SALDOS"
        ws["D6"].font = Font(bold=True, size=10, color="666666")
        ws["D6"].alignment = Alignment(horizontal='center')

        ref_tot_c = f"C{f_c-1}" if not creditos.empty else "0"
        ref_tot_d = f"G{f_d-1}" if not debitos.empty else "0"
        ws["D7"] = f"=ROUND(B3+{ref_tot_c}-{ref_tot_d}-B4, 2)"
        ws["D7"].number_format = '"$ "#,##0.00'
        ws["D7"].font = Font(bold=True)
        ws["D7"].alignment = Alignment(horizontal='center')
        ws["D7"].border = thin_border

        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        red_font = Font(color='9C0006', bold=True)
        ws.conditional_formatting.add('D7', CellIsRule(operator='notEqual', formula=['0'], stopIfTrue=True, fill=red_fill, font=red_font))

        # --- Anchos de columnas ---
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 45
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 25
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 45
        ws.column_dimensions["G"].width = 18

        wb.save(output)
        output.seek(0)
        return output.getvalue()

    except Exception as e:
        import traceback
        st.error(f"Error al procesar: {e}")
        print(traceback.format_exc())
        return None
