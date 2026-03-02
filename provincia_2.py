import streamlit as st
import PyPDF2
import re
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule

ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

def clean_for_excel(text):
    """Elimina caracteres ilegales para Excel y espacios extra"""
    if not text: return ""
    text = str(text)
    text = ILLEGAL_CHARACTERS_RE.sub("", text)
    return text.strip()

def parse_numero_ar(texto):
    """Convierte número con formato argentino (1.234.567,89) a float"""
    texto = texto.strip().replace(".", "").replace(",", ".")
    return float(texto)

def extraer_importe_del_final(texto):
    """Extrae un número AR del final del texto, eligiendo el grupo inicial MÁS PEQUEÑO.
    Esto evita incluir dígitos de números de referencia concatenados.
    Ej: '...9651.663,74' → 1.663,74 (no 651.663,74)"""
    # Buscar la parte de miles+decimales al final: (.ddd)* ,dd
    m = re.search(r'((?:\.\d{3})*,\d{2})$', texto)
    if not m:
        return None
    suffix = m.group(1)  # ej: ".663,74" o ",74"
    prefix_end = m.start()

    # Intentar 1, 2, luego 3 dígitos para el primer grupo (el más chico primero)
    for n_digits in (1, 2, 3):
        start = prefix_end - n_digits
        if start < 0:
            continue
        first_group = texto[start:prefix_end]
        if re.match(r'^\d{' + str(n_digits) + r'}$', first_group):
            # Ver si hay signo negativo antes
            if start > 0 and texto[start - 1] == '-':
                return parse_numero_ar('-' + first_group + suffix)
            return parse_numero_ar(first_group + suffix)
    return None

def procesar_provincia_formato_2(archivo_pdf):
    """Procesa archivos PDF del banco Provincia (Formato 2) con Estilo Dashboard"""
    st.info("Procesando archivo del banco Provincia (Formato 2)...")

    try:
        archivo_pdf.seek(0)
        with io.BytesIO(archivo_pdf.read()) as pdf_file:
            reader = PyPDF2.PdfReader(pdf_file)
            texto_completo = "".join(page.extract_text() + "\n" for page in reader.pages)
            lineas = texto_completo.splitlines()

        # Extraer cuenta
        cuenta = "Sin Especificar"
        for l in lineas[:10]:
            match_cuenta = re.search(r"Cuenta:\s*(.+)", l)
            if match_cuenta:
                cuenta = match_cuenta.group(1).strip()
                break

        # Filtrar headers/footers
        skip_patterns = [
            re.compile(r"^Fecha:\d{2}/\d{2}/\d{4}"),
            re.compile(r"^Detalle de Movimientos"),
            re.compile(r"^Fecha\s+Descripci"),
            re.compile(r"^\d*\s*Esta informaci"),
            re.compile(r"^y est. supeditada"),
            re.compile(r"^\d*\s*P.gina \d+"),
            re.compile(r"^\d+$"),
        ]

        lineas_mov = []
        for linea in lineas:
            stripped = linea.strip()
            if not stripped:
                continue
            if any(pat.match(stripped) for pat in skip_patterns):
                continue
            lineas_mov.append(stripped)

        # Agrupar líneas por movimiento (cada uno empieza con fecha dd-mmm-yyyy)
        date_start = re.compile(r"^\d{2}-\w{3}-\d{4}")
        movimientos_raw = []
        current = ""

        for linea in lineas_mov:
            if date_start.match(linea):
                if current:
                    movimientos_raw.append(current)
                current = linea
            else:
                if current:
                    current += " " + linea
        if current:
            movimientos_raw.append(current)

        # Parsear cada movimiento: extraer fecha y saldo (siempre separado por espacio).
        # El importe se calcula después desde diferencias de saldos.
        saldo_re = re.compile(r"\s+(-?\d{1,3}(?:\.\d{3})*,\d{2})$")

        movimientos = []
        for raw in movimientos_raw:
            date_match = re.match(r"^(\d{2}-\w{3}-\d{4})\s+(.*)$", raw, re.DOTALL)
            if not date_match:
                continue
            fecha = date_match.group(1)
            rest = date_match.group(2)

            # Extraer saldo (último número, siempre separado por espacio)
            saldo_match = saldo_re.search(rest)
            if not saldo_match:
                continue
            saldo = parse_numero_ar(saldo_match.group(1))
            rest_sin_saldo = rest[:saldo_match.start()]

            # Extraer descripción: sacar el importe del final del texto restante
            importe_match = saldo_re.search(rest_sin_saldo)
            if importe_match:
                # Importe separado por espacio → descripción limpia
                descripcion = rest_sin_saldo[:importe_match.start()].strip()
            else:
                # Importe pegado a texto (ej: referencia concatenada) → limpiar
                desc_clean = re.sub(r"-?\d{1,3}(?:\.\d{3})*,\d{2}$", "", rest_sin_saldo)
                descripcion = desc_clean.strip()

            movimientos.append({
                "Fecha": fecha,
                "Descripcion": clean_for_excel(descripcion),
                "Saldo": saldo
            })

        # Mergear duplicados por corte de página (mismo saldo + fecha consecutivos)
        merged = []
        for mov in movimientos:
            if (merged
                and merged[-1]["Saldo"] == mov["Saldo"]
                and merged[-1]["Fecha"] == mov["Fecha"]):
                merged[-1]["Descripcion"] += " " + mov["Descripcion"]
            else:
                merged.append(mov)
        movimientos = merged

        if not movimientos:
            st.warning("No se encontraron movimientos en el PDF")
            return None

        # Calcular importes desde diferencias de saldos (orden inverso: más reciente primero)
        for i in range(len(movimientos) - 1):
            movimientos[i]["Importe"] = round(movimientos[i]["Saldo"] - movimientos[i + 1]["Saldo"], 2)

        # Para el último movimiento: extraer importe con heurística de grupo mínimo
        last_idx = len(movimientos) - 1
        last_raw = movimientos_raw[-1] if movimientos_raw else ""
        last_saldo_match = saldo_re.search(last_raw)
        if last_saldo_match:
            texto_sin_saldo = last_raw[:last_saldo_match.start()].strip()
            importe_last = extraer_importe_del_final(texto_sin_saldo)
            movimientos[last_idx]["Importe"] = importe_last if importe_last is not None else 0.0
        else:
            movimientos[last_idx]["Importe"] = 0.0

        # Saldos
        saldo_final = movimientos[0]["Saldo"]
        saldo_inicial = round(movimientos[-1]["Saldo"] - movimientos[-1]["Importe"], 2)

        # --- GENERACIÓN EXCEL (DASHBOARD) ---
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Provincia F2"
        ws.sheet_view.showGridLines = False

        color_bg_main = "00703C"
        color_txt_main = "FFFFFF"

        thin_border = Border(
            left=Side(style='thin', color="A6A6A6"),
            right=Side(style='thin', color="A6A6A6"),
            top=Side(style='thin', color="A6A6A6"),
            bottom=Side(style='thin', color="A6A6A6")
        )

        fill_head_deb = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        fill_col_deb = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
        fill_row_deb = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")

        fill_head_cred = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        fill_col_cred = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")
        fill_row_cred = PatternFill(start_color="F2F9F1", end_color="F2F9F1", fill_type="solid")

        df = pd.DataFrame(movimientos)
        if not df.empty:
            creditos = df[df["Importe"] > 0].copy()
            debitos = df[df["Importe"] < 0].copy()
            debitos["Importe"] = debitos["Importe"].abs()
        else:
            creditos = pd.DataFrame(columns=["Fecha", "Descripcion", "Importe"])
            debitos = pd.DataFrame(columns=["Fecha", "Descripcion", "Importe"])

        # 1. Header
        ws.merge_cells("A1:G1")
        tit = ws["A1"]
        tit.value = f"REPORTE PROVINCIA (F2) - Cuenta: {clean_for_excel(cuenta)}"
        tit.font = Font(size=14, bold=True, color=color_txt_main)
        tit.fill = PatternFill(start_color=color_bg_main, end_color=color_bg_main, fill_type="solid")
        tit.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25

        # 2. Metadata y Saldos
        ws["A3"] = "SALDO INICIAL"
        ws["A3"].font = Font(bold=True, size=10, color="666666")
        ws["B3"] = saldo_inicial
        ws["B3"].number_format = '"$ "#,##0.00'
        ws["B3"].font = Font(bold=True, size=11)
        ws["B3"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

        ws["A4"] = "SALDO FINAL"
        ws["A4"].font = Font(bold=True, size=10, color="666666")
        ws["B4"] = saldo_final
        ws["B4"].number_format = '"$ "#,##0.00'
        ws["B4"].font = Font(bold=True, size=11)
        ws["B4"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

        ws["D3"] = "CUENTA"
        ws["D3"].alignment = Alignment(horizontal='right')
        ws["D3"].font = Font(bold=True, color="666666", size=10)
        ws["E3"] = clean_for_excel(cuenta)
        ws["E3"].font = Font(bold=True, size=11)
        ws["E3"].alignment = Alignment(horizontal='center')
        ws.merge_cells("E3:G3")
        for c in ["E","F","G"]: ws[f"{c}3"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

        ws["D6"] = "CONTROL DE SALDOS"
        ws["D6"].font = Font(bold=True, size=10, color="666666")
        ws["D6"].alignment = Alignment(horizontal='center')

        cell_ctl = ws["D7"]
        cell_ctl.font = Font(bold=True, size=12)
        cell_ctl.alignment = Alignment(horizontal='center')
        cell_ctl.border = thin_border

        # 3. Tablas Paralelas
        fila_inicio = 10
        f_header = fila_inicio

        ws.merge_cells(f"A{f_header}:C{f_header}")
        ws[f"A{f_header}"] = "CRÉDITOS"
        ws[f"A{f_header}"].fill = fill_head_cred
        ws[f"A{f_header}"].font = Font(bold=True, color="FFFFFF")
        ws[f"A{f_header}"].alignment = Alignment(horizontal='center')
        ws[f"A{f_header}"].border = thin_border

        headers = ["Fecha", "Descripción", "Importe"]
        cols_cred = ["A", "B", "C"]
        f_sub = f_header + 1
        for i, h in enumerate(headers):
            c = ws[f"{cols_cred[i]}{f_sub}"]
            c.value = h
            c.fill = fill_col_cred
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal='center')
            c.border = thin_border

        ws.merge_cells(f"E{f_header}:G{f_header}")
        ws[f"E{f_header}"] = "DÉBITOS"
        ws[f"E{f_header}"].fill = fill_head_deb
        ws[f"E{f_header}"].font = Font(bold=True, color="FFFFFF")
        ws[f"E{f_header}"].alignment = Alignment(horizontal='center')
        ws[f"E{f_header}"].border = thin_border

        cols_deb = ["E", "F", "G"]
        for i, h in enumerate(headers):
            c = ws[f"{cols_deb[i]}{f_sub}"]
            c.value = h
            c.fill = fill_col_deb
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal='center')
            c.border = thin_border

        fila_dato_start = f_sub + 1

        # Créditos
        f_cred = fila_dato_start
        if creditos.empty:
            ws.merge_cells(f"A{f_cred}:C{f_cred}")
            ws[f"A{f_cred}"] = "SIN MOVIMIENTOS"
            ws[f"A{f_cred}"].font = Font(italic=True, color="666666")
            ws[f"A{f_cred}"].alignment = Alignment(horizontal='center')
            ws[f"A{f_cred}"].border = thin_border
            f_cred += 1
        else:
            start_c = f_cred
            for _, r in creditos.iterrows():
                ws[f"A{f_cred}"] = clean_for_excel(r["Fecha"])
                ws[f"A{f_cred}"].fill = fill_row_cred
                ws[f"A{f_cred}"].alignment = Alignment(horizontal='center')
                ws[f"A{f_cred}"].border = thin_border
                ws[f"B{f_cred}"] = clean_for_excel(r["Descripcion"])
                ws[f"B{f_cred}"].fill = fill_row_cred
                ws[f"B{f_cred}"].border = thin_border
                ws[f"C{f_cred}"] = r["Importe"]
                ws[f"C{f_cred}"].number_format = '"$ "#,##0.00'
                ws[f"C{f_cred}"].fill = fill_row_cred
                ws[f"C{f_cred}"].border = thin_border
                f_cred += 1
            ws.merge_cells(f"A{f_cred}:B{f_cred}")
            ws[f"A{f_cred}"] = "TOTAL CRÉDITOS"
            ws[f"A{f_cred}"].font = Font(bold=True)
            ws[f"A{f_cred}"].alignment = Alignment(horizontal='right')
            ws[f"A{f_cred}"].border = thin_border
            ws[f"C{f_cred}"] = f"=SUM(C{start_c}:C{f_cred-1})"
            ws[f"C{f_cred}"].number_format = '"$ "#,##0.00'
            ws[f"C{f_cred}"].font = Font(bold=True)
            ws[f"C{f_cred}"].border = thin_border
            f_cred += 1

        # Débitos
        f_deb = fila_dato_start
        if debitos.empty:
            ws.merge_cells(f"E{f_deb}:G{f_deb}")
            ws[f"E{f_deb}"] = "SIN MOVIMIENTOS"
            ws[f"E{f_deb}"].font = Font(italic=True, color="666666")
            ws[f"E{f_deb}"].alignment = Alignment(horizontal='center')
            ws[f"E{f_deb}"].border = thin_border
            f_deb += 1
        else:
            start_d = f_deb
            for _, r in debitos.iterrows():
                ws[f"E{f_deb}"] = clean_for_excel(r["Fecha"])
                ws[f"E{f_deb}"].fill = fill_row_deb
                ws[f"E{f_deb}"].alignment = Alignment(horizontal='center')
                ws[f"E{f_deb}"].border = thin_border
                ws[f"F{f_deb}"] = clean_for_excel(r["Descripcion"])
                ws[f"F{f_deb}"].fill = fill_row_deb
                ws[f"F{f_deb}"].border = thin_border
                ws[f"G{f_deb}"] = r["Importe"]
                ws[f"G{f_deb}"].number_format = '"$ "#,##0.00'
                ws[f"G{f_deb}"].fill = fill_row_deb
                ws[f"G{f_deb}"].border = thin_border
                f_deb += 1
            ws.merge_cells(f"E{f_deb}:F{f_deb}")
            ws[f"E{f_deb}"] = "TOTAL DÉBITOS"
            ws[f"E{f_deb}"].font = Font(bold=True)
            ws[f"E{f_deb}"].alignment = Alignment(horizontal='right')
            ws[f"E{f_deb}"].border = thin_border
            ws[f"G{f_deb}"] = f"=SUM(G{start_d}:G{f_deb-1})"
            ws[f"G{f_deb}"].number_format = '"$ "#,##0.00'
            ws[f"G{f_deb}"].font = Font(bold=True)
            ws[f"G{f_deb}"].border = thin_border
            f_deb += 1

        # Control de saldos formula
        f_tot_cred = f"C{f_cred-1}" if not creditos.empty else "0"
        f_tot_deb = f"G{f_deb-1}" if not debitos.empty else "0"
        ws["D7"] = f"=ROUND(B3+{f_tot_cred}-{f_tot_deb}-B4, 2)"
        ws["D7"].number_format = '"$ "#,##0.00'

        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        red_font = Font(color='9C0006', bold=True)
        ws.conditional_formatting.add('D7', CellIsRule(operator='notEqual', formula=['0'], stopIfTrue=True, fill=red_fill, font=red_font))

        # Anchos
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 45
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 25
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 45
        ws.column_dimensions["G"].width = 18

        wb.save(output)
        output.seek(0)
        return output.getvalue()

    except Exception as e:
        import traceback
        st.error(f"Error al procesar el archivo: {str(e)}")
        print(traceback.format_exc())
        return None
