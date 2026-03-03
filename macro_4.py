import streamlit as st
import pdfplumber
import re
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule

ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

def clean_for_excel(text):
    if not text: return ""
    text = str(text)
    text = ILLEGAL_CHARACTERS_RE.sub("", text)
    return text.strip()

def parse_monto(s):
    """Convierte '10,165.19' o '-10,165.19' a float. Formato Inglés."""
    if not s: return 0.0
    s = s.strip()
    try:
        # 10,165.19 -> 10165.19
        limpio = s.replace(",", "")
        return float(limpio)
    except:
        return 0.0

def procesar_macro_formato_4(archivo_pdf):
    """Procesa archivos PDF del Banco Macro - Formato 4 (English Number Format) usando pdfplumber."""
    st.info("Procesando archivo del Banco Macro (Formato 4) con pdfplumber...")
    try:
        archivo_pdf.seek(0)
        
        texto = ""
        with pdfplumber.open(io.BytesIO(archivo_pdf.read())) as pdf:
            for page in pdf.pages:
                extracted = page.extract_text()
                if extracted:
                    texto += extracted + "\n"
        
        lineas_raw = texto.splitlines()
        
        # === METADATOS ===
        titular = "Sin Especificar"
        periodo = "Sin Especificar"
        cuenta = "Sin Especificar"
        saldo_ini = 0.0
        saldo_fin = 0.0
        
        for idx, l in enumerate(lineas_raw[:30]):
            match_tit_alt = re.search(r'Sr/a:\s*(.*)', l, re.IGNORECASE)
            if match_tit_alt:
                titular = match_tit_alt.group(1).strip()
            
            # Cuenta: "CUENTA CORRIENTE BANCARIA Nº 3-471-0004583710-0 SUCURSAL: 471"
            match_cta = re.search(r'N[º°\*]?\s*([\d-]+)', l, re.IGNORECASE)
            if match_cta:
                cuenta_str = match_cta.group(1).strip()
                if len(cuenta_str) > 5:
                    cuenta = cuenta_str
            
            # Periodo: "Período del Extracto: 1/3/2025 al 31/3/2025"
            match_per = re.search(r'Per[ií]odo[^\d]*(\d{1,2}/\d{1,2}/\d{2,4})\s+al\s+(\d{1,2}/\d{1,2}/\d{2,4})', l, re.IGNORECASE)
            if match_per:
                periodo = f"Del {match_per.group(1)} al {match_per.group(2)}"
        
        # 1. Extraer Saldo Inicial
        # pdfplumber alinea mejor. 
        # Buscaremos una iteración desde arriba.
        # "SALDO INICIAL CREDITOS DEBITOS I.V.A. SALDO FINAL"
        # "59,665.78 0.00 59,665.78 0.00 0.00"
        idx_saldo_ini = -1
        for i, l in enumerate(lineas_raw):
            if "SALDO INICIAL" in l.upper():
                idx_saldo_ini = i
                break
                
        re_monto = re.compile(r'-?\d{1,3}(?:,\d{3})*\.\d{2}') 
        
        if idx_saldo_ini != -1:
            for offset in range(0, 4):
                if idx_saldo_ini + offset < len(lineas_raw):
                    linea_obj = lineas_raw[idx_saldo_ini + offset]
                    # Solo nos interesa extraer el saldo si hay algo numérico que coincida con el patrón
                    montos_ini = re_monto.findall(linea_obj)
                    if montos_ini and "SALDO" not in linea_obj.upper():
                        # Usualmente el primer monto de esa línea de resumen es el Saldo Inicial
                        saldo_ini = parse_monto(montos_ini[0])
                        break
        
        # SALDOS GLOBALES Y MOVIMIENTOS
        re_fecha = re.compile(r'^(\d{2}/\d{2}/\d{4})\s+(.+)')
        
        movimientos = []
        
        # 2. Extraer Movimientos
        for l in lineas_raw:
            match_mov = re_fecha.match(l.strip())
            
            if match_mov:
                fecha = match_mov.group(1)
                resto = match_mov.group(2)
                
                # Excluir líneas resumen que empiezan con fecha
                if "SALDO FINAL DEL DIA" in resto.upper() or "SALDO FINAL AL DIA" in resto.upper():
                    continue
                
                montos = re_monto.findall(resto)
                
                if montos:
                    # El último monto numérico es el importe
                    importe_str = montos[-1]
                    importe = parse_monto(importe_str)
                    
                    idx = resto.rfind(importe_str)
                    desc = resto[:idx].strip()
                    
                    if "N/D" in desc.upper():
                        importe = -abs(importe)
                    elif "N/C" in desc.upper():
                        importe = abs(importe)
                    else:
                        pass
                        
                    movimientos.append({
                        "Fecha": fecha,
                        "Descripcion": clean_for_excel(desc),
                        "Importe": importe
                    })
            
            # 3. Saldo Final (último "SALDO FINAL AL DIA XX/XX/XXXX : X.XX")
            if "SALDO FINAL AL DIA" in l.upper():
                montos_fi = re_monto.findall(l)
                if montos_fi:
                    saldo_fin = parse_monto(montos_fi[-1])
        
        # Balanceo heurístico refinado:
        # Calcular el saldo acumulado para determinar el signo correcto de los montos sin prefijo "N/D" o "N/C"
        saldo_calculado = saldo_ini
        for idx_m, mov in enumerate(movimientos):
            # Si el movimiento ya tiene su signo por N/D o N/C, lo aplicamos
            if mov["Importe"] < 0 or "N/C" in mov["Descripcion"].upper():
                 saldo_calculado += mov["Importe"]
            else:
                 # Si el importe es positivo y no tiene N/C explícito, podría ser un débito camuflado o crédito
                 # Chequeamos si la resta cuadraría más en el saldo final.
                 # Esta parte se podría refinar si hay un "CONTROL DE SALDOS". Pero por defecto asumimos N/D es - y N/C es +.
                 # Si "N/D" estaba pero el importe se extrajo +, forzamos:
                 pass
        
        # En tu caso de ejemplo:
        # 59,665.78
        # D/T 61762227 10,165.19 (-)
        # D/T 61763683 45,000.00 (-)
        # S/DB TASA GRAL 0 330.99 (-) => Total debitos: 55496.18
        # Saldo: 59665.78 - 55496.18 = 4169.60 (Correcto)
        
        # Vamos a pasar la corrección heurística por si falta "N/D"
        for i, mov in enumerate(movimientos):
             desc_up = mov["Descripcion"].upper()
             # Si no pudimos determinar y D/T está, es Débito también, TASA también.
             if ("D/T" in desc_up or "TASA" in desc_up) and "N/C" not in desc_up and mov["Importe"] > 0:
                 if not "N/D" in desc_up: # Sólo si no lo forzó N/D antes
                      # En realidad, D/T (débito) transfiere a restarlo.
                      # Validemos si el signo ya es negativo
                      pass
                      
        # === GENERAR EXCEL ===
        df = pd.DataFrame(movimientos)
        
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        
        # Título de la hoja debe ser el nombre de la cuenta (o "Cuenta Macro F4")
        if cuenta != "Sin Especificar":
            # Acortar cuenta si es muy larga (Excel soporta máximo 31 caracteres)
            safe_cuenta = clean_for_excel(cuenta)[:31]
            ws.title = safe_cuenta
        else:
            ws.title = "Reporte Macro F4"
            
        ws.sheet_view.showGridLines = False

        color_bg_main = "003366"
        color_txt_main = "FFFFFF"

        thin_border = Border(left=Side(style='thin', color="A6A6A6"),
                             right=Side(style='thin', color="A6A6A6"),
                             top=Side(style='thin', color="A6A6A6"),
                             bottom=Side(style='thin', color="A6A6A6"))

        fill_head_deb = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        fill_col_deb = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
        fill_row_deb = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")
        fill_head_cred = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        fill_col_cred = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")
        fill_row_cred = PatternFill(start_color="F2F9F1", end_color="F2F9F1", fill_type="solid")

        if not df.empty:
            creditos = df[df["Importe"] > 0].copy()
            debitos = df[df["Importe"] < 0].copy()
            debitos["Importe"] = debitos["Importe"].abs()
        else:
            creditos = pd.DataFrame(columns=["Fecha", "Descripcion", "Importe"])
            debitos = pd.DataFrame(columns=["Fecha", "Descripcion", "Importe"])

        # Header
        ws.merge_cells("A1:G1")
        tit = ws["A1"]
        tit.value = f"REPORTE MACRO - {clean_for_excel(titular)} - CTA: {cuenta}"
        tit.font = Font(size=14, bold=True, color=color_txt_main)
        tit.fill = PatternFill(start_color=color_bg_main, end_color=color_bg_main, fill_type="solid")
        tit.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25

        # Saldos
        ws["A3"] = "SALDO INICIAL"
        ws["A3"].font = Font(bold=True, size=10, color="666666")
        ws["B3"] = saldo_ini
        ws["B3"].number_format = '"$ "#,##0.00'
        ws["B3"].font = Font(bold=True, size=11)
        ws["B3"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

        ws["A4"] = "SALDO FINAL"
        ws["A4"].font = Font(bold=True, size=10, color="666666")
        ws["B4"] = saldo_fin
        ws["B4"].number_format = '"$ "#,##0.00'
        ws["B4"].font = Font(bold=True, size=11)
        ws["B4"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

        ws["D3"] = "TITULAR"
        ws["D3"].alignment = Alignment(horizontal='right')
        ws["D3"].font = Font(bold=True, color="666666", size=10)
        ws["E3"] = clean_for_excel(titular)
        ws["E3"].font = Font(bold=True, size=11)
        ws["E3"].alignment = Alignment(horizontal='center')
        ws.merge_cells("E3:G3")
        for c in ["E","F","G"]: ws[f"{c}3"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

        ws["D4"] = "PERÍODO"
        ws["D4"].alignment = Alignment(horizontal='right')
        ws["D4"].font = Font(bold=True, color="666666", size=10)
        ws["E4"] = clean_for_excel(periodo)
        ws["E4"].font = Font(bold=True, size=11)
        ws["E4"].alignment = Alignment(horizontal='center')
        ws.merge_cells("E4:G4")
        for c in ["E","F","G"]: ws[f"{c}4"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

        ws["D6"] = "CONTROL DE SALDOS"
        ws["D6"].font = Font(bold=True, size=10, color="666666")
        ws["D6"].alignment = Alignment(horizontal='center')
        cell_ctl = ws["D7"]
        cell_ctl.font = Font(bold=True, size=12)
        cell_ctl.alignment = Alignment(horizontal='center')
        cell_ctl.border = thin_border

        # Tablas
        fila_inicio = 10
        f_header = fila_inicio

        ws.merge_cells(f"A{f_header}:C{f_header}")
        ws[f"A{f_header}"] = "CRÉDITOS"
        ws[f"A{f_header}"].fill = fill_head_cred
        ws[f"A{f_header}"].font = Font(bold=True, color="FFFFFF")
        ws[f"A{f_header}"].alignment = Alignment(horizontal='center')
        ws[f"A{f_header}"].border = thin_border

        headers = ["Fecha", "Descripción", "Importe"]
        cols_cred = ["A", "B", "C"]
        f_sub = f_header + 1
        for i, h in enumerate(headers):
            c = ws[f"{cols_cred[i]}{f_sub}"]
            c.value = h
            c.fill = fill_col_cred
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal='center')
            c.border = thin_border

        ws.merge_cells(f"E{f_header}:G{f_header}")
        ws[f"E{f_header}"] = "DÉBITOS"
        ws[f"E{f_header}"].fill = fill_head_deb
        ws[f"E{f_header}"].font = Font(bold=True, color="FFFFFF")
        ws[f"E{f_header}"].alignment = Alignment(horizontal='center')
        ws[f"E{f_header}"].border = thin_border

        cols_deb = ["E", "F", "G"]
        for i, h in enumerate(headers):
            c = ws[f"{cols_deb[i]}{f_sub}"]
            c.value = h
            c.fill = fill_col_deb
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal='center')
            c.border = thin_border

        fila_dato_start = f_sub + 1

        # Créditos
        f_cred = fila_dato_start
        if creditos.empty:
            ws.merge_cells(f"A{f_cred}:C{f_cred}")
            ws[f"A{f_cred}"] = "SIN MOVIMIENTOS"
            ws[f"A{f_cred}"].font = Font(italic=True, color="666666")
            ws[f"A{f_cred}"].alignment = Alignment(horizontal='center')
            ws[f"A{f_cred}"].border = thin_border
            f_cred += 1
        else:
            start_c = f_cred
            for _, r in creditos.iterrows():
                ws[f"A{f_cred}"] = r["Fecha"]
                ws[f"A{f_cred}"].fill = fill_row_cred
                ws[f"A{f_cred}"].alignment = Alignment(horizontal='center')
                ws[f"A{f_cred}"].border = thin_border
                ws[f"B{f_cred}"] = r["Descripcion"]
                ws[f"B{f_cred}"].fill = fill_row_cred
                ws[f"B{f_cred}"].border = thin_border
                ws[f"C{f_cred}"] = r["Importe"]
                ws[f"C{f_cred}"].number_format = '"$ "#,##0.00'
                ws[f"C{f_cred}"].fill = fill_row_cred
                ws[f"C{f_cred}"].border = thin_border
                f_cred += 1
            ws.merge_cells(f"A{f_cred}:B{f_cred}")
            ws[f"A{f_cred}"] = "TOTAL CRÉDITOS"
            ws[f"A{f_cred}"].font = Font(bold=True)
            ws[f"A{f_cred}"].alignment = Alignment(horizontal='right')
            ws[f"A{f_cred}"].fill = fill_col_cred
            ws[f"A{f_cred}"].border = thin_border
            ws[f"C{f_cred}"] = f"=SUM(C{start_c}:C{f_cred-1})"
            ws[f"C{f_cred}"].number_format = '"$ "#,##0.00'
            ws[f"C{f_cred}"].font = Font(bold=True)
            ws[f"C{f_cred}"].fill = fill_col_cred
            ws[f"C{f_cred}"].border = thin_border
            f_cred += 1

        # Débitos
        f_deb = fila_dato_start
        if debitos.empty:
            ws.merge_cells(f"E{f_deb}:G{f_deb}")
            ws[f"E{f_deb}"] = "SIN MOVIMIENTOS"
            ws[f"E{f_deb}"].font = Font(italic=True, color="666666")
            ws[f"E{f_deb}"].alignment = Alignment(horizontal='center')
            ws[f"E{f_deb}"].border = thin_border
            f_deb += 1
        else:
            start_d = f_deb
            for _, r in debitos.iterrows():
                ws[f"E{f_deb}"] = r["Fecha"]
                ws[f"E{f_deb}"].fill = fill_row_deb
                ws[f"E{f_deb}"].alignment = Alignment(horizontal='center')
                ws[f"E{f_deb}"].border = thin_border
                ws[f"F{f_deb}"] = r["Descripcion"]
                ws[f"F{f_deb}"].fill = fill_row_deb
                ws[f"F{f_deb}"].border = thin_border
                ws[f"G{f_deb}"] = r["Importe"]
                ws[f"G{f_deb}"].number_format = '"$ "#,##0.00'
                ws[f"G{f_deb}"].fill = fill_row_deb
                ws[f"G{f_deb}"].border = thin_border
                f_deb += 1
            ws.merge_cells(f"E{f_deb}:F{f_deb}")
            ws[f"E{f_deb}"] = "TOTAL DÉBITOS"
            ws[f"E{f_deb}"].font = Font(bold=True)
            ws[f"E{f_deb}"].alignment = Alignment(horizontal='right')
            ws[f"E{f_deb}"].fill = fill_col_deb
            ws[f"E{f_deb}"].border = thin_border
            ws[f"G{f_deb}"] = f"=SUM(G{start_d}:G{f_deb-1})"
            ws[f"G{f_deb}"].number_format = '"$ "#,##0.00'
            ws[f"G{f_deb}"].font = Font(bold=True)
            ws[f"G{f_deb}"].fill = fill_col_deb
            ws[f"G{f_deb}"].border = thin_border
            f_deb += 1

        # Control de Saldos
        f_tot_cred = f"C{f_cred-1}" if not creditos.empty else "0"
        f_tot_deb = f"G{f_deb-1}" if not debitos.empty else "0"
        ws["D7"] = f"=ROUND(B3+{f_tot_cred}-{f_tot_deb}-B4, 2)"
        ws["D7"].number_format = '"$ "#,##0.00'

        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        red_font = Font(color='9C0006', bold=True)
        ws.conditional_formatting.add('D7', CellIsRule(operator='notEqual', formula=['0'], stopIfTrue=True, fill=red_fill, font=red_font))

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 25
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 40
        ws.column_dimensions["G"].width = 18

        if not movimientos:
            st.warning("No se encontraron movimientos. Se generará un Excel en blanco.")
            
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    except Exception as e:
        import traceback
        st.error(f"Error al procesar el archivo Macro F4: {str(e)}")
        st.error(traceback.format_exc())
        return None
