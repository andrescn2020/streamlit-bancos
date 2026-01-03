import streamlit as st
import pdfplumber
import re
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule

# Regex
ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

def clean_for_excel(text):
    if not text: return ""
    text = str(text)
    text = ILLEGAL_CHARACTERS_RE.sub("", text)
    return text.strip()

def parse_amount(s):
    if not s: return 0.0
    s = s.strip()
    sign = 1.0
    if s.endswith("-"):
        sign = -1.0
        s = s[:-1]
    elif s.startswith("-"):
        sign = -1.0
        s = s[1:]
    
    try:
        val = float(s.replace(",", ""))
        return val * sign
    except:
        return 0.0

def procesar_hipotecario(archivo_pdf):
    st.info("Procesando archivo del Banco Hipotecario...")
    try:
        archivo_pdf.seek(0)
        with pdfplumber.open(io.BytesIO(archivo_pdf.read())) as pdf:
            texto_completo = ""
            for page in pdf.pages:
                texto_completo += page.extract_text() + "\n"
        
        lineas = texto_completo.splitlines()

        # 1. Metadatos
        titular = "Sin Especificar"
        periodo = "Sin Especificar"
        cuenta = "Sin Especificar"
        saldo_inicial = 0.0
        saldo_final = 0.0

        # Titular: "Sr(es): JAMAN SRL..."
        match_titular = re.search(r"Sr\(es\):\s+(.*)", texto_completo)
        if match_titular:
            titular = match_titular.group(1).strip()
        
        # Periodo: "Período del Extracto: 01/01/2025 al 31/01/2025"
        match_periodo = re.search(r"Período del Extracto:\s+(\d{2}/\d{2}/\d{4})\s+al\s+(\d{2}/\d{2}/\d{4})", texto_completo)
        if match_periodo:
            periodo = f"Del {match_periodo.group(1)} al {match_periodo.group(2)}"
            
        # Cuenta
        match_cuenta = re.search(r"CUENTA CORRIENTE EN PESOS Nº\s+([\d-]+)", texto_completo)
        if match_cuenta:
            cuenta = match_cuenta.group(1)

        # Saldos Header
        # Regex mejorado para capturar negativos (-X.XX o X.XX-)
        # Patrón flexible para buscar la línea de saldos
        # Busca 5 montos en una línea, permitiendo signos negativos
        
        # Iterar buscando línea que contenga montos monetarios coherentes
        saldos_linea_raw = None
        for line in lineas:
             if "SALDO INICIAL" in line and "SALDO FINAL" in line:
                 continue # Header texto
             
             # Buscar linea con multiples montos
             matches = re.findall(r"-?[\d,]+\.\d{2}-?", line)
             if len(matches) >= 5:
                 # Asumimos que esta es la linea de valores de saldos
                 # [SaldoIni, Creditos, Debitos, IVA, SaldoFin] (o similar)
                 saldos_linea_raw = matches
                 break
        
        if saldos_linea_raw:
             saldo_inicial = parse_amount(saldos_linea_raw[0])
             saldo_final = parse_amount(saldos_linea_raw[-1])
        else:
             # Fallback regex directo si no encuentra por linea
             match_saldos = re.search(r"\$\s+(-?[\d,]+\.\d{2}-?)\s+\$\s+-?[\d,]+\.\d{2}-?\s+\$\s+-?[\d,]+\.\d{2}-?\s+\$\s+-?[\d,]+\.\d{2}-?\s+\$\s+(-?[\d,]+\.\d{2}-?)", texto_completo)
             if match_saldos:
                saldo_inicial = parse_amount(match_saldos.group(1))
                saldo_final = parse_amount(match_saldos.group(2))

        # 2. Movimientos
        transactions = []
        
        keywords_credito = ["N/C", "ACRED", "CREDITO", "DEVOLUCION", "DEPOSITO", "RESCATE", "RECIBISTE"] 
        keywords_debito = ["N/D", "DEBITO", "RETENCION", "IMPUESTO", "COMISION", "DB TRF", "CHEQUE", "EXTRACCION", "PAGO", "DB ", "COMIS"]
        
        for line in lineas:
            line = line.strip()
            # Detectar fecha
            if not re.match(r"^\d{2}/\d{2}/\d{4}", line):
                continue
            
            # Ignorar lineas de saldo diario explicitas
            if "SALDO FINAL DEL DIA" in line or "SALDO FINAL AL DIA" in line or "SALDO INICIAL" in line:
                continue
            
            parts = line.split()
            fecha = parts[0]
            
            monto_str = parts[-1]
            try:
                importe_abs = parse_amount(monto_str)
            except:
                continue
            
            descripcion_full = " ".join(parts[1:-1])
            
            es_credito = False
            desc_upper = descripcion_full.upper()
            
            if any(k in desc_upper for k in keywords_credito):
                es_credito = True
            elif any(k in desc_upper for k in keywords_debito):
                es_credito = False
            
            # Refinamiento Cheques
            if "CHEQUE" in desc_upper and "ACREDITACION" in desc_upper:
                es_credito = True
            
            amount = importe_abs if es_credito else -importe_abs
            
            transactions.append({
                "Fecha": fecha,
                "Descripcion": clean_for_excel(descripcion_full),
                "Importe": amount
            })
            
        if not transactions:
            st.warning("No se encontraron movimientos")
            return None

        # --- EXCEL ---
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Hipotecario"
        ws.sheet_view.showGridLines = False
        
        # Paleta Hipotecario
        color_bg_main = "F37021"
        color_txt_main = "FFFFFF"
        
        thin_border = Border(left=Side(style='thin', color="A6A6A6"), 
                             right=Side(style='thin', color="A6A6A6"), 
                             top=Side(style='thin', color="A6A6A6"), 
                             bottom=Side(style='thin', color="A6A6A6"))
                             
        fill_head_deb = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        fill_col_deb = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
        fill_row_deb = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")

        fill_head_cred = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        fill_col_cred = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")
        fill_row_cred = PatternFill(start_color="F2F9F1", end_color="F2F9F1", fill_type="solid")

        df = pd.DataFrame(transactions)
        creditos = df[df["Importe"] > 0].copy()
        debitos = df[df["Importe"] < 0].copy()
        debitos["Importe"] = debitos["Importe"].abs()

        # Header
        ws.merge_cells("A1:G1")
        tit = ws["A1"]
        tit.value = f"REPORTE HIPOTECARIO - CTA {clean_for_excel(cuenta)}"
        tit.font = Font(size=14, bold=True, color=color_txt_main)
        tit.fill = PatternFill(start_color=color_bg_main, end_color=color_bg_main, fill_type="solid")
        tit.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25

        # Metadata
        ws["A3"] = "SALDO INICIAL"
        ws["A3"].font = Font(bold=True, size=10, color="666666")
        ws["B3"] = saldo_inicial
        ws["B3"].number_format = '"$ "#,##0.00'
        ws["B3"].font = Font(bold=True, size=11)
        ws["B3"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

        ws["A4"] = "SALDO FINAL"
        ws["A4"].font = Font(bold=True, size=10, color="666666")
        ws["B4"] = saldo_final
        ws["B4"].number_format = '"$ "#,##0.00'
        ws["B4"].font = Font(bold=True, size=11)
        ws["B4"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

        ws["D3"] = "TITULAR"
        ws["D3"].alignment = Alignment(horizontal='right')
        ws["D3"].font = Font(bold=True, color="666666", size=10)
        ws["E3"] = clean_for_excel(titular)
        ws["E3"].font = Font(bold=True, size=11)
        ws["E3"].alignment = Alignment(horizontal='center')
        ws.merge_cells("E3:G3")
        for c in ["E","F","G"]: ws[f"{c}3"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

        ws["D4"] = "PERÍODO"
        ws["D4"].alignment = Alignment(horizontal='right')
        ws["D4"].font = Font(bold=True, color="666666", size=10)
        ws["E4"] = clean_for_excel(periodo)
        ws["E4"].font = Font(bold=True, size=11)
        ws["E4"].alignment = Alignment(horizontal='center')
        ws.merge_cells("E4:G4")
        for c in ["E","F","G"]: ws[f"{c}4"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

        # Tablas
        fila_inicio = 10
        f_header = fila_inicio
        
        ws.merge_cells(f"A{f_header}:C{f_header}")
        ws[f"A{f_header}"] = "CRÉDITOS" 
        ws[f"A{f_header}"].fill = fill_head_cred
        ws[f"A{f_header}"].font = Font(bold=True, color="FFFFFF")
        ws[f"A{f_header}"].alignment = Alignment(horizontal='center')
        ws[f"A{f_header}"].border = thin_border
        
        ws.merge_cells(f"E{f_header}:G{f_header}")
        ws[f"E{f_header}"] = "DÉBITOS" 
        ws[f"E{f_header}"].fill = fill_head_deb
        ws[f"E{f_header}"].font = Font(bold=True, color="FFFFFF")
        ws[f"E{f_header}"].alignment = Alignment(horizontal='center')
        ws[f"E{f_header}"].border = thin_border

        headers = ["Fecha", "Descripción", "Importe"]
        cols_cred = ["A", "B", "C"]
        cols_deb = ["E", "F", "G"]
        f_sub = f_header + 1
        
        for i, h in enumerate(headers):
            c = ws[f"{cols_cred[i]}{f_sub}"]
            c.value = h
            c.fill = fill_col_cred
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal='center')
            c.border = thin_border

            d = ws[f"{cols_deb[i]}{f_sub}"]
            d.value = h
            d.fill = fill_col_deb
            d.font = Font(bold=True)
            d.alignment = Alignment(horizontal='center')
            d.border = thin_border

        fila_a_llenar = f_sub + 1
        
        # Creditos
        f_c = fila_a_llenar
        if creditos.empty:
            ws.merge_cells(f"A{f_c}:C{f_c}")
            ws[f"A{f_c}"] = "SIN MOVIMIENTOS"
            ws[f"A{f_c}"].border = thin_border
            f_c += 1
        else:
            start_c = f_c
            for _, r in creditos.iterrows():
                ws[f"A{f_c}"] = r["Fecha"]
                ws[f"A{f_c}"].fill = fill_row_cred
                ws[f"A{f_c}"].border = thin_border
                ws[f"A{f_c}"].alignment = Alignment(horizontal='center')
                ws[f"B{f_c}"] = r["Descripcion"]
                ws[f"B{f_c}"].fill = fill_row_cred
                ws[f"B{f_c}"].border = thin_border
                ws[f"C{f_c}"] = r["Importe"]
                ws[f"C{f_c}"].number_format = '"$ "#,##0.00'
                ws[f"C{f_c}"].fill = fill_row_cred
                ws[f"C{f_c}"].border = thin_border
                f_c += 1
            ws.merge_cells(f"A{f_c}:B{f_c}")
            ws[f"A{f_c}"] = "TOTAL CRÉDITOS"
            ws[f"A{f_c}"].font = Font(bold=True)
            ws[f"A{f_c}"].alignment = Alignment(horizontal='right')
            ws[f"C{f_c}"] = f"=SUM(C{start_c}:C{f_c-1})"
            ws[f"C{f_c}"].font = Font(bold=True)
            ws[f"C{f_c}"].number_format = '"$ "#,##0.00'
            f_c += 1

        # Debitos
        f_d = fila_a_llenar
        if debitos.empty:
            ws.merge_cells(f"E{f_d}:G{f_d}")
            ws[f"E{f_d}"] = "SIN MOVIMIENTOS"
            ws[f"E{f_d}"].border = thin_border
            f_d += 1
        else:
            start_d = f_d
            for _, r in debitos.iterrows():
                ws[f"E{f_d}"] = r["Fecha"]
                ws[f"E{f_d}"].fill = fill_row_deb
                ws[f"E{f_d}"].border = thin_border
                ws[f"E{f_d}"].alignment = Alignment(horizontal='center')
                ws[f"F{f_d}"] = r["Descripcion"]
                ws[f"F{f_d}"].fill = fill_row_deb
                ws[f"F{f_d}"].border = thin_border
                ws[f"G{f_d}"] = r["Importe"]
                ws[f"G{f_d}"].number_format = '"$ "#,##0.00'
                ws[f"G{f_d}"].fill = fill_row_deb
                ws[f"G{f_d}"].border = thin_border
                f_d += 1
            ws.merge_cells(f"E{f_d}:F{f_d}")
            ws[f"E{f_d}"] = "TOTAL DÉBITOS"
            ws[f"E{f_d}"].font = Font(bold=True)
            ws[f"E{f_d}"].alignment = Alignment(horizontal='right')
            ws[f"G{f_d}"] = f"=SUM(G{start_d}:G{f_d-1})"
            ws[f"G{f_d}"].font = Font(bold=True)
            ws[f"G{f_d}"].number_format = '"$ "#,##0.00'
            f_d += 1

        # Control
        ws["D6"] = "CONTROL DE SALDOS"
        ws["D6"].font = Font(bold=True, size=10, color="666666")
        ws["D6"].alignment = Alignment(horizontal='center')
        
        ref_tot_c = f"C{f_c-1}" if not creditos.empty else "0"
        ref_tot_d = f"G{f_d-1}" if not debitos.empty else "0"
        ws["D7"] = f"=ROUND(B3+{ref_tot_c}-{ref_tot_d}-B4, 2)"
        ws["D7"].number_format = '"$ "#,##0.00'
        ws["D7"].font = Font(bold=True)
        ws["D7"].alignment = Alignment(horizontal='center')
        ws["D7"].border = thin_border
        
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        red_font = Font(color='9C0006', bold=True)
        ws.conditional_formatting.add('D7', CellIsRule(operator='notEqual', formula=['0'], stopIfTrue=True, fill=red_fill, font=red_font))

        # Anchos
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 25
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 40
        ws.column_dimensions["G"].width = 18

        wb.save(output)
        output.seek(0)
        return output.getvalue()
        
    except Exception as e:
        import traceback
        st.error(f"Error al procesar: {e}")
        print(traceback.format_exc())
        return None
