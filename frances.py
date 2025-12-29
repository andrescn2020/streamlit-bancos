import streamlit as st
import io
import pdfplumber
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule

# Regex para caracteres ilegales en Excel
ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

def clean_for_excel(text):
    """Elimina caracteres ilegales para Excel y espacios extra"""
    if not text: return ""
    text = str(text)
    text = ILLEGAL_CHARACTERS_RE.sub("", text)
    return text.strip()

def procesar_bbva_frances(archivo_pdf):
    """Procesa archivos PDF de BBVA Frances con Estilo Dashboard"""
    st.info("Procesando archivo de BBVA Frances...")

    try:
        # Leer el PDF usando pdfplumber
        with pdfplumber.open(io.BytesIO(archivo_pdf.read())) as pdf:
            texto_completo = ""
            for page in pdf.pages:
                texto_completo += page.extract_text() + "\n"
            
        lineas = texto_completo.splitlines()

        # 1. Extracción de Metadata Global (Titular, Período)
        titular_global = "Sin Especificar"
        periodo_global = "Sin Especificar"
        
        # Extracción Titular: Buscar sección 'Intervinientes'
        for i, l in enumerate(lineas[:100]):
            if "Intervinientes" in l:
                # El titular suele estar en la línea siguiente (i+1)
                if i + 1 < len(lineas):
                    raw_tit = lineas[i+1].strip()
                    # Quitar CUIT entre paréntesis si existe: "JUAN PEREZ (20-123...)"
                    raw_tit = re.sub(r"\(\d{2}-\d{8,}-\d\)", "", raw_tit)
                    titular_global = raw_tit.strip()
                break
        
        # Extracción Período (Intento Regex Global)
        match_per = re.search(r"del\s+(\d{2}/\d{2}/\d{4})\s+al\s+(\d{2}/\d{2}/\d{4})", texto_completo, re.IGNORECASE)
        if match_per:
            periodo_global = f"Del {match_per.group(1)} al {match_per.group(2)}"
        
        # (Si falla el regex, calcularemos el período basado en las fechas de movimientos más adelante)

        # 2. Lógica de Extracción de Movimientos (Existente)
        inicio = next((i for i, line in enumerate(lineas) if "Movimientos en cuentas" in line), None)
        fin = next((i for i, line in enumerate(lineas) if "Transferencias" in line), None)

        if inicio is None:
             st.error("No se encontró la sección 'Movimientos en cuentas'")
             return None
        
        # Si no encuentra "Transferencias", usar el final del archivo
        movimientos_extraidos = lineas[inicio + 1 : fin] if fin else lineas[inicio+1:]

        pattern_cuenta = r"^(CA|CC)\s"
        cuentas = []

        for index, movimiento in enumerate(movimientos_extraidos):
            if re.match(pattern_cuenta, movimiento):
                corte = movimiento.find("(") - 1 if "(" in movimiento else len(movimiento)
                cuenta = {
                    "cuenta": movimiento[:corte].strip(),
                    "inicio": index,
                    "saldo_inicial": 0.0,
                    "saldo_final": 0.0,
                    "fin": len(movimientos_extraidos)
                }
                # Buscar fin del bloque
                for j in range(index + 1, len(movimientos_extraidos)):
                    if "TOTAL MOVIMIENTOS" in movimientos_extraidos[j] or re.match(pattern_cuenta, movimientos_extraidos[j]):
                         # Si encontramos TOTAL MOVIMIENTOS o el inicio de OTRA cuenta
                        if "TOTAL MOVIMIENTOS" in movimientos_extraidos[j]:
                             cuenta["fin"] = j
                        else:
                             cuenta["fin"] = j # Cortar antes de la siguiente cuenta
                        cuentas.append(cuenta)
                        break
                else:
                    # Si llega al final sin encontrar cierre explícito
                    cuentas.append(cuenta)

        # Deduplicar cuentas por índice de inicio (por si la lógica anterior falló)
        cuentas_unicas = {c['inicio']: c for c in cuentas}.values()

        if not cuentas_unicas:
            st.warning("No se encontraron cuentas en el PDF")
            return None

        # --- GENERACIÓN EXCEL (ESTILO DASHBOARD) ---
        output = io.BytesIO()
        wb = Workbook()
        # Eliminar hoja default
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        
        # Estilos
        thin_border = Border(left=Side(style='thin', color="A6A6A6"), 
                             right=Side(style='thin', color="A6A6A6"), 
                             top=Side(style='thin', color="A6A6A6"), 
                             bottom=Side(style='thin', color="A6A6A6"))
        
        color_bg_main = "004481" # Azul BBVA aproximado
        color_txt_main = "FFFFFF"
        
        fill_head_deb = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        fill_col_deb = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
        fill_row_deb = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")

        fill_head_cred = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        fill_col_cred = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")
        fill_row_cred = PatternFill(start_color="F2F9F1", end_color="F2F9F1", fill_type="solid")

        # Procesar cada cuenta
        for cuenta_info in cuentas_unicas:
            nombre_hoja = clean_for_excel(cuenta_info["cuenta"].replace("/", "-"))[:30]
            ws = wb.create_sheet(title=nombre_hoja)
            ws.sheet_view.showGridLines = False
            
            # Extraer Movimientos de esta cuenta
            pattern = r"(\d{2}/\d{2})\s([A-Za-z0-9\s\./,\-+]+)\s([-]?\d{1,3}(?:[\.,]\d{3})*(?:[\.,]\d{2}))\s"
            resultados = []
            
            raw_lines = movimientos_extraidos[cuenta_info["inicio"] + 1 : cuenta_info["fin"]]
            
            saldo_inicial = 0.0
            saldo_final = 0.0

            for linea in raw_lines:
                # Extraer saldos
                if "SALDO ANTERIOR" in linea:
                    matches = re.findall(r"[-]?\d{1,3}(?:\.\d{3})*,\d{2}", linea)
                    if matches:
                        saldo_inicial = float(matches[0].replace(".", "").replace(",", "."))
                
                if "SALDO AL" in linea:
                    matches = re.findall(r"[-]?\d{1,3}(?:\.\d{3})*,\d{2}", linea)
                    if matches:
                        saldo_final = float(matches[0].replace(".", "").replace(",", "."))

                # Limpieza SIRCREB
                if "SIRCREB" in linea and "F:" in linea:
                     # Lógica original conservada
                     try:
                        inicio_mov = linea.index("F:")
                        fin_mov = inicio_mov + 10
                        linea = linea[:inicio_mov] + linea[fin_mov:]
                     except: pass

                match = re.match(pattern, linea)
                if match:
                    fecha = match.group(1)
                    descripcion = match.group(2).strip()
                    importe_str = match.group(3).replace(",", ".") # Ojo, BBVA suele usar 1.000,00 -> replace . then , -> .
                    # La logica original era: .replace(".", "", count-1)
                    # Vamos a usar una logica más robusta
                    # Formato esperado: -1.234,56
                    try:
                        importe_clean = match.group(3).replace(".", "").replace(",", ".")
                        importe = float(importe_clean)
                    except:
                        importe = 0.0
                        
                    resultados.append({
                        "Fecha": fecha, 
                        "Descripcion": descripcion, 
                        "Importe": importe
                    })
            
            # DataFrames
            df = pd.DataFrame(resultados)
            
            if not df.empty:
                creditos = df[df["Importe"] > 0].copy()
                debitos = df[df["Importe"] < 0].copy()
                debitos["Importe"] = debitos["Importe"].abs()
            else:
                creditos = pd.DataFrame(columns=["Fecha", "Descripcion", "Importe"])
                debitos = pd.DataFrame(columns=["Fecha", "Descripcion", "Importe"])

            # --- DIBUJAR DASHBOARD EN HOJA ---
            
            # 1. Título
            ws.merge_cells("A1:G1")
            tit = ws["A1"]
            tit.value = f"REPORTE BBVA FRANCES - {clean_for_excel(titular_global)}"
            tit.font = Font(size=14, bold=True, color=color_txt_main)
            tit.fill = PatternFill(start_color=color_bg_main, end_color=color_bg_main, fill_type="solid")
            tit.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 25

            # 2. Saldos y Metadata
            ws["A3"] = "SALDO INICIAL"
            ws["A3"].font = Font(bold=True, size=10, color="666666")
            ws["B3"] = saldo_inicial
            ws["B3"].number_format = '"$ "#,##0.00'
            ws["B3"].font = Font(bold=True, size=11)
            ws["B3"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

            ws["A4"] = "SALDO FINAL"
            ws["A4"].font = Font(bold=True, size=10, color="666666")
            ws["B4"] = saldo_final
            ws["B4"].number_format = '"$ "#,##0.00'
            ws["B4"].font = Font(bold=True, size=11)
            ws["B4"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

            ws["D3"] = "TITULAR"
            ws["D3"].alignment = Alignment(horizontal='right')
            ws["D3"].font = Font(bold=True, color="666666", size=10)
            ws["E3"] = clean_for_excel(titular_global)
            ws["E3"].font = Font(bold=True, size=11)
            ws["E3"].alignment = Alignment(horizontal='center')
            ws.merge_cells("E3:G3")
            for c in ["E","F","G"]: ws[f"{c}3"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

            ws["D4"] = "PERÍODO"
            ws["D4"].alignment = Alignment(horizontal='right')
            ws["D4"].font = Font(bold=True, color="666666", size=10)
            ws["E4"] = clean_for_excel(periodo_global)
            ws["E4"].font = Font(bold=True, size=11)
            ws["E4"].alignment = Alignment(horizontal='center')
            ws.merge_cells("E4:G4")
            for c in ["E","F","G"]: ws[f"{c}4"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

            ws["D6"] = "CONTROL DE SALDOS"
            ws["D6"].font = Font(bold=True, size=10, color="666666")
            ws["D6"].alignment = Alignment(horizontal='center')
            
            cell_ctl = ws["D7"]
            cell_ctl.font = Font(bold=True, size=12)
            cell_ctl.alignment = Alignment(horizontal='center')
            cell_ctl.border = thin_border

            # 3. Tablas Paralelas
            fila_inicio = 10
            
            # Headers Créditos
            f_header = fila_inicio
            ws.merge_cells(f"A{f_header}:C{f_header}")
            ws[f"A{f_header}"] = "CRÉDITOS" 
            ws[f"A{f_header}"].fill = fill_head_cred
            ws[f"A{f_header}"].font = Font(bold=True, color="FFFFFF")
            ws[f"A{f_header}"].alignment = Alignment(horizontal='center')
            ws[f"A{f_header}"].border = thin_border
            
            headers = ["Fecha", "Descripción", "Importe"]
            cols_cred = ["A", "B", "C"]
            f_sub = f_header + 1
            for i, h in enumerate(headers):
                c = ws[f"{cols_cred[i]}{f_sub}"]
                c.value = h
                c.fill = fill_col_cred
                c.font = Font(bold=True)
                c.alignment = Alignment(horizontal='center')
                c.border = thin_border

            # Headers Débitos
            ws.merge_cells(f"E{f_header}:G{f_header}")
            ws[f"E{f_header}"] = "DÉBITOS" 
            ws[f"E{f_header}"].fill = fill_head_deb
            ws[f"E{f_header}"].font = Font(bold=True, color="FFFFFF")
            ws[f"E{f_header}"].alignment = Alignment(horizontal='center')
            ws[f"E{f_header}"].border = thin_border
            
            cols_deb = ["E", "F", "G"]
            for i, h in enumerate(headers):
                c = ws[f"{cols_deb[i]}{f_sub}"]
                c.value = h
                c.fill = fill_col_deb
                c.font = Font(bold=True)
                c.alignment = Alignment(horizontal='center')
                c.border = thin_border

            # Datos
            fila_dato_start = f_sub + 1
            
            # Llenar Créditos
            f_cred = fila_dato_start
            if creditos.empty:
                ws.merge_cells(f"A{f_cred}:C{f_cred}")
                ws[f"A{f_cred}"] = "SIN MOVIMIENTOS"
                ws[f"A{f_cred}"].font = Font(italic=True, color="666666")
                ws[f"A{f_cred}"].alignment = Alignment(horizontal='center')
                ws[f"A{f_cred}"].border = thin_border
                f_cred += 1
            else:
                start_c = f_cred
                for _, r in creditos.iterrows():
                    ws[f"A{f_cred}"] = clean_for_excel(r["Fecha"])
                    ws[f"A{f_cred}"].fill = fill_row_cred
                    ws[f"A{f_cred}"].alignment = Alignment(horizontal='center')
                    ws[f"A{f_cred}"].border = thin_border

                    ws[f"B{f_cred}"] = clean_for_excel(r["Descripcion"])
                    ws[f"B{f_cred}"].fill = fill_row_cred
                    ws[f"B{f_cred}"].border = thin_border

                    ws[f"C{f_cred}"] = r["Importe"]
                    ws[f"C{f_cred}"].number_format = '"$ "#,##0.00'
                    ws[f"C{f_cred}"].fill = fill_row_cred
                    ws[f"C{f_cred}"].border = thin_border
                    f_cred += 1
                
                # Total
                ws.merge_cells(f"A{f_cred}:B{f_cred}")
                ws[f"A{f_cred}"] = "TOTAL CRÉDITOS"
                ws[f"A{f_cred}"].font = Font(bold=True)
                ws[f"A{f_cred}"].alignment = Alignment(horizontal='right')
                ws[f"A{f_cred}"].border = thin_border
                
                ws[f"C{f_cred}"] = f"=SUM(C{start_c}:C{f_cred-1})"
                ws[f"C{f_cred}"].number_format = '"$ "#,##0.00'
                ws[f"C{f_cred}"].font = Font(bold=True)
                ws[f"C{f_cred}"].border = thin_border
                f_cred += 1

            # Llenar Débitos
            f_deb = fila_dato_start
            if debitos.empty:
                ws.merge_cells(f"E{f_deb}:G{f_deb}")
                ws[f"E{f_deb}"] = "SIN MOVIMIENTOS"
                ws[f"E{f_deb}"].font = Font(italic=True, color="666666")
                ws[f"E{f_deb}"].alignment = Alignment(horizontal='center')
                ws[f"E{f_deb}"].border = thin_border
                f_deb += 1
            else:
                start_d = f_deb
                for _, r in debitos.iterrows():
                    ws[f"E{f_deb}"] = clean_for_excel(r["Fecha"])
                    ws[f"E{f_deb}"].fill = fill_row_deb
                    ws[f"E{f_deb}"].alignment = Alignment(horizontal='center')
                    ws[f"E{f_deb}"].border = thin_border

                    ws[f"F{f_deb}"] = clean_for_excel(r["Descripcion"])
                    ws[f"F{f_deb}"].fill = fill_row_deb
                    ws[f"F{f_deb}"].border = thin_border

                    ws[f"G{f_deb}"] = r["Importe"]
                    ws[f"G{f_deb}"].number_format = '"$ "#,##0.00'
                    ws[f"G{f_deb}"].fill = fill_row_deb
                    ws[f"G{f_deb}"].border = thin_border
                    f_deb += 1
                
                # Total
                ws.merge_cells(f"E{f_deb}:F{f_deb}")
                ws[f"E{f_deb}"] = "TOTAL DÉBITOS"
                ws[f"E{f_deb}"].font = Font(bold=True)
                ws[f"E{f_deb}"].alignment = Alignment(horizontal='right')
                ws[f"E{f_deb}"].border = thin_border
                
                ws[f"G{f_deb}"] = f"=SUM(G{start_d}:G{f_deb-1})"
                ws[f"G{f_deb}"].number_format = '"$ "#,##0.00'
                ws[f"G{f_deb}"].font = Font(bold=True)
                ws[f"G{f_deb}"].border = thin_border
                f_deb += 1

            # Formula Control
            f_ini = "B3"
            f_tot_cred = f"C{f_cred-1}" if not creditos.empty else "0"
            f_tot_deb = f"G{f_deb-1}" if not debitos.empty else "0"
            f_fin = "B4"
            
            ws["D7"] = f"=ROUND({f_ini}+{f_tot_cred}-{f_tot_deb}-{f_fin}, 2)"
            ws["D7"].number_format = '"$ "#,##0.00'
            
            red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            red_font = Font(color='9C0006', bold=True)
            ws.conditional_formatting.add('D7', CellIsRule(operator='notEqual', formula=['0'], stopIfTrue=True, fill=red_fill, font=red_font))

            # Anchos
            ws.column_dimensions["A"].width = 12
            ws.column_dimensions["B"].width = 40
            ws.column_dimensions["C"].width = 18
            ws.column_dimensions["D"].width = 25
            ws.column_dimensions["E"].width = 12
            ws.column_dimensions["F"].width = 40
            ws.column_dimensions["G"].width = 18

        wb.save(output)
        output.seek(0)
        return output.getvalue()

    except Exception as e:
        import traceback
        st.error(f"Error al procesar el archivo: {str(e)}")
        print(traceback.format_exc())
        return None
