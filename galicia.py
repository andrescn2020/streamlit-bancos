import streamlit as st
import io
import PyPDF2
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule

# Regex para caracteres ilegales en Excel
ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

def clean_for_excel(text):
    """Elimina caracteres ilegales para Excel y espacios extra"""
    if not text: return ""
    text = str(text)
    text = ILLEGAL_CHARACTERS_RE.sub("", text)
    return text.strip()

def procesar_galicia(archivo_pdf):
    """Procesa archivos PDF del banco Galicia con Estilo Dashboard"""
    st.info("Procesando archivo del banco Galicia...")

    try:
        archivo_pdf.seek(0)
        with io.BytesIO(archivo_pdf.read()) as pdf_file:
            reader = PyPDF2.PdfReader(pdf_file)
            texto_completo = "".join(page.extract_text() + "\n" for page in reader.pages)
            texto = texto_completo.splitlines()

        # Eliminar líneas vacías y espacios extra
        lineas = [line.strip() for line in texto if line.strip()]

        # 1. Extracción Metadata (Titular, Período, Saldos)
        titular_global = "Sin Especificar"
        periodo_global = "Sin Especificar"
        
        # Regex Titular
        for l in lineas[:20]:
            # Caso 1: Formato "IVA: Consumidor FinalNOMBRE APELLIDOResumen..."
            if "IVA:" in l and "Resumen" in l:
                # Tomamos todo antes de "Resumen"
                parte_izq = l.split("Resumen")[0]
                # Buscamos texto en mayúsculas después de la última minúscula (ej: finaL NOMBRE)
                match_nombre = re.search(r"[a-z]([A-Z\s\.]+)$", parte_izq)
                if match_nombre:
                   titular_global = match_nombre.group(1).strip()
                   break
            
            # Caso 2: Formato "Cuenta: ... NOMBRE ... Resumen" (Anterior)
            if l.startswith("Cuenta:"):
                match_tit = re.search(r"Cuenta:.*?\d+([A-Z\s]+)Resumen", l)
                if match_tit:
                    titular_global = match_tit.group(1).strip()
                break
        
        # Regex Período: "...24/02/2023 27/01/2023Período..."
        for l in lineas[:15]:
            if "Período" in l or "Periodo" in l:
                fechas = re.findall(r"(\d{2}/\d{2}/\d{4})", l)
                if len(fechas) >= 2:
                    # Ordenar cronológicamente (DD/MM/YYYY)
                    fechas_obj = sorted(fechas, key=lambda x: x.split("/")[::-1])
                    periodo_global = f"Del {fechas_obj[0]} al {fechas_obj[-1]}"
                break

        # 2. Extracción Saldos (Ajustado: [0]=Final, [1]=Inicial)
        saldo_cuenta = 0.0
        saldo_inicial = 0.0
        saldo_final = 0.0

        for l in lineas:
            if "Saldos" in l:
                patron = r"([+-]?\$\s*\d{1,3}(?:\.\d{3})*,\d{2})"
                valores = re.findall(patron, l)
                if len(valores) >= 2:
                    # Según análisis: $0,00(Final)$0,05(Inicial)Saldos
                    val_final_raw = valores[0]
                    val_inicial_raw = valores[1]
                    
                    def clean_val(v):
                        v = v.replace("$", "").replace(" ", "")
                        sign = -1 if "-" in v else 1
                        v = v.replace("-", "")
                        return float(v.replace(".", "").replace(",", ".")) * sign

                    saldo_final = clean_val(val_final_raw)
                    saldo_inicial = clean_val(val_inicial_raw)
                    saldo_cuenta = saldo_inicial # Para el cálculo incremental

            # Lógica alternativa Saldos (Original) - Mantenida por compatibilidad
            if re.search(r"\$\d{1,3}(?:\.\d{3})*(,\d{2})?-\$\d{1,3}(?:\.\d{3})*(,\d{2})?-Saldos", l):
                 pass # (Omitida, confiamos en la primera)

        # 3. Extracción Movimientos (Lógica Original)
        inicio = next((i for i, line in enumerate(lineas) if "Movimientos" in line), None)
        fin = next((i for i, line in enumerate(lineas) if "Total" in line), None)

        if inicio is None:
            # Fallback si no encuentra Movimientos
            st.error("No se encontró sección Movimientos")
            return None
        
        movimientos_extraidos = lineas[inicio + 1 : fin] if fin else lineas[inicio+1:]
        
        # Unir líneas
        movimientos_unidos = []
        linea_actual = ""
        for linea in movimientos_extraidos:
            if re.match(r"\d{2}/\d{2}/\d{2}", linea):
                if linea_actual: movimientos_unidos.append(linea_actual.strip())
                linea_actual = linea
            else:
                linea_actual += " " + linea
        if linea_actual: movimientos_unidos.append(linea_actual.strip())

        movimientos_procesados = []
        
        # El saldo_inicial para el cálculo iterativo debe ser el Saldo Inicial del periodo
        saldo_iterativo = saldo_inicial 
        
        for linea in movimientos_unidos:
             # Ignorar encabezados internos si se colaron
             if "Fecha" in linea and "Concepto" in linea: continue
             
             matches = re.findall(r"-?\d{1,3}(?:\.\d{3})*,\d{2}-?", linea)
             match_fecha = re.match(r"(\d{2}/\d{2}/\d{2})", linea)
             
             if match_fecha and matches:
                fecha = match_fecha.group(1)
                
                # Descripción
                linea_sin_fecha = linea[len(fecha):].strip()
                descripcion = re.split(r"\d+\.\d+", linea_sin_fecha, maxsplit=1)[0]
                descripcion = re.sub(r"-?\d+[\.,]\d+", "", descripcion).strip()
                if "-" in descripcion: descripcion = descripcion.replace("-", "")
                
                # Saldo de la línea (último número)
                saldo_str = matches[-1]
                saldo_linea = float(saldo_str.replace(".", "").replace(",", ".").replace("-", ""))
                if "-" in saldo_str: saldo_linea *= -1
                
                # Calculo importe por diferencia
                importe = round(saldo_linea - saldo_iterativo, 2)
                
                # Actualizar saldo previo
                saldo_iterativo = saldo_linea
                
                movimientos_procesados.append({
                    "Fecha": fecha,
                    "Descripcion": descripcion,
                    "Importe": importe
                })
        
        saldo_final_reporte = saldo_iterativo # El último saldo calculado es el final

        # --- GENERACIÓN EXCEL (DASHBOARD) ---
        
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Galicia"
        ws.sheet_view.showGridLines = False
        
        # Paleta Galicia (Naranja/Amarillo) - Aproximación Premium
        color_bg_main = "FF6900" 
        color_txt_main = "FFFFFF"
        
        # Bordes y Fills Estandar
        thin_border = Border(left=Side(style='thin', color="A6A6A6"), 
                             right=Side(style='thin', color="A6A6A6"), 
                             top=Side(style='thin', color="A6A6A6"), 
                             bottom=Side(style='thin', color="A6A6A6"))
                             
        fill_head_deb = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        fill_col_deb = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
        fill_row_deb = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")

        fill_head_cred = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        fill_col_cred = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")
        fill_row_cred = PatternFill(start_color="F2F9F1", end_color="F2F9F1", fill_type="solid")

        # DataFrames
        df = pd.DataFrame(movimientos_procesados)
        if not df.empty:
            creditos = df[df["Importe"] > 0].copy()
            debitos = df[df["Importe"] < 0].copy()
            debitos["Importe"] = debitos["Importe"].abs()
        else:
            creditos = pd.DataFrame(columns=["Fecha", "Descripcion", "Importe"])
            debitos = pd.DataFrame(columns=["Fecha", "Descripcion", "Importe"])

        # 1. Header Global
        ws.merge_cells("A1:G1")
        tit = ws["A1"]
        tit.value = f"REPORTE GALICIA - {clean_for_excel(titular_global)}"
        tit.font = Font(size=14, bold=True, color=color_txt_main)
        tit.fill = PatternFill(start_color=color_bg_main, end_color=color_bg_main, fill_type="solid")
        tit.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25

        # 2. Metadata y Saldos
        ws["A3"] = "SALDO INICIAL"
        ws["A3"].font = Font(bold=True, size=10, color="666666")
        ws["B3"] = saldo_inicial
        ws["B3"].number_format = '"$ "#,##0.00'
        ws["B3"].font = Font(bold=True, size=11)
        ws["B3"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

        ws["A4"] = "SALDO FINAL"
        ws["A4"].font = Font(bold=True, size=10, color="666666")
        ws["B4"] = saldo_final_reporte
        ws["B4"].number_format = '"$ "#,##0.00'
        ws["B4"].font = Font(bold=True, size=11)
        ws["B4"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

        ws["D3"] = "TITULAR"
        ws["D3"].alignment = Alignment(horizontal='right')
        ws["D3"].font = Font(bold=True, color="666666", size=10)
        ws["E3"] = clean_for_excel(titular_global)
        ws["E3"].font = Font(bold=True, size=11)
        ws["E3"].alignment = Alignment(horizontal='center')
        ws.merge_cells("E3:G3")
        for c in ["E","F","G"]: ws[f"{c}3"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

        ws["D4"] = "PERÍODO"
        ws["D4"].alignment = Alignment(horizontal='right')
        ws["D4"].font = Font(bold=True, color="666666", size=10)
        ws["E4"] = clean_for_excel(periodo_global)
        ws["E4"].font = Font(bold=True, size=11)
        ws["E4"].alignment = Alignment(horizontal='center')
        ws.merge_cells("E4:G4")
        for c in ["E","F","G"]: ws[f"{c}4"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

        ws["D6"] = "CONTROL DE SALDOS"
        ws["D6"].font = Font(bold=True, size=10, color="666666")
        ws["D6"].alignment = Alignment(horizontal='center')
        
        cell_ctl = ws["D7"]
        cell_ctl.font = Font(bold=True, size=12)
        cell_ctl.alignment = Alignment(horizontal='center')
        cell_ctl.border = thin_border

        # 3. Tablas Paralelas
        fila_inicio = 10
        
        # Headers Créditos
        f_header = fila_inicio
        ws.merge_cells(f"A{f_header}:C{f_header}")
        ws[f"A{f_header}"] = "CRÉDITOS" 
        ws[f"A{f_header}"].fill = fill_head_cred
        ws[f"A{f_header}"].font = Font(bold=True, color="FFFFFF")
        ws[f"A{f_header}"].alignment = Alignment(horizontal='center')
        ws[f"A{f_header}"].border = thin_border
        
        headers = ["Fecha", "Descripción", "Importe"]
        cols_cred = ["A", "B", "C"]
        f_sub = f_header + 1
        for i, h in enumerate(headers):
            c = ws[f"{cols_cred[i]}{f_sub}"]
            c.value = h
            c.fill = fill_col_cred
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal='center')
            c.border = thin_border

        # Headers Débitos
        ws.merge_cells(f"E{f_header}:G{f_header}")
        ws[f"E{f_header}"] = "DÉBITOS" 
        ws[f"E{f_header}"].fill = fill_head_deb
        ws[f"E{f_header}"].font = Font(bold=True, color="FFFFFF")
        ws[f"E{f_header}"].alignment = Alignment(horizontal='center')
        ws[f"E{f_header}"].border = thin_border
        
        cols_deb = ["E", "F", "G"]
        for i, h in enumerate(headers):
            c = ws[f"{cols_deb[i]}{f_sub}"]
            c.value = h
            c.fill = fill_col_deb
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal='center')
            c.border = thin_border

        # Llenado de Datos
        fila_dato_start = f_sub + 1
        
        # Créditos
        f_cred = fila_dato_start
        if creditos.empty:
            ws.merge_cells(f"A{f_cred}:C{f_cred}")
            ws[f"A{f_cred}"] = "SIN MOVIMIENTOS"
            ws[f"A{f_cred}"].font = Font(italic=True, color="666666")
            ws[f"A{f_cred}"].alignment = Alignment(horizontal='center')
            ws[f"A{f_cred}"].border = thin_border
            f_cred += 1
        else:
            start_c = f_cred
            for _, r in creditos.iterrows():
                ws[f"A{f_cred}"] = clean_for_excel(r["Fecha"])
                ws[f"A{f_cred}"].fill = fill_row_cred
                ws[f"A{f_cred}"].alignment = Alignment(horizontal='center')
                ws[f"A{f_cred}"].border = thin_border

                ws[f"B{f_cred}"] = clean_for_excel(r["Descripcion"])
                ws[f"B{f_cred}"].fill = fill_row_cred
                ws[f"B{f_cred}"].border = thin_border

                ws[f"C{f_cred}"] = r["Importe"]
                ws[f"C{f_cred}"].number_format = '"$ "#,##0.00'
                ws[f"C{f_cred}"].fill = fill_row_cred
                ws[f"C{f_cred}"].border = thin_border
                f_cred += 1
            
            # Total
            ws.merge_cells(f"A{f_cred}:B{f_cred}")
            ws[f"A{f_cred}"] = "TOTAL CRÉDITOS"
            ws[f"A{f_cred}"].font = Font(bold=True)
            ws[f"A{f_cred}"].alignment = Alignment(horizontal='right')
            ws[f"A{f_cred}"].border = thin_border
            
            ws[f"C{f_cred}"] = f"=SUM(C{start_c}:C{f_cred-1})"
            ws[f"C{f_cred}"].number_format = '"$ "#,##0.00'
            ws[f"C{f_cred}"].font = Font(bold=True)
            ws[f"C{f_cred}"].border = thin_border
            f_cred += 1

        # Débitos
        f_deb = fila_dato_start
        if debitos.empty:
            ws.merge_cells(f"E{f_deb}:G{f_deb}")
            ws[f"E{f_deb}"] = "SIN MOVIMIENTOS"
            ws[f"E{f_deb}"].font = Font(italic=True, color="666666")
            ws[f"E{f_deb}"].alignment = Alignment(horizontal='center')
            ws[f"E{f_deb}"].border = thin_border
            f_deb += 1
        else:
            start_d = f_deb
            for _, r in debitos.iterrows():
                ws[f"E{f_deb}"] = clean_for_excel(r["Fecha"])
                ws[f"E{f_deb}"].fill = fill_row_deb
                ws[f"E{f_deb}"].alignment = Alignment(horizontal='center')
                ws[f"E{f_deb}"].border = thin_border

                ws[f"F{f_deb}"] = clean_for_excel(r["Descripcion"])
                ws[f"F{f_deb}"].fill = fill_row_deb
                ws[f"F{f_deb}"].border = thin_border

                ws[f"G{f_deb}"] = r["Importe"]
                ws[f"G{f_deb}"].number_format = '"$ "#,##0.00'
                ws[f"G{f_deb}"].fill = fill_row_deb
                ws[f"G{f_deb}"].border = thin_border
                f_deb += 1
            
            # Total
            ws.merge_cells(f"E{f_deb}:F{f_deb}")
            ws[f"E{f_deb}"] = "TOTAL DÉBITOS"
            ws[f"E{f_deb}"].font = Font(bold=True)
            ws[f"E{f_deb}"].alignment = Alignment(horizontal='right')
            ws[f"E{f_deb}"].border = thin_border
            
            ws[f"G{f_deb}"] = f"=SUM(G{start_d}:G{f_deb-1})"
            ws[f"G{f_deb}"].number_format = '"$ "#,##0.00'
            ws[f"G{f_deb}"].font = Font(bold=True)
            ws[f"G{f_deb}"].border = thin_border
            f_deb += 1

        # Formula Control
        f_ini = "B3"
        f_tot_cred = f"C{f_cred-1}" if not creditos.empty else "0"
        f_tot_deb = f"G{f_deb-1}" if not debitos.empty else "0"
        f_fin = "B4"
        
        ws["D7"] = f"=ROUND({f_ini}+{f_tot_cred}-{f_tot_deb}-{f_fin}, 2)"
        ws["D7"].number_format = '"$ "#,##0.00'
        
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        red_font = Font(color='9C0006', bold=True)
        ws.conditional_formatting.add('D7', CellIsRule(operator='notEqual', formula=['0'], stopIfTrue=True, fill=red_fill, font=red_font))

        # Anchos
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 25
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 40
        ws.column_dimensions["G"].width = 18

        wb.save(output)
        output.seek(0)
        return output.getvalue()

    except Exception as e:
        import traceback
        st.error(f"Error al procesar el archivo: {str(e)}")
        print(traceback.format_exc())
        return None
