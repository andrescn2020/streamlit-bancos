import streamlit as st
import PyPDF2
import re
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule

ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

def clean_for_excel(text):
    if not text: return ""
    text = str(text)
    text = ILLEGAL_CHARACTERS_RE.sub("", text)
    return text.strip()

def parse_monto(s):
    """Convierte '1.234,56' o '-1.234,56' a float"""
    if not s: return 0.0
    s = s.strip()
    try:
        limpio = s.replace(".", "").replace(",", ".")
        return float(limpio)
    except:
        return 0.0

def _detectar_umbral(lineas):
    """Busca linea header limpia (FECHA ... DEBITOS ... CREDITOS) para calibrar posiciones."""
    for l in lineas:
        l_upper = l.upper()
        # Solo considerar headers que empiecen con FECHA (no los que tienen DETALLE DE MOVIMIENTO prefix)
        stripped = l.strip().upper()
        if stripped.startswith("FECHA") and "DEBITOS" in l_upper and "CREDITOS" in l_upper:
            i_deb = l_upper.find("DEBITOS")
            i_cred = l_upper.find("CREDITOS")
            if i_deb != -1 and i_cred != -1 and i_cred > i_deb:
                fin_deb = i_deb + len("DEBITOS")
                fin_cred = i_cred + len("CREDITOS")
                return (fin_deb + fin_cred) // 2
    return 90  # Fallback

def _split_lineas_fusionadas(lineas):
    """Separa líneas que tienen 2+ movimientos pegados por la extracción PDF."""
    resultado = []
    re_fecha_interna = re.compile(r'(\s{2,})(\d{2}/\d{2}/\d{2}\s)')
    
    for linea in lineas:
        # Buscar fechas internas (no al inicio de la línea)
        # Una fecha interna es precedida por 2+ espacios y aparece después de posición 20
        partes = []
        pos = 0
        for m in re.finditer(r'\s{2,}\d{2}/\d{2}/\d{2}\s', linea):
            start = m.start()
            if start < 20:  # La primera fecha puede empezar cerca del inicio
                continue
            # Encontrar donde empieza la fecha dentro del match
            fecha_start = m.start() + len(m.group()) - len(m.group().lstrip())
            # Buscar posición real de la fecha
            fecha_match = re.search(r'\d{2}/\d{2}/\d{2}', m.group())
            if fecha_match:
                real_start = m.start() + fecha_match.start()
                if real_start > 20:  # No es la primera fecha de la línea
                    partes.append(linea[pos:real_start].rstrip())
                    pos = real_start
        
        partes.append(linea[pos:])
        resultado.extend([p for p in partes if p.strip()])
    
    return resultado

def _nombre_hoja(nombre_cuenta, idx):
    """Genera un nombre de hoja Excel válido (max 31 chars)."""
    nombre = nombre_cuenta.upper()
    if "DOLAR" in nombre:
        short = "CC Dolares"
    elif "ESPECIAL" in nombre and "PESOS" in nombre:
        short = "CC Esp Pesos"
    elif "BANCARIA" in nombre:
        short = "CC Bancaria"
    elif "PESOS" in nombre:
        short = "CC Pesos"
    else:
        short = f"Cuenta {idx+1}"
    
    # Asegurar max 31 chars
    return short[:31]

def _crear_hoja_cuenta(wb, nombre_hoja, titular, periodo, saldo_ini, saldo_fin, movimientos_df):
    """Crea una hoja con layout dashboard para una cuenta."""
    ws = wb.create_sheet(title=nombre_hoja)
    ws.sheet_view.showGridLines = False

    color_bg_main = "003366"
    color_txt_main = "FFFFFF"

    thin_border = Border(left=Side(style='thin', color="A6A6A6"),
                         right=Side(style='thin', color="A6A6A6"),
                         top=Side(style='thin', color="A6A6A6"),
                         bottom=Side(style='thin', color="A6A6A6"))

    fill_head_deb = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    fill_col_deb = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
    fill_row_deb = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")
    fill_head_cred = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    fill_col_cred = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")
    fill_row_cred = PatternFill(start_color="F2F9F1", end_color="F2F9F1", fill_type="solid")

    df = movimientos_df
    if not df.empty:
        creditos = df[df["Importe"] > 0].copy()
        debitos = df[df["Importe"] < 0].copy()
        debitos["Importe"] = debitos["Importe"].abs()
    else:
        creditos = pd.DataFrame(columns=["Fecha", "Descripcion", "Importe"])
        debitos = pd.DataFrame(columns=["Fecha", "Descripcion", "Importe"])

    # Header
    ws.merge_cells("A1:G1")
    tit = ws["A1"]
    tit.value = f"REPORTE MACRO - {clean_for_excel(titular)} - {nombre_hoja}"
    tit.font = Font(size=14, bold=True, color=color_txt_main)
    tit.fill = PatternFill(start_color=color_bg_main, end_color=color_bg_main, fill_type="solid")
    tit.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25

    # Saldos
    ws["A3"] = "SALDO INICIAL"
    ws["A3"].font = Font(bold=True, size=10, color="666666")
    ws["B3"] = saldo_ini
    ws["B3"].number_format = '"$ "#,##0.00'
    ws["B3"].font = Font(bold=True, size=11)
    ws["B3"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

    ws["A4"] = "SALDO FINAL"
    ws["A4"].font = Font(bold=True, size=10, color="666666")
    ws["B4"] = saldo_fin
    ws["B4"].number_format = '"$ "#,##0.00'
    ws["B4"].font = Font(bold=True, size=11)
    ws["B4"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

    ws["D3"] = "TITULAR"
    ws["D3"].alignment = Alignment(horizontal='right')
    ws["D3"].font = Font(bold=True, color="666666", size=10)
    ws["E3"] = clean_for_excel(titular)
    ws["E3"].font = Font(bold=True, size=11)
    ws["E3"].alignment = Alignment(horizontal='center')
    ws.merge_cells("E3:G3")
    for c in ["E","F","G"]: ws[f"{c}3"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

    ws["D4"] = "PERÍODO"
    ws["D4"].alignment = Alignment(horizontal='right')
    ws["D4"].font = Font(bold=True, color="666666", size=10)
    ws["E4"] = clean_for_excel(periodo)
    ws["E4"].font = Font(bold=True, size=11)
    ws["E4"].alignment = Alignment(horizontal='center')
    ws.merge_cells("E4:G4")
    for c in ["E","F","G"]: ws[f"{c}4"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

    ws["D6"] = "CONTROL DE SALDOS"
    ws["D6"].font = Font(bold=True, size=10, color="666666")
    ws["D6"].alignment = Alignment(horizontal='center')
    cell_ctl = ws["D7"]
    cell_ctl.font = Font(bold=True, size=12)
    cell_ctl.alignment = Alignment(horizontal='center')
    cell_ctl.border = thin_border

    # Tablas
    fila_inicio = 10
    f_header = fila_inicio

    ws.merge_cells(f"A{f_header}:C{f_header}")
    ws[f"A{f_header}"] = "CRÉDITOS"
    ws[f"A{f_header}"].fill = fill_head_cred
    ws[f"A{f_header}"].font = Font(bold=True, color="FFFFFF")
    ws[f"A{f_header}"].alignment = Alignment(horizontal='center')
    ws[f"A{f_header}"].border = thin_border

    headers = ["Fecha", "Descripción", "Importe"]
    cols_cred = ["A", "B", "C"]
    f_sub = f_header + 1
    for i, h in enumerate(headers):
        c = ws[f"{cols_cred[i]}{f_sub}"]
        c.value = h
        c.fill = fill_col_cred
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal='center')
        c.border = thin_border

    ws.merge_cells(f"E{f_header}:G{f_header}")
    ws[f"E{f_header}"] = "DÉBITOS"
    ws[f"E{f_header}"].fill = fill_head_deb
    ws[f"E{f_header}"].font = Font(bold=True, color="FFFFFF")
    ws[f"E{f_header}"].alignment = Alignment(horizontal='center')
    ws[f"E{f_header}"].border = thin_border

    cols_deb = ["E", "F", "G"]
    for i, h in enumerate(headers):
        c = ws[f"{cols_deb[i]}{f_sub}"]
        c.value = h
        c.fill = fill_col_deb
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal='center')
        c.border = thin_border

    fila_dato_start = f_sub + 1

    # Créditos
    f_cred = fila_dato_start
    if creditos.empty:
        ws.merge_cells(f"A{f_cred}:C{f_cred}")
        ws[f"A{f_cred}"] = "SIN MOVIMIENTOS"
        ws[f"A{f_cred}"].font = Font(italic=True, color="666666")
        ws[f"A{f_cred}"].alignment = Alignment(horizontal='center')
        ws[f"A{f_cred}"].border = thin_border
        f_cred += 1
    else:
        start_c = f_cred
        for _, r in creditos.iterrows():
            ws[f"A{f_cred}"] = clean_for_excel(r["Fecha"])
            ws[f"A{f_cred}"].fill = fill_row_cred
            ws[f"A{f_cred}"].alignment = Alignment(horizontal='center')
            ws[f"A{f_cred}"].border = thin_border
            ws[f"B{f_cred}"] = clean_for_excel(r["Descripcion"])
            ws[f"B{f_cred}"].fill = fill_row_cred
            ws[f"B{f_cred}"].border = thin_border
            ws[f"C{f_cred}"] = r["Importe"]
            ws[f"C{f_cred}"].number_format = '"$ "#,##0.00'
            ws[f"C{f_cred}"].fill = fill_row_cred
            ws[f"C{f_cred}"].border = thin_border
            f_cred += 1
        ws.merge_cells(f"A{f_cred}:B{f_cred}")
        ws[f"A{f_cred}"] = "TOTAL CRÉDITOS"
        ws[f"A{f_cred}"].font = Font(bold=True)
        ws[f"A{f_cred}"].alignment = Alignment(horizontal='right')
        ws[f"A{f_cred}"].fill = fill_col_cred
        ws[f"A{f_cred}"].border = thin_border
        ws[f"C{f_cred}"] = f"=SUM(C{start_c}:C{f_cred-1})"
        ws[f"C{f_cred}"].number_format = '"$ "#,##0.00'
        ws[f"C{f_cred}"].font = Font(bold=True)
        ws[f"C{f_cred}"].fill = fill_col_cred
        ws[f"C{f_cred}"].border = thin_border
        f_cred += 1

    # Débitos
    f_deb = fila_dato_start
    if debitos.empty:
        ws.merge_cells(f"E{f_deb}:G{f_deb}")
        ws[f"E{f_deb}"] = "SIN MOVIMIENTOS"
        ws[f"E{f_deb}"].font = Font(italic=True, color="666666")
        ws[f"E{f_deb}"].alignment = Alignment(horizontal='center')
        ws[f"E{f_deb}"].border = thin_border
        f_deb += 1
    else:
        start_d = f_deb
        for _, r in debitos.iterrows():
            ws[f"E{f_deb}"] = clean_for_excel(r["Fecha"])
            ws[f"E{f_deb}"].fill = fill_row_deb
            ws[f"E{f_deb}"].alignment = Alignment(horizontal='center')
            ws[f"E{f_deb}"].border = thin_border
            ws[f"F{f_deb}"] = clean_for_excel(r["Descripcion"])
            ws[f"F{f_deb}"].fill = fill_row_deb
            ws[f"F{f_deb}"].border = thin_border
            ws[f"G{f_deb}"] = r["Importe"]
            ws[f"G{f_deb}"].number_format = '"$ "#,##0.00'
            ws[f"G{f_deb}"].fill = fill_row_deb
            ws[f"G{f_deb}"].border = thin_border
            f_deb += 1
        ws.merge_cells(f"E{f_deb}:F{f_deb}")
        ws[f"E{f_deb}"] = "TOTAL DÉBITOS"
        ws[f"E{f_deb}"].font = Font(bold=True)
        ws[f"E{f_deb}"].alignment = Alignment(horizontal='right')
        ws[f"E{f_deb}"].fill = fill_col_deb
        ws[f"E{f_deb}"].border = thin_border
        ws[f"G{f_deb}"] = f"=SUM(G{start_d}:G{f_deb-1})"
        ws[f"G{f_deb}"].number_format = '"$ "#,##0.00'
        ws[f"G{f_deb}"].font = Font(bold=True)
        ws[f"G{f_deb}"].fill = fill_col_deb
        ws[f"G{f_deb}"].border = thin_border
        f_deb += 1

    # Control de Saldos
    f_tot_cred = f"C{f_cred-1}" if not creditos.empty else "0"
    f_tot_deb = f"G{f_deb-1}" if not debitos.empty else "0"
    ws["D7"] = f"=ROUND(B3+{f_tot_cred}-{f_tot_deb}-B4, 2)"
    ws["D7"].number_format = '"$ "#,##0.00'

    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    red_font = Font(color='9C0006', bold=True)
    ws.conditional_formatting.add('D7', CellIsRule(operator='notEqual', formula=['0'], stopIfTrue=True, fill=red_fill, font=red_font))

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 40
    ws.column_dimensions["G"].width = 18


def procesar_macro_formato_3(archivo_pdf):
    """Procesa archivos PDF del Banco Macro - Formato Multi-Cuenta (Extracto Sucursal).
    Genera una hoja por cada cuenta encontrada."""
    st.info("Procesando archivo del Banco Macro (Formato 3 - Multi-Cuenta)...")
    try:
        archivo_pdf.seek(0)
        reader = PyPDF2.PdfReader(io.BytesIO(archivo_pdf.read()))
        texto = "".join(page.extract_text() + "\n" for page in reader.pages)
        texto = texto.replace('\x00', '')
        
        lineas_raw = texto.splitlines()
        
        # Pre-procesamiento: separar líneas fusionadas (2 movimientos en 1 línea)
        lineas = _split_lineas_fusionadas(lineas_raw)
        
        # === METADATOS ===
        titular = "Sin Especificar"
        periodo = "Sin Especificar"
        
        for l in lineas_raw[:20]:
            match_tit = re.search(r'C\.U\.I\.T\s+\d+\s+(.*)', l)
            if match_tit:
                titular = match_tit.group(1).strip()
                break
        
        for l in lineas_raw[:20]:
            match_per = re.search(r'Per[ií]odo\s+del\s+Extracto:\s*(\d{2}/\d{2}/\d{4})\s+al\s+(\d{2}/\d{2}/\d{4})', l, re.IGNORECASE)
            if match_per:
                periodo = f"Del {match_per.group(1)} al {match_per.group(2)}"
                break
        
        # === UMBRAL DINÁMICO ===
        umbral = _detectar_umbral(lineas)
        
        # === PARSEO POR CUENTAS ===
        re_cuenta_header = re.compile(r'(CUENTA\s+CORRIENTE.*?)NRO\.:\s*(\S+)', re.IGNORECASE)
        re_fecha = re.compile(r'^\s*(\d{2}/\d{2}/\d{2})\s+(.*)')
        re_monto = re.compile(r'-?\d{1,3}(?:\.\d{3})*,\d{2}')
        
        # Diccionario de cuentas: {nro: {nombre, saldo_ini, saldo_fin, movimientos}}
        cuentas = {}
        orden_cuentas = []  # Para preservar orden de aparición
        
        cuenta_actual_nro = None
        
        for i, linea in enumerate(lineas):
            l_upper = linea.upper().strip()
            
            # Detectar inicio de sección de cuenta
            match_cta = re_cuenta_header.search(linea)
            if match_cta:
                nombre_cta = match_cta.group(1).strip()
                nro_cta = match_cta.group(2).strip()
                cuenta_actual_nro = nro_cta
                
                # Inicializar cuenta si es la primera vez que la vemos
                if nro_cta not in cuentas:
                    cuentas[nro_cta] = {
                        "nombre": nombre_cta + " NRO.: " + nro_cta,
                        "nombre_corto": nombre_cta,
                        "saldo_ini": 0.0,
                        "saldo_fin": 0.0,
                        "saldo_ini_set": False,
                        "saldo_fin_set": False,
                        "movimientos": []
                    }
                    orden_cuentas.append(nro_cta)
                continue
            
            if not cuenta_actual_nro:
                continue
            
            cta = cuentas[cuenta_actual_nro]
            
            # Líneas a ignorar
            if not l_upper:
                continue
            if "DETALLE DE MOVIMIENTO" in l_upper:
                continue
            if l_upper.startswith("FECHA") and "DESCRIPCION" in l_upper:
                continue
            if "CLAVE BANCARIA" in l_upper:
                continue
            if "TASA NOM" in l_upper:
                continue
            if "INFORMACION DE SU" in l_upper:
                continue
            if "RESUMEN GENERAL" in l_upper:
                cuenta_actual_nro = None
                continue
            if "SALDOS CONSOLIDADOS" in l_upper:
                continue
            if "HOJA NRO" in l_upper:
                continue
            if "TIPO CUENTA" in l_upper:
                continue
            if "SUCURSAL" in l_upper and "MONEDA" in l_upper:
                continue
            if re.match(r'^\s*-\s+-\s+-', linea):
                cuenta_actual_nro = None  # Separador = fin de esta sección
                continue
            
            # Footer legal → fin de procesamiento de cuenta
            if "LOS DEPOSITOS EN PESOS" in l_upper or "LOS DEPÓSITOS EN PESOS" in l_upper:
                cuenta_actual_nro = None
                continue
            
            # Info fiscal (no es movimiento)
            if "TOTAL COBRADO" in l_upper or "D. 409" in l_upper or \
               "ESTIMADO CLIENTE" in l_upper or "IIBB SIRCREB" in l_upper or \
               "LE HABILITEN" in l_upper:
                continue
            
            # Saldo Anterior
            if "SALDO ULTIMO EXTRACTO" in l_upper:
                if not cta["saldo_ini_set"]:
                    montos = re_monto.findall(linea)
                    if montos:
                        val = parse_monto(montos[-1])
                        monto_str = montos[-1]
                        idx_m = linea.rfind(monto_str)
                        prefix = linea[:idx_m].rstrip()
                        if prefix.endswith('-'):
                            val = -abs(val)
                        cta["saldo_ini"] = val
                        cta["saldo_ini_set"] = True
                continue
            
            # Saldo Final
            if "SALDO FINAL" in l_upper:
                # Siempre tomar el último (por si aparece en varias páginas)
                montos = re_monto.findall(linea)
                if montos:
                    val = parse_monto(montos[-1])
                    monto_str = montos[-1]
                    idx_m = linea.rfind(monto_str)
                    prefix = linea[:idx_m].rstrip()
                    if prefix.endswith('-'):
                        val = -abs(val)
                    cta["saldo_fin"] = val
                    cta["saldo_fin_set"] = True
                continue
            
            # Movimientos: fecha dd/mm/yy
            match_mov = re_fecha.match(linea)
            if match_mov:
                fecha = match_mov.group(1)
                
                montos = re_monto.findall(linea)
                if not montos:
                    continue
                
                # Descripción
                desc = match_mov.group(2)
                for m in montos:
                    desc = desc.replace(m, "", 1)
                desc = re.sub(r'\s+0\s*$', '', desc).strip()
                desc = re.sub(r'\s{2,}', ' ', desc).strip()
                
                # Primer monto = importe, determinar signo por posición
                primer_monto_str = montos[0]
                idx_fin_monto = linea.find(primer_monto_str) + len(primer_monto_str)
                
                importe = parse_monto(primer_monto_str)
                
                if idx_fin_monto <= umbral:
                    importe = -abs(importe)  # Débito
                else:
                    importe = abs(importe)   # Crédito
                
                cta["movimientos"].append({
                    "Fecha": fecha,
                    "Descripcion": desc,
                    "Importe": importe
                })
        
        # === GENERAR EXCEL ===
        output = io.BytesIO()
        wb = Workbook()
        # Eliminar hoja por defecto
        wb.remove(wb.active)
        
        hojas_creadas = 0
        nombres_usados = set()
        
        for idx, nro_cta in enumerate(orden_cuentas):
            cta = cuentas[nro_cta]
            
            # Generar nombre de hoja único
            nombre_h = _nombre_hoja(cta["nombre_corto"], idx)
            base = nombre_h
            counter = 2
            while nombre_h in nombres_usados:
                nombre_h = f"{base} {counter}"
                counter += 1
            nombres_usados.add(nombre_h)
            
            df = pd.DataFrame(cta["movimientos"]) if cta["movimientos"] else pd.DataFrame(columns=["Fecha", "Descripcion", "Importe"])
            
            _crear_hoja_cuenta(
                wb, 
                nombre_h, 
                titular, 
                periodo, 
                cta["saldo_ini"], 
                cta["saldo_fin"], 
                df
            )
            hojas_creadas += 1
        
        if hojas_creadas == 0:
            st.warning("No se encontraron cuentas en el PDF")
            return None
        
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    except Exception as e:
        import traceback
        st.error(f"Error al procesar el archivo Macro F3: {str(e)}")
        st.error(traceback.format_exc())
        return None
