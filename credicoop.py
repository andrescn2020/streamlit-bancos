import streamlit as st
import re
import pandas as pd
import PyPDF2
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule

# Regex para caracteres ilegales en Excel (ASCII Control characters excepto \t, \n, \r)
ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

def clean_for_excel(text):
    """Elimina caracteres ilegales para Excel y espacios extra"""
    if not text: return ""
    text = str(text)
    text = ILLEGAL_CHARACTERS_RE.sub("", text)
    return text.strip()

def convertir_a_numerico(valor_str):
    """Convierte string de moneda '1.234,56' a float"""
    if not valor_str: return 0.0
    try:
        # Manejo de caracteres extraños como − (U+2212)
        valor_str = valor_str.replace("−", "-") 
        limpio = valor_str.replace(".", "").replace(",", ".")
        return float(limpio)
    except:
        return 0.0

def procesar_credicoop(archivo_pdf):
    st.info("Procesando archivo Credicoop (Formato Estandarizado)...")
    try:
        archivo_pdf.seek(0)
        reader = PyPDF2.PdfReader(io.BytesIO(archivo_pdf.read()))
        texto = "".join(page.extract_text() + "\n" for page in reader.pages)
        texto = texto.replace('\x00', '')
        
        # Limpieza específica de basura intercalada (cubre CONTINUA/CONTINUAR y PÁGINA/PAGINA)
        texto = re.sub(r"(?i)CONTINUA(R)?\s+EN\s+P(A|Á)GINA\s+SIGUIENTE.*", "", texto)
        
        # Limpieza footer Credicoop infiltrado
        texto = re.sub(r"(?i)Banco\s+Credicoop\s+Cooperativo\s+Limitado.*", "", texto)
        texto = re.sub(r"(?i)Ctro\.\s+de\s+Contacto\s+Telefonico.*", "", texto)
        # Limpieza brutal de residuos Credicoop
        texto = re.sub(r"(?i)Calidad\s+de\s+Servicios.*", "", texto)
        texto = re.sub(r"(?i)Sitio\s+de\s+Internet.*", "", texto)
        texto = re.sub(r"(?i).*@bancocredicoop\.coop.*", "", texto)
        texto = re.sub(r"(?i).*www\.bancocredicoop\.coop.*", "", texto)
        
        lineas = texto.splitlines()
        
        nombre_titular = None
        periodo = None
        saldo_inicial = None
        saldo_final = None 
        ultimo_saldo_acumulado = None
        movimientos = []
        
        # Regex corregido y mejorado para capturar Código Opcional
        # Grupo 1: Fecha, Grupo 2: Código (opcional), Grupo 3: Resto
        regex_mov = re.compile(r"^\s*(\d{2}/\d{2}/\d{2})\s+(?:(\d+)\s+)?(.*)")

        # Detección de Umbral Dinámico basada en encabezados
        idx_umbral = 95 # Valor fallback
        for l in lineas[:20]:
            l_upper = l.upper()
            if "DEBITO" in l_upper and "CREDITO" in l_upper:
                try:
                    i_deb = l_upper.find("DEBITO")
                    i_cred = l_upper.find("CREDITO")
                    # El umbral está entre el fin de DEBITO y el inicio de CREDITO
                    # DEBITO suele estar centrado sobre su columna, CREDITO igual.
                    # Punto medio simple:
                    if i_deb != -1 and i_cred != -1 and i_cred > i_deb:
                        # Usamos el final de las palabras para alinear mejor con montos a la derecha
                        fin_deb = i_deb + len("DEBITO")
                        fin_cred = i_cred + len("CREDITO")
                        idx_umbral = (fin_deb + fin_cred) // 2
                        print(f"Umbral dinámico (Fin Monto): {idx_umbral}")
                        break
                except:
                    pass

        i = 0
        while i < len(lineas):
            linea = lineas[i]
            linea_strip = linea.strip()
            
            if not linea_strip:
                i += 1
                continue

            # 1. Titular
            if i == 1 and not nombre_titular:
                parts = re.split(r"\s{4,}", linea_strip)
                nombre_titular = parts[0]

            # 2. Período
            if "Resumen:" in linea and "del:" in linea:
                match_per = re.search(r"del:\s*(\d{2}/\d{2}/\d{4})\s+al:\s*(\d{2}/\d{2}/\d{4})", linea)
                if match_per:
                    periodo = f"Del {match_per.group(1)} al {match_per.group(2)}"

            # 3. Saldo Anterior
            if "SALDO ANTERIOR" in linea:
                match_saldo = re.findall(r"([\d\.,]+)", linea)
                if match_saldo:
                    saldo_inicial = match_saldo[-1]

            # DETECCIÓN DE FIN DE MOVIMIENTOS (Saldo final o inicio de tablas resumen)
            l_upper = linea.upper()
            # Simplificamos las condiciones para ser más agresivos con el footer
            if (re.search(r"SALDO\s+AL", l_upper) and "IMPUESTO" not in l_upper) or \
               "TOTAL IMPUESTO" in l_upper or \
               "PERCIBIDO" in l_upper or \
               "CRE FISC" in l_upper or \
               "INFORMACION ADICIONAL" in l_upper or \
               "DETALLE DE TRANSFERENCIAS" in l_upper or \
               "VIENE DE PAGINA" in l_upper or \
               "DENOMINACION" in l_upper:
                
                # Intentar extraer saldo si es la línea de saldo
                if "SALDO" in l_upper and "PERCIBIDO" not in l_upper:
                    match_saldo_final = re.findall(r"([\d\.,]+)", linea)
                    if match_saldo_final:
                        posibles = [x for x in match_saldo_final if ',' in x]
                        if posibles:
                            saldo_final = posibles[-1]
                
                # print(f"DEBUG BREAK: Cortando por footer en '{linea.strip()}'")
                break # TERMINAR DEFINTIVAMENTE

            # 4. Movimientos
            match_mov = regex_mov.match(linea) # Usar linea original para regex que maneja espacios
            if match_mov:
                fecha = match_mov.group(1)
                codigo_mov = match_mov.group(2) # Puede ser None
                resto = match_mov.group(3)
                
                tokens = re.split(r"\s{2,}", resto)
                
                montos_candidatos = []
                descripcion_parts = []
                
                for token in tokens:
                    # Regex estricto moneda: debe tener coma decimal (evita CUITs)
                    t_clean = token.strip().replace("−", "-")
                    if re.match(r"^-?(\d{1,3}(\.\d{3})*|\d+),\d{2}$", t_clean):
                         montos_candidatos.append(t_clean)
                    else:
                        descripcion_parts.append(token.strip())
                
                descripcion = " ".join(descripcion_parts)
                
                # Filtro anti-footer por contenido (evitar tablas finales que empiecen con fecha)
                desc_upper = descripcion.upper()
                if "PERCIBIDO" in desc_upper or \
                   "TOTAL" in desc_upper or \
                   "INFORMACION ADICIONAL" in desc_upper:
                    i += 1 # IMPORTANTE: Avanzar índice antes de continue para evitar loop infinito
                    continue 

                importe = 0.0
                
                if montos_candidatos:
                    monto_str = montos_candidatos[0]
                    # Si hay más de un monto, el último suele ser el saldo acumulado de la línea
                    if len(montos_candidatos) >= 2:
                        ultimo_saldo_acumulado = montos_candidatos[-1]
                    
                    idx_monto = linea.rfind(monto_str)
                    idx_fin = idx_monto + len(monto_str)
                    val = convertir_a_numerico(monto_str)
                    
                    # Heurística de posición DINÁMICA basada en ALINEACIÓN DERECHA (Final del monto)
                    if idx_fin < idx_umbral: 
                        importe = -abs(val) # Débito
                    else:
                        importe = abs(val) # Crédito
                    
                    # Descripción extra
                    desc_extra = ""
                    j = i + 1
                    while j < len(lineas):
                        l_next = lineas[j]
                        if not regex_mov.match(l_next) and re.match(r"^\s{10,}", l_next):
                            desc_extra += " " + l_next.strip()
                            j += 1
                        else:
                            break
                    
                    descripcion += desc_extra
                    i = j - 1
                    
                    # Limpieza FINAL de descripción (por si se pegó SALDO AL de la línea siguiente)
                    descripcion = re.sub(r"\s*SALDO\s+AL.*", "", descripcion, flags=re.IGNORECASE).strip()
                    
                    movimientos.append({
                        "Fecha": fecha,
                        "Descripcion": descripcion,
                        "Importe": importe
                    })


            i += 1
        

        if saldo_final is None and ultimo_saldo_acumulado:
             saldo_final = ultimo_saldo_acumulado
            
        if saldo_final is None and saldo_inicial:
            saldo_calc = convertir_a_numerico(saldo_inicial)
            for m in movimientos:
                saldo_calc += m["Importe"]
            saldo_final = saldo_calc 

        # --- EXCEL ---
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Movimientos"
        ws.sheet_view.showGridLines = False

        thin_border = Border(left=Side(style='thin', color="A6A6A6"), 
                             right=Side(style='thin', color="A6A6A6"), 
                             top=Side(style='thin', color="A6A6A6"), 
                             bottom=Side(style='thin', color="A6A6A6"))
        
        color_bg_main = "2C3E50"
        color_txt_main = "FFFFFF"
        
        fill_head_deb = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        fill_col_deb = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
        fill_row_deb = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")

        fill_head_cred = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        fill_col_cred = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")
        fill_row_cred = PatternFill(start_color="F2F9F1", end_color="F2F9F1", fill_type="solid")

        ws.merge_cells("A1:G1")
        tit = ws["A1"]
        clean_tit = clean_for_excel(nombre_titular) or 'Desconocido'
        tit.value = f"REPORTE CREDICOOP - {clean_tit}"
        tit.font = Font(size=14, bold=True, color=color_txt_main)
        tit.fill = PatternFill(start_color=color_bg_main, end_color=color_bg_main, fill_type="solid")
        tit.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25

        ws["A3"] = "SALDO INICIAL"
        ws["A3"].font = Font(bold=True, size=10, color="666666")
        ws["B3"] = convertir_a_numerico(saldo_inicial) if saldo_inicial else 0
        ws["B3"].number_format = '"$ "#,##0.00'
        ws["B3"].font = Font(bold=True, size=11)
        ws["B3"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

        ws["A4"] = "SALDO FINAL"
        ws["A4"].font = Font(bold=True, size=10, color="666666")
        ws["B4"] = convertir_a_numerico(saldo_final) if saldo_final else 0
        ws["B4"].number_format = '"$ "#,##0.00'
        ws["B4"].font = Font(bold=True, size=11)
        ws["B4"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

        ws["D3"] = "TITULAR"
        ws["D3"].alignment = Alignment(horizontal='right')
        ws["D3"].font = Font(bold=True, color="666666", size=10)
        
        ws["E3"] = clean_for_excel(nombre_titular)
        ws["E3"].font = Font(bold=True, size=11)
        ws["E3"].alignment = Alignment(horizontal='center')
        ws.merge_cells("E3:G3")
        for c in ["E","F","G"]: ws[f"{c}3"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

        ws["D4"] = "PERÍODO"
        ws["D4"].alignment = Alignment(horizontal='right')
        ws["D4"].font = Font(bold=True, color="666666", size=10)
        
        ws["E4"] = clean_for_excel(periodo)
        ws["E4"].font = Font(bold=True, size=11)
        ws["E4"].alignment = Alignment(horizontal='center')
        ws.merge_cells("E4:G4")
        for c in ["E","F","G"]: ws[f"{c}4"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

        ws["D6"] = "CONTROL DE SALDOS"
        ws["D6"].font = Font(bold=True, size=10, color="666666")
        ws["D6"].alignment = Alignment(horizontal='center')
        
        cell_ctl = ws["D7"]
        cell_ctl.font = Font(bold=True, size=12)
        cell_ctl.alignment = Alignment(horizontal='center')
        cell_ctl.border = thin_border
        
        fila_inicio = 10
        df = pd.DataFrame(movimientos)
        if not df.empty:
            creditos = df[df["Importe"] > 0].copy()
            debitos = df[df["Importe"] < 0].copy()
            debitos["Importe"] = debitos["Importe"].abs()
        else:
            creditos = pd.DataFrame(columns=["Fecha", "Descripcion", "Importe"])
            debitos = pd.DataFrame(columns=["Fecha", "Descripcion", "Importe"])

        # CRÉDITOS (A-C)
        f_header = fila_inicio
        ws.merge_cells(f"A{f_header}:C{f_header}")
        ws[f"A{f_header}"] = "CRÉDITOS" 
        ws[f"A{f_header}"].fill = fill_head_cred
        ws[f"A{f_header}"].font = Font(bold=True, color="FFFFFF")
        ws[f"A{f_header}"].alignment = Alignment(horizontal='center')
        ws[f"A{f_header}"].border = thin_border
        
        headers = ["Fecha", "Descripción", "Importe"]
        cols_cred = ["A", "B", "C"]
        f_sub = f_header + 1
        for i, h in enumerate(headers):
            c = ws[f"{cols_cred[i]}{f_sub}"]
            c.value = h
            c.fill = fill_col_cred
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal='center')
            c.border = thin_border
        
        # DÉBITOS (E-G)
        ws.merge_cells(f"E{f_header}:G{f_header}")
        ws[f"E{f_header}"] = "DÉBITOS" 
        ws[f"E{f_header}"].fill = fill_head_deb
        ws[f"E{f_header}"].font = Font(bold=True, color="FFFFFF")
        ws[f"E{f_header}"].alignment = Alignment(horizontal='center')
        ws[f"E{f_header}"].border = thin_border
        
        cols_deb = ["E", "F", "G"]
        for i, h in enumerate(headers):
            c = ws[f"{cols_deb[i]}{f_sub}"]
            c.value = h
            c.fill = fill_col_deb
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal='center')
            c.border = thin_border
            
        # --- LLENADO DE DATOS (PARALELO) ---
        fila_dato_start = f_sub + 1
        
        # 1. CRÉDITOS
        f_cred = fila_dato_start
        if creditos.empty:
            ws.merge_cells(f"A{f_cred}:C{f_cred}")
            ws[f"A{f_cred}"] = "SIN MOVIMIENTOS"
            ws[f"A{f_cred}"].font = Font(italic=True, color="666666")
            ws[f"A{f_cred}"].alignment = Alignment(horizontal='center')
            ws[f"A{f_cred}"].border = thin_border
            f_cred += 1
        else:
            start_c = f_cred
            for _, r in creditos.iterrows():
                ws[f"A{f_cred}"] = r["Fecha"]
                ws[f"A{f_cred}"].fill = fill_row_cred
                ws[f"A{f_cred}"].alignment = Alignment(horizontal='center')
                ws[f"A{f_cred}"].border = thin_border

                ws[f"B{f_cred}"] = clean_for_excel(r["Descripcion"])
                ws[f"B{f_cred}"].fill = fill_row_cred
                ws[f"B{f_cred}"].border = thin_border

                ws[f"C{f_cred}"] = r["Importe"]
                ws[f"C{f_cred}"].number_format = '"$ "#,##0.00'
                ws[f"C{f_cred}"].fill = fill_row_cred
                ws[f"C{f_cred}"].border = thin_border
                f_cred += 1
            
            # Total Créditos
            ws.merge_cells(f"A{f_cred}:B{f_cred}")
            ws[f"A{f_cred}"] = "TOTAL CRÉDITOS"
            ws[f"A{f_cred}"].font = Font(bold=True)
            ws[f"A{f_cred}"].alignment = Alignment(horizontal='right')
            ws[f"A{f_cred}"].fill = fill_col_cred
            ws[f"A{f_cred}"].border = thin_border
            
            ws[f"C{f_cred}"] = f"=SUM(C{start_c}:C{f_cred-1})"
            ws[f"C{f_cred}"].number_format = '"$ "#,##0.00'
            ws[f"C{f_cred}"].font = Font(bold=True)
            ws[f"C{f_cred}"].fill = fill_col_cred
            ws[f"C{f_cred}"].border = thin_border
            f_cred += 1

        # 2. DÉBITOS
        f_deb = fila_dato_start
        if debitos.empty:
            ws.merge_cells(f"E{f_deb}:G{f_deb}")
            ws[f"E{f_deb}"] = "SIN MOVIMIENTOS"
            ws[f"E{f_deb}"].font = Font(italic=True, color="666666")
            ws[f"E{f_deb}"].alignment = Alignment(horizontal='center')
            ws[f"E{f_deb}"].border = thin_border
            f_deb += 1
        else:
            start_d = f_deb
            for _, r in debitos.iterrows():
                ws[f"E{f_deb}"] = r["Fecha"]
                ws[f"E{f_deb}"].fill = fill_row_deb
                ws[f"E{f_deb}"].alignment = Alignment(horizontal='center')
                ws[f"E{f_deb}"].border = thin_border

                ws[f"F{f_deb}"] = clean_for_excel(r["Descripcion"])
                ws[f"F{f_deb}"].fill = fill_row_deb
                ws[f"F{f_deb}"].border = thin_border

                ws[f"G{f_deb}"] = r["Importe"]
                ws[f"G{f_deb}"].number_format = '"$ "#,##0.00'
                ws[f"G{f_deb}"].fill = fill_row_deb
                ws[f"G{f_deb}"].border = thin_border
                f_deb += 1
            
            # Total Débitos
            ws.merge_cells(f"E{f_deb}:F{f_deb}")
            ws[f"E{f_deb}"] = "TOTAL DÉBITOS"
            ws[f"E{f_deb}"].font = Font(bold=True)
            ws[f"E{f_deb}"].alignment = Alignment(horizontal='right')
            ws[f"E{f_deb}"].fill = fill_col_deb
            ws[f"E{f_deb}"].border = thin_border
            
            ws[f"G{f_deb}"] = f"=SUM(G{start_d}:G{f_deb-1})"
            ws[f"G{f_deb}"].number_format = '"$ "#,##0.00'
            ws[f"G{f_deb}"].font = Font(bold=True)
            ws[f"G{f_deb}"].fill = fill_col_deb
            ws[f"G{f_deb}"].border = thin_border
            f_deb += 1

        f_ini = "B3"
        f_tot_cred = f"C{f_cred-1}" if not creditos.empty else "0"
        f_tot_deb = f"G{f_deb-1}" if not debitos.empty else "0"
        f_fin = "B4"
        
        ws["D7"] = f"=ROUND({f_ini}+{f_tot_cred}-{f_tot_deb}-{f_fin}, 2)"
        ws["D7"].number_format = '"$ "#,##0.00'
        
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        red_font = Font(color='9C0006', bold=True)
        ws.conditional_formatting.add('D7', CellIsRule(operator='notEqual', formula=['0'], stopIfTrue=True, fill=red_fill, font=red_font))

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 25
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 40
        ws.column_dimensions["G"].width = 18

        wb.save(output)
        output.seek(0)
        return output.getvalue()

    except Exception as e:
        import traceback
        st.error(f"Error al procesar el archivo Credicoop: {str(e)}")
        st.error(traceback.format_exc())
        return None
