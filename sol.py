import streamlit as st
import pdfplumber
import re
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule

# Regex para limpiar caracteres ilegales de Excel
ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

def clean_for_excel(text):
    """Elimina caracteres ilegales para Excel y espacios extra"""
    if not text: return ""
    text = str(text)
    text = ILLEGAL_CHARACTERS_RE.sub("", text)
    return text.strip()

def parse_numero_ar(s):
    """Parsea número en formato argentino (1.234,56) a float. Soporta sufijo - para negativos."""
    if not s: return 0.0
    s = s.strip()
    neg = False
    if s.endswith("-"):
        neg = True
        s = s[:-1]
    elif s.startswith("-"):
        neg = True
        s = s[1:]
    s = s.replace(".", "").replace(",", ".")
    try:
        val = float(s)
        return -val if neg else val
    except:
        return 0.0

# Mapeo de meses abreviados en español a número
MESES = {
    "ENE": "01", "FEB": "02", "MAR": "03", "ABR": "04",
    "MAY": "05", "JUN": "06", "JUL": "07", "AGO": "08",
    "SEP": "09", "OCT": "10", "NOV": "11", "DIC": "12"
}

def convertir_fecha_espacio(fecha_str):
    """Convierte '01 JUL 2026' a '01/07/2026'"""
    try:
        partes = fecha_str.split()
        dia = partes[0]
        mes = MESES.get(partes[1].upper(), partes[1])
        anio = partes[2]
        return f"{dia}/{mes}/{anio}"
    except:
        return fecha_str

def procesar_sol(archivo_pdf):
    st.info("Procesando archivo del Banco del Sol...")
    try:
        archivo_pdf.seek(0)
        with pdfplumber.open(io.BytesIO(archivo_pdf.read())) as pdf:
            texto_completo = ""
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    texto_completo += t + "\n"

        lineas = texto_completo.splitlines()

        # ============================================================
        # 1. METADATOS
        # ============================================================
        titular = "Sin Especificar"
        periodo = "Sin Especificar"
        cuenta = "Sin Especificar"
        saldo_inicial = 0.0
        saldo_final = 0.0

        # Titular: línea siguiente a "PÁGINA 1 DE N" (el PDF tiene problemas de
        # encoding en vocales acentuadas, por eso "P.GINA" con comodín)
        for i, line in enumerate(lineas):
            if re.match(r'^P.GINA 1 DE \d+', line.strip().upper()):
                if i + 1 < len(lineas):
                    titular = lineas[i + 1].strip()
                break

        # Cuenta: "Número de Cuenta:1000032618" (también con vocal rota)
        m_cuenta = re.search(r'N.mero\s+de\s+Cuenta:?\s*(\d+)', texto_completo)
        if m_cuenta:
            cuenta = m_cuenta.group(1).strip()

        # Fechas de inicio/cierre y saldo inicial/final: aparecen juntos en cada
        # página como un rótulo fijo seguido de sus valores
        m_header = re.search(
            r'FECHA DE INICIO FECHA DE CIERRE SALDO INICIAL SALDO AL CIERRE\s*\r?\n'
            r'\s*(\d{2}\s+[A-Z]{3}\s+\d{4})\s+(\d{2}\s+[A-Z]{3}\s+\d{4})\s+\$([\d.,]+)\s+\$([\d.,]+)',
            texto_completo
        )
        if m_header:
            fecha_inicio_raw = m_header.group(1)
            fecha_cierre_raw = m_header.group(2)
            saldo_inicial = parse_numero_ar(m_header.group(3))
            saldo_final = parse_numero_ar(m_header.group(4))
            periodo = f"Del {convertir_fecha_espacio(fecha_inicio_raw)} al {convertir_fecha_espacio(fecha_cierre_raw)}"

        # ============================================================
        # 2. MOVIMIENTOS - una línea con fecha+código+descripción+montos,
        #    seguida de 0 a 3 líneas de detalle sin montos.
        #    El importe se calcula como diferencia de saldos (igual que Ciudad).
        # ============================================================
        patron_fecha = re.compile(r'^(\d{2}/\d{2}/\d{4})\s+([\d.]+)\s+(.+)$')
        patron_monto = re.compile(r'\$\s?([\d.]+,\d{2}-?)')

        movimientos_raw = []
        bloque = None

        def cerrar_bloque():
            nonlocal bloque
            if bloque:
                movimientos_raw.append(bloque)
            bloque = None

        i = 0
        n = len(lineas)
        while i < n:
            line = lineas[i].strip()
            if not line:
                i += 1
                continue
            up = line.upper()

            if re.match(r'^P.GINA \d+ DE \d+', up):
                i += 1
                continue
            if up.startswith("FECHA DE INICIO"):
                # Rótulo + línea de valores, se repiten en cada página
                i += 2
                continue
            if re.match(r'^FECHA COD DESCRIPCI', up):
                i += 1
                continue
            if up.startswith("BANCO DEL SOL"):
                i += 1
                continue
            if up.startswith("TOTAL DEBITADO"):
                cerrar_bloque()
                break

            match = patron_fecha.match(line)
            if match:
                cerrar_bloque()
                fecha_raw = match.group(1)
                resto = match.group(3)
                montos = patron_monto.findall(resto)
                saldo = parse_numero_ar(montos[-1]) if montos else None
                concepto = patron_monto.sub('', resto).strip().rstrip(":")
                # Fix puntual: única palabra acentuada corrompida en el
                # contenido de las transacciones ("Pago con Tarj. D�bito")
                concepto = re.sub(r'D.bito', 'Débito', concepto)
                bloque = {
                    "fecha": fecha_raw,
                    "concepto": concepto,
                    "detalle": [],
                    "saldo": saldo,
                }
            else:
                if bloque:
                    bloque["detalle"].append(line)
            i += 1
        cerrar_bloque()

        transactions = []
        saldo_previo = saldo_inicial
        for mov in movimientos_raw:
            if mov["saldo"] is None:
                continue
            importe = round(mov["saldo"] - saldo_previo, 2)
            saldo_previo = mov["saldo"]

            if mov["detalle"]:
                descripcion = f"{mov['concepto']} - {' | '.join(mov['detalle'])}"
            else:
                descripcion = mov["concepto"]

            transactions.append({
                "Fecha": mov["fecha"],
                "Descripcion": clean_for_excel(descripcion),
                "Importe": importe
            })

        if not transactions:
            st.info("No se encontraron movimientos. Se generará el Excel solo con los saldos.")

        if transactions:
            st.success(f"Se encontraron {len(transactions)} movimientos.")

        # ============================================================
        # 3. GENERAR EXCEL
        # ============================================================
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Banco del Sol"
        ws.sheet_view.showGridLines = False

        # Paleta Banco del Sol (naranja)
        color_bg_main = "F26522"
        color_txt_main = "FFFFFF"

        thin_border = Border(
            left=Side(style='thin', color="A6A6A6"),
            right=Side(style='thin', color="A6A6A6"),
            top=Side(style='thin', color="A6A6A6"),
            bottom=Side(style='thin', color="A6A6A6")
        )

        fill_head_deb = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        fill_col_deb = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
        fill_row_deb = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")

        fill_head_cred = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        fill_col_cred = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")
        fill_row_cred = PatternFill(start_color="F2F9F1", end_color="F2F9F1", fill_type="solid")

        df = pd.DataFrame(transactions) if transactions else pd.DataFrame(columns=["Fecha", "Descripcion", "Importe"])
        creditos = df[df["Importe"] > 0].copy() if not df.empty else pd.DataFrame(columns=["Fecha", "Descripcion", "Importe"])
        debitos = df[df["Importe"] < 0].copy() if not df.empty else pd.DataFrame(columns=["Fecha", "Descripcion", "Importe"])
        if not debitos.empty:
            debitos["Importe"] = debitos["Importe"].abs()

        # --- Header ---
        ws.merge_cells("A1:G1")
        tit = ws["A1"]
        tit.value = f"REPORTE BANCO DEL SOL - CTA {clean_for_excel(cuenta)}"
        tit.font = Font(size=14, bold=True, color=color_txt_main)
        tit.fill = PatternFill(start_color=color_bg_main, end_color=color_bg_main, fill_type="solid")
        tit.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25

        # --- Metadata ---
        ws["A3"] = "SALDO INICIAL"
        ws["A3"].font = Font(bold=True, size=10, color="666666")
        ws["B3"] = saldo_inicial
        ws["B3"].number_format = '"$ "#,##0.00'
        ws["B3"].font = Font(bold=True, size=11)
        ws["B3"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

        ws["A4"] = "SALDO FINAL"
        ws["A4"].font = Font(bold=True, size=10, color="666666")
        ws["B4"] = saldo_final
        ws["B4"].number_format = '"$ "#,##0.00'
        ws["B4"].font = Font(bold=True, size=11)
        ws["B4"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

        ws["D3"] = "TITULAR"
        ws["D3"].alignment = Alignment(horizontal='right')
        ws["D3"].font = Font(bold=True, color="666666", size=10)
        ws["E3"] = clean_for_excel(titular)
        ws["E3"].font = Font(bold=True, size=11)
        ws["E3"].alignment = Alignment(horizontal='center')
        ws.merge_cells("E3:G3")
        for c in ["E","F","G"]: ws[f"{c}3"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

        ws["D4"] = "PERÍODO"
        ws["D4"].alignment = Alignment(horizontal='right')
        ws["D4"].font = Font(bold=True, color="666666", size=10)
        ws["E4"] = clean_for_excel(periodo)
        ws["E4"].font = Font(bold=True, size=11)
        ws["E4"].alignment = Alignment(horizontal='center')
        ws.merge_cells("E4:G4")
        for c in ["E","F","G"]: ws[f"{c}4"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

        # --- Tablas Créditos / Débitos ---
        fila_inicio = 10
        f_header = fila_inicio

        ws.merge_cells(f"A{f_header}:C{f_header}")
        ws[f"A{f_header}"] = "CRÉDITOS"
        ws[f"A{f_header}"].fill = fill_head_cred
        ws[f"A{f_header}"].font = Font(bold=True, color="FFFFFF")
        ws[f"A{f_header}"].alignment = Alignment(horizontal='center')
        ws[f"A{f_header}"].border = thin_border

        ws.merge_cells(f"E{f_header}:G{f_header}")
        ws[f"E{f_header}"] = "DÉBITOS"
        ws[f"E{f_header}"].fill = fill_head_deb
        ws[f"E{f_header}"].font = Font(bold=True, color="FFFFFF")
        ws[f"E{f_header}"].alignment = Alignment(horizontal='center')
        ws[f"E{f_header}"].border = thin_border

        headers = ["Fecha", "Descripción", "Importe"]
        cols_cred = ["A", "B", "C"]
        cols_deb = ["E", "F", "G"]
        f_sub = f_header + 1

        for i, h in enumerate(headers):
            c = ws[f"{cols_cred[i]}{f_sub}"]
            c.value = h
            c.fill = fill_col_cred
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal='center')
            c.border = thin_border

            d = ws[f"{cols_deb[i]}{f_sub}"]
            d.value = h
            d.fill = fill_col_deb
            d.font = Font(bold=True)
            d.alignment = Alignment(horizontal='center')
            d.border = thin_border

        fila_a_llenar = f_sub + 1

        # Créditos
        f_c = fila_a_llenar
        if creditos.empty:
            ws.merge_cells(f"A{f_c}:C{f_c}")
            ws[f"A{f_c}"] = "SIN MOVIMIENTOS"
            ws[f"A{f_c}"].border = thin_border
            f_c += 1
        else:
            start_c = f_c
            for _, r in creditos.iterrows():
                ws[f"A{f_c}"] = r["Fecha"]
                ws[f"A{f_c}"].fill = fill_row_cred
                ws[f"A{f_c}"].border = thin_border
                ws[f"A{f_c}"].alignment = Alignment(horizontal='center')
                ws[f"B{f_c}"] = r["Descripcion"]
                ws[f"B{f_c}"].fill = fill_row_cred
                ws[f"B{f_c}"].border = thin_border
                ws[f"C{f_c}"] = r["Importe"]
                ws[f"C{f_c}"].number_format = '"$ "#,##0.00'
                ws[f"C{f_c}"].fill = fill_row_cred
                ws[f"C{f_c}"].border = thin_border
                f_c += 1
            ws.merge_cells(f"A{f_c}:B{f_c}")
            ws[f"A{f_c}"] = "TOTAL CRÉDITOS"
            ws[f"A{f_c}"].font = Font(bold=True)
            ws[f"A{f_c}"].alignment = Alignment(horizontal='right')
            ws[f"C{f_c}"] = f"=SUM(C{start_c}:C{f_c-1})"
            ws[f"C{f_c}"].font = Font(bold=True)
            ws[f"C{f_c}"].number_format = '"$ "#,##0.00'
            f_c += 1

        # Débitos
        f_d = fila_a_llenar
        if debitos.empty:
            ws.merge_cells(f"E{f_d}:G{f_d}")
            ws[f"E{f_d}"] = "SIN MOVIMIENTOS"
            ws[f"E{f_d}"].border = thin_border
            f_d += 1
        else:
            start_d = f_d
            for _, r in debitos.iterrows():
                ws[f"E{f_d}"] = r["Fecha"]
                ws[f"E{f_d}"].fill = fill_row_deb
                ws[f"E{f_d}"].border = thin_border
                ws[f"E{f_d}"].alignment = Alignment(horizontal='center')
                ws[f"F{f_d}"] = r["Descripcion"]
                ws[f"F{f_d}"].fill = fill_row_deb
                ws[f"F{f_d}"].border = thin_border
                ws[f"G{f_d}"] = r["Importe"]
                ws[f"G{f_d}"].number_format = '"$ "#,##0.00'
                ws[f"G{f_d}"].fill = fill_row_deb
                ws[f"G{f_d}"].border = thin_border
                f_d += 1
            ws.merge_cells(f"E{f_d}:F{f_d}")
            ws[f"E{f_d}"] = "TOTAL DÉBITOS"
            ws[f"E{f_d}"].font = Font(bold=True)
            ws[f"E{f_d}"].alignment = Alignment(horizontal='right')
            ws[f"G{f_d}"] = f"=SUM(G{start_d}:G{f_d-1})"
            ws[f"G{f_d}"].font = Font(bold=True)
            ws[f"G{f_d}"].number_format = '"$ "#,##0.00'
            f_d += 1

        # --- Control de Saldos ---
        ws["D6"] = "CONTROL DE SALDOS"
        ws["D6"].font = Font(bold=True, size=10, color="666666")
        ws["D6"].alignment = Alignment(horizontal='center')

        ref_tot_c = f"C{f_c-1}" if not creditos.empty else "0"
        ref_tot_d = f"G{f_d-1}" if not debitos.empty else "0"
        ws["D7"] = f"=ROUND(B3+{ref_tot_c}-{ref_tot_d}-B4, 2)"
        ws["D7"].number_format = '"$ "#,##0.00'
        ws["D7"].font = Font(bold=True)
        ws["D7"].alignment = Alignment(horizontal='center')
        ws["D7"].border = thin_border

        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        red_font = Font(color='9C0006', bold=True)
        ws.conditional_formatting.add('D7', CellIsRule(operator='notEqual', formula=['0'], stopIfTrue=True, fill=red_fill, font=red_font))

        # --- Anchos de columnas ---
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 25
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 40
        ws.column_dimensions["G"].width = 18

        wb.save(output)
        output.seek(0)
        return output.getvalue()

    except Exception as e:
        import traceback
        st.error(f"Error al procesar: {e}")
        print(traceback.format_exc())
        return None
