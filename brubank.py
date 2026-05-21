import streamlit as st
import io
import PyPDF2
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule

ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')


def clean_for_excel(text):
    if not text:
        return ""
    text = str(text)
    text = ILLEGAL_CHARACTERS_RE.sub("", text)
    return text.strip()


def _parse_monto(s):
    """Parsea '$ 1.234,56', 'U$S 1.234,56', '- $ 1.234,56' o '-'. Devuelve None si es '-' solo."""
    if s is None:
        return None
    s = s.strip()
    if s == "-" or s == "":
        return None
    negativo = False
    if s.startswith("-"):
        negativo = True
        s = s[1:].strip()
    if s.endswith("-"):
        negativo = True
        s = s[:-1].strip()
    s = s.replace("U$S", "").replace("$", "").replace(" ", "")
    if not s:
        return None
    try:
        valor = float(s.replace(".", "").replace(",", "."))
    except ValueError:
        return None
    return -valor if negativo else valor


def _es_monto(linea):
    """True si la línea representa un monto ($, U$S, o '-' solo)."""
    l = linea.strip()
    if l == "-":
        return True
    if "$" in l or "U$S" in l:
        return True
    return False


def _es_fecha_mov(linea):
    return bool(re.match(r"^\d{2}-\d{2}-\d{2}$", linea.strip()))


def _es_ref(linea):
    return bool(re.match(r"^\d{6,}$", linea.strip()))


def _parsear_cuentas(lineas):
    """Segmenta el texto en bloques de cuenta (uno por cada 'Mi cuenta' encontrado)."""
    indices_inicio = []
    for i, l in enumerate(lineas):
        if l.strip() == "Mi cuenta":
            if i + 1 < len(lineas) and lineas[i + 1].strip() == "Resumen":
                indices_inicio.append(i)

    if not indices_inicio:
        return []

    bloques = []
    for k, inicio in enumerate(indices_inicio):
        fin = indices_inicio[k + 1] if k + 1 < len(indices_inicio) else len(lineas)
        bloques.append(lineas[inicio:fin])

    titular_global, periodo_global = _extraer_titular_periodo(lineas)

    cuentas = []
    for bloque in bloques:
        cuenta = _parsear_bloque_cuenta(bloque)
        cuenta["titular"] = titular_global
        cuenta["periodo"] = periodo_global
        cuentas.append(cuenta)
    return cuentas


def _extraer_titular_periodo(lineas):
    """Busca el footer 'N MES YYYY al N MES YYYY' y toma la línea siguiente como titular."""
    titular = "Sin Especificar"
    periodo = "Sin Especificar"
    patron_periodo = re.compile(r"^\d{1,2}\s+[A-Z]{3}\s+\d{4}\s+al\s+\d{1,2}\s+[A-Z]{3}\s+\d{4}$")
    for i, l in enumerate(lineas):
        if patron_periodo.match(l.strip()):
            periodo = l.strip()
            if i + 1 < len(lineas):
                titular = lineas[i + 1].strip()
            break
    return titular, periodo


def _parsear_bloque_cuenta(lineas_bloque):
    """Extrae metadata y movimientos de un bloque 'Mi cuenta'."""
    cuenta = {
        "moneda": "ARS",
        "simbolo": "$",
        "tipo": "",
        "saldo_inicial": 0.0,
        "saldo_final": 0.0,
        "creditos_total": 0.0,
        "debitos_total": 0.0,
        "cuit": "",
        "numero": "",
        "cbu": "",
        "movimientos": [],
    }

    etiquetas = {
        "Tipo": "tipo",
        "Saldo Inicial": "saldo_inicial",
        "Moneda": "moneda",
        "CUIT": "cuit",
        "Número": "numero",
        "Numero": "numero",
        "CBU": "cbu",
        "Saldo Final": "saldo_final",
    }
    re_creditos = re.compile(r"^Cr.?ditos$")
    re_debitos = re.compile(r"^D.?bitos$")
    re_numero = re.compile(r"^N.?mero$")

    idx_movimientos = None
    for i, raw in enumerate(lineas_bloque):
        l = raw.strip()
        if l == "Movimientos":
            idx_movimientos = i
            break
        siguiente = lineas_bloque[i + 1].strip() if i + 1 < len(lineas_bloque) else ""
        if l in etiquetas and siguiente:
            campo = etiquetas[l]
            if campo in ("saldo_inicial", "saldo_final"):
                valor = _parse_monto(siguiente)
                if valor is not None:
                    cuenta[campo] = valor
            elif campo == "moneda":
                if "USD" in siguiente or "lar" in siguiente.lower():
                    cuenta["moneda"] = "USD"
                else:
                    cuenta["moneda"] = "ARS"
            else:
                cuenta[campo] = siguiente
        elif re_creditos.match(l) and siguiente:
            v = _parse_monto(siguiente)
            if v is not None:
                cuenta["creditos_total"] = v
        elif re_debitos.match(l) and siguiente:
            v = _parse_monto(siguiente)
            if v is not None:
                cuenta["debitos_total"] = v
        elif re_numero.match(l) and siguiente:
            cuenta["numero"] = siguiente

    cuenta["simbolo"] = "U$S" if cuenta["moneda"] == "USD" else "$"

    if idx_movimientos is not None:
        cuenta["movimientos"] = _parsear_movimientos(lineas_bloque[idx_movimientos + 1:])

    return cuenta


def _parsear_movimientos(lineas):
    """Recorre las líneas y arma movimientos de a 6 campos (Fecha, Ref, Desc..., Deb, Cred, Saldo)."""
    # Líneas a ignorar (encabezados y footers repetidos por página).
    skip_exactos = {"Movimientos", "Fecha", "#Ref", "Débito", "Crédito", "Saldo"}
    re_skip_descripcion = re.compile(r"^Descripci.?n$")
    re_skip_debito = re.compile(r"^D.?bito$")
    re_skip_credito = re.compile(r"^Cr.?dito$")
    re_periodo = re.compile(r"^\d{1,2}\s+[A-Z]{3}\s+\d{4}\s+al\s+\d{1,2}\s+[A-Z]{3}\s+\d{4}$")

    # Filtramos líneas de relleno conservando el orden original.
    lineas_utiles = []
    saltar_siguientes = 0
    for l in lineas:
        s = l.strip()
        if not s:
            continue
        if saltar_siguientes > 0:
            saltar_siguientes -= 1
            continue
        if s in skip_exactos:
            continue
        if re_skip_descripcion.match(s) or re_skip_debito.match(s) or re_skip_credito.match(s):
            continue
        if re_periodo.match(s):
            # Después del período viene el titular y, opcionalmente, líneas de dirección
            # hasta llegar a la próxima fecha de movimiento. Las descartamos a medida que aparecen.
            continue
        lineas_utiles.append(s)

    movimientos = []
    i = 0
    n = len(lineas_utiles)
    while i < n:
        if not _es_fecha_mov(lineas_utiles[i]):
            i += 1
            continue
        fecha = lineas_utiles[i]
        i += 1
        if i >= n or not _es_ref(lineas_utiles[i]):
            # Estructura inesperada; intentamos seguir.
            continue
        i += 1  # consumimos ref
        # Acumular descripción hasta encontrar un monto.
        desc_partes = []
        while i < n and not _es_monto(lineas_utiles[i]) and not _es_fecha_mov(lineas_utiles[i]):
            desc_partes.append(lineas_utiles[i])
            i += 1
        if i + 2 >= n:
            break
        debito_raw = lineas_utiles[i]; i += 1
        credito_raw = lineas_utiles[i]; i += 1
        # El saldo a veces se parte: ej. "-" en una línea y "$ 60.662,36" en la siguiente.
        # En la mayoría de los casos viene completo: "- $ 60.662,36" o "$ 92.824,54".
        saldo_raw = lineas_utiles[i]; i += 1
        if saldo_raw.strip() == "-" and i < n and _es_monto(lineas_utiles[i]):
            saldo_raw = "- " + lineas_utiles[i]
            i += 1

        debito = _parse_monto(debito_raw)
        credito = _parse_monto(credito_raw)
        saldo = _parse_monto(saldo_raw)

        if credito is not None:
            importe = credito
        elif debito is not None:
            importe = -debito
        else:
            importe = 0.0

        descripcion = " ".join(p for p in desc_partes if p).strip()
        # Limpiar repeticiones de espacios.
        descripcion = re.sub(r"\s+", " ", descripcion)

        movimientos.append({
            "Fecha": fecha,
            "Descripcion": descripcion,
            "Importe": round(importe, 2),
            "Saldo": saldo if saldo is not None else 0.0,
        })

    return movimientos


def _render_dashboard(ws, cuenta):
    """Pinta el dashboard estilo galicia.py en la hoja `ws` con los datos de `cuenta`."""
    ws.sheet_view.showGridLines = False
    simbolo = cuenta["simbolo"]
    fmt_num = f'"{simbolo} "#,##0.00'

    color_bg_main = "5B2EFF"  # violeta Brubank
    color_txt_main = "FFFFFF"

    thin_border = Border(
        left=Side(style='thin', color="A6A6A6"),
        right=Side(style='thin', color="A6A6A6"),
        top=Side(style='thin', color="A6A6A6"),
        bottom=Side(style='thin', color="A6A6A6"),
    )

    fill_head_deb = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    fill_col_deb = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
    fill_row_deb = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")

    fill_head_cred = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    fill_col_cred = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")
    fill_row_cred = PatternFill(start_color="F2F9F1", end_color="F2F9F1", fill_type="solid")

    df = pd.DataFrame(cuenta["movimientos"])
    if not df.empty:
        creditos = df[df["Importe"] > 0].copy()
        debitos = df[df["Importe"] < 0].copy()
        debitos["Importe"] = debitos["Importe"].abs()
    else:
        creditos = pd.DataFrame(columns=["Fecha", "Descripcion", "Importe"])
        debitos = pd.DataFrame(columns=["Fecha", "Descripcion", "Importe"])

    sufijo_moneda = f" {cuenta['moneda']}" if cuenta["moneda"] != "ARS" else ""
    ws.merge_cells("A1:G1")
    tit = ws["A1"]
    tit.value = f"REPORTE BRUBANK{sufijo_moneda} - {clean_for_excel(cuenta['titular'])}"
    tit.font = Font(size=14, bold=True, color=color_txt_main)
    tit.fill = PatternFill(start_color=color_bg_main, end_color=color_bg_main, fill_type="solid")
    tit.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25

    ws["A3"] = "SALDO INICIAL"
    ws["A3"].font = Font(bold=True, size=10, color="666666")
    ws["B3"] = cuenta["saldo_inicial"]
    ws["B3"].number_format = fmt_num
    ws["B3"].font = Font(bold=True, size=11)
    ws["B3"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

    ws["A4"] = "SALDO FINAL"
    ws["A4"].font = Font(bold=True, size=10, color="666666")
    ws["B4"] = cuenta["saldo_final"]
    ws["B4"].number_format = fmt_num
    ws["B4"].font = Font(bold=True, size=11)
    ws["B4"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

    ws["D3"] = "TITULAR"
    ws["D3"].alignment = Alignment(horizontal='right')
    ws["D3"].font = Font(bold=True, color="666666", size=10)
    ws["E3"] = clean_for_excel(cuenta["titular"])
    ws["E3"].font = Font(bold=True, size=11)
    ws["E3"].alignment = Alignment(horizontal='center')
    ws.merge_cells("E3:G3")
    for c in ["E", "F", "G"]:
        ws[f"{c}3"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

    ws["D4"] = "PERÍODO"
    ws["D4"].alignment = Alignment(horizontal='right')
    ws["D4"].font = Font(bold=True, color="666666", size=10)
    ws["E4"] = clean_for_excel(cuenta["periodo"])
    ws["E4"].font = Font(bold=True, size=11)
    ws["E4"].alignment = Alignment(horizontal='center')
    ws.merge_cells("E4:G4")
    for c in ["E", "F", "G"]:
        ws[f"{c}4"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

    ws["A6"] = "CUENTA"
    ws["A6"].font = Font(bold=True, size=10, color="666666")
    ws["B6"] = f"{cuenta['tipo']} ({cuenta['moneda']}) - N° {cuenta['numero']}"
    ws["B6"].font = Font(size=10)
    ws.merge_cells("B6:C6")

    ws["D6"] = "CONTROL DE SALDOS"
    ws["D6"].font = Font(bold=True, size=10, color="666666")
    ws["D6"].alignment = Alignment(horizontal='center')

    cell_ctl = ws["D7"]
    cell_ctl.font = Font(bold=True, size=12)
    cell_ctl.alignment = Alignment(horizontal='center')
    cell_ctl.border = thin_border

    fila_inicio = 10
    f_header = fila_inicio

    ws.merge_cells(f"A{f_header}:C{f_header}")
    ws[f"A{f_header}"] = "CRÉDITOS"
    ws[f"A{f_header}"].fill = fill_head_cred
    ws[f"A{f_header}"].font = Font(bold=True, color="FFFFFF")
    ws[f"A{f_header}"].alignment = Alignment(horizontal='center')
    ws[f"A{f_header}"].border = thin_border

    headers = ["Fecha", "Descripción", "Importe"]
    cols_cred = ["A", "B", "C"]
    f_sub = f_header + 1
    for i, h in enumerate(headers):
        c = ws[f"{cols_cred[i]}{f_sub}"]
        c.value = h
        c.fill = fill_col_cred
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal='center')
        c.border = thin_border

    ws.merge_cells(f"E{f_header}:G{f_header}")
    ws[f"E{f_header}"] = "DÉBITOS"
    ws[f"E{f_header}"].fill = fill_head_deb
    ws[f"E{f_header}"].font = Font(bold=True, color="FFFFFF")
    ws[f"E{f_header}"].alignment = Alignment(horizontal='center')
    ws[f"E{f_header}"].border = thin_border

    cols_deb = ["E", "F", "G"]
    for i, h in enumerate(headers):
        c = ws[f"{cols_deb[i]}{f_sub}"]
        c.value = h
        c.fill = fill_col_deb
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal='center')
        c.border = thin_border

    fila_dato_start = f_sub + 1

    f_cred = fila_dato_start
    if creditos.empty:
        ws.merge_cells(f"A{f_cred}:C{f_cred}")
        ws[f"A{f_cred}"] = "SIN MOVIMIENTOS"
        ws[f"A{f_cred}"].font = Font(italic=True, color="666666")
        ws[f"A{f_cred}"].alignment = Alignment(horizontal='center')
        ws[f"A{f_cred}"].border = thin_border
        f_cred += 1
        f_tot_cred_ref = "0"
    else:
        start_c = f_cred
        for _, r in creditos.iterrows():
            ws[f"A{f_cred}"] = clean_for_excel(r["Fecha"])
            ws[f"A{f_cred}"].fill = fill_row_cred
            ws[f"A{f_cred}"].alignment = Alignment(horizontal='center')
            ws[f"A{f_cred}"].border = thin_border

            ws[f"B{f_cred}"] = clean_for_excel(r["Descripcion"])
            ws[f"B{f_cred}"].fill = fill_row_cred
            ws[f"B{f_cred}"].border = thin_border

            ws[f"C{f_cred}"] = r["Importe"]
            ws[f"C{f_cred}"].number_format = fmt_num
            ws[f"C{f_cred}"].fill = fill_row_cred
            ws[f"C{f_cred}"].border = thin_border
            f_cred += 1

        ws.merge_cells(f"A{f_cred}:B{f_cred}")
        ws[f"A{f_cred}"] = "TOTAL CRÉDITOS"
        ws[f"A{f_cred}"].font = Font(bold=True)
        ws[f"A{f_cred}"].alignment = Alignment(horizontal='right')
        ws[f"A{f_cred}"].border = thin_border

        ws[f"C{f_cred}"] = f"=SUM(C{start_c}:C{f_cred-1})"
        ws[f"C{f_cred}"].number_format = fmt_num
        ws[f"C{f_cred}"].font = Font(bold=True)
        ws[f"C{f_cred}"].border = thin_border
        f_tot_cred_ref = f"C{f_cred}"
        f_cred += 1

    f_deb = fila_dato_start
    if debitos.empty:
        ws.merge_cells(f"E{f_deb}:G{f_deb}")
        ws[f"E{f_deb}"] = "SIN MOVIMIENTOS"
        ws[f"E{f_deb}"].font = Font(italic=True, color="666666")
        ws[f"E{f_deb}"].alignment = Alignment(horizontal='center')
        ws[f"E{f_deb}"].border = thin_border
        f_deb += 1
        f_tot_deb_ref = "0"
    else:
        start_d = f_deb
        for _, r in debitos.iterrows():
            ws[f"E{f_deb}"] = clean_for_excel(r["Fecha"])
            ws[f"E{f_deb}"].fill = fill_row_deb
            ws[f"E{f_deb}"].alignment = Alignment(horizontal='center')
            ws[f"E{f_deb}"].border = thin_border

            ws[f"F{f_deb}"] = clean_for_excel(r["Descripcion"])
            ws[f"F{f_deb}"].fill = fill_row_deb
            ws[f"F{f_deb}"].border = thin_border

            ws[f"G{f_deb}"] = r["Importe"]
            ws[f"G{f_deb}"].number_format = fmt_num
            ws[f"G{f_deb}"].fill = fill_row_deb
            ws[f"G{f_deb}"].border = thin_border
            f_deb += 1

        ws.merge_cells(f"E{f_deb}:F{f_deb}")
        ws[f"E{f_deb}"] = "TOTAL DÉBITOS"
        ws[f"E{f_deb}"].font = Font(bold=True)
        ws[f"E{f_deb}"].alignment = Alignment(horizontal='right')
        ws[f"E{f_deb}"].border = thin_border

        ws[f"G{f_deb}"] = f"=SUM(G{start_d}:G{f_deb-1})"
        ws[f"G{f_deb}"].number_format = fmt_num
        ws[f"G{f_deb}"].font = Font(bold=True)
        ws[f"G{f_deb}"].border = thin_border
        f_tot_deb_ref = f"G{f_deb}"
        f_deb += 1

    ws["D7"] = f"=ROUND(B3+{f_tot_cred_ref}-{f_tot_deb_ref}-B4, 2)"
    ws["D7"].number_format = fmt_num

    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    red_font = Font(color='9C0006', bold=True)
    ws.conditional_formatting.add(
        'D7',
        CellIsRule(operator='notEqual', formula=['0'], stopIfTrue=True, fill=red_fill, font=red_font),
    )

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 40
    ws.column_dimensions["G"].width = 18


def procesar_brubank(archivo_pdf):
    """Procesa archivos PDF del banco Brubank con Estilo Dashboard (una hoja por cuenta)."""
    st.info("Procesando archivo del banco Brubank...")

    try:
        archivo_pdf.seek(0)
        with io.BytesIO(archivo_pdf.read()) as pdf_file:
            reader = PyPDF2.PdfReader(pdf_file)
            texto_completo = "".join(page.extract_text() + "\n" for page in reader.pages)
        lineas = [l for l in texto_completo.splitlines()]

        cuentas = _parsear_cuentas(lineas)
        if not cuentas:
            st.error("No se encontró ninguna cuenta ('Mi cuenta') en el PDF.")
            return None

        for c in cuentas:
            suma_cred = sum(m["Importe"] for m in c["movimientos"] if m["Importe"] > 0)
            suma_deb = sum(-m["Importe"] for m in c["movimientos"] if m["Importe"] < 0)
            if c["creditos_total"] and abs(suma_cred - c["creditos_total"]) > 1:
                st.warning(
                    f"Cuenta {c['moneda']}: total créditos parseado ({suma_cred:.2f}) "
                    f"difiere del declarado en el PDF ({c['creditos_total']:.2f})."
                )
            if c["debitos_total"] and abs(suma_deb - c["debitos_total"]) > 1:
                st.warning(
                    f"Cuenta {c['moneda']}: total débitos parseado ({suma_deb:.2f}) "
                    f"difiere del declarado en el PDF ({c['debitos_total']:.2f})."
                )

        output = io.BytesIO()
        wb = Workbook()
        wb.remove(wb.active)
        for c in cuentas:
            ws = wb.create_sheet(title=f"Brubank {c['moneda']}")
            _render_dashboard(ws, c)

        wb.save(output)
        output.seek(0)
        return output.getvalue()

    except Exception as e:
        import traceback
        st.error(f"Error al procesar el archivo: {str(e)}")
        print(traceback.format_exc())
        return None
