import streamlit as st
import io
import pdfplumber
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule

# Regex para caracteres ilegales en Excel
ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

def clean_for_excel(text):
    if not text: return ""
    text = str(text)
    text = ILLEGAL_CHARACTERS_RE.sub("", text)
    return text.strip()

def parse_float(val_str):
    try:
        if not val_str: return 0.0
        # Credicoop F2 usa punto decimal y sin separador miles visibles en tabla
        # "4342.67"
        return float(val_str)
    except:
        # Fallback por si acaso viene con coma
        try:
            return float(val_str.replace(".", "").replace(",", "."))
        except:
            return 0.0

def procesar_credicoop_formato_2(archivo_pdf):
    """Procesa Credicoop Formato 2 -> Estilo Dashboard"""
    st.info("Procesando archivo Credicoop (Formato 2)...")

    try:
        texto_completo = ""
        with pdfplumber.open(io.BytesIO(archivo_pdf.read())) as pdf:
            for page in pdf.pages:
                texto_completo += page.extract_text() + "\n"
        
        lineas = texto_completo.splitlines()
        
        # --- 1. Metadata ---
        titular = "Sin Especificar"
        cuenta = "Sin Especificar"
        periodo = "Sin Especificar"
        
        # Unir las primeras líneas para búsqueda más fácil
        header_text = "\n".join(lineas[:25])
        
        # Búsqueda Titular (Adherente)
        match_tit = re.search(r"Adherente:\s*(.*)", header_text, re.IGNORECASE)
        if match_tit:
            titular = match_tit.group(1).strip()

        # Búsqueda Cuenta
        match_cta = re.search(r"Nro\.?\s*de\s*Cuenta:\s*(.*)", header_text, re.IGNORECASE)
        if match_cta:
            cuenta = match_cta.group(1).strip()

        # Búsqueda Periodo
        if "Saldos y movimientos" in header_text:
             match_per = re.search(r"Saldos y movimientos\s+(.*)", header_text)
             if match_per:
                 periodo = match_per.group(1).strip()

        # --- 2. Movimientos y Saldos ---
        # Regex update: Permitir negativos con "-?" en los grupos de montos.
        # Estructura: Fecha | ...Desc... | Deb | Cred | Saldo | Cod
        
        regex_mov_flexible = r"^(\d{2}/\d{2}/\d{4})\s+(.+?)\s+(-?\d+\.\d{2})\s+(-?\d+\.\d{2})\s+(-?\d+\.\d{2})\s+([A-Za-z0-9]+)$"
        
        movimientos = []
        
        for l in lineas:
            l = l.strip()
            # Headers ignorar
            if "Fecha" in l and "Concepto" in l: continue
            if "Saldos y movimientos" in l: continue
            if "Saldo Disponible" in l: continue
            if "Saldo Contable" in l: continue
            if not l: continue

            match = re.match(regex_mov_flexible, l)
            if match:
                fecha = match.group(1)
                desc_raw = match.group(2).strip()
                deb_str = match.group(3)
                cred_str = match.group(4)
                saldo_str = match.group(5)
                # codigo = match.group(6)

                debito = parse_float(deb_str)
                credito = parse_float(cred_str)
                saldo = parse_float(saldo_str)
                
                # Cálculo Importe (Neto)
                # Ojo: Débito suele ser positivo en columna debito.
                # Si Credito > 0 -> Importe positivo.
                # Si Debito > 0 -> Importe negativo.
                importe = 0.0
                if abs(credito) > 0:
                    importe = abs(credito)
                elif abs(debito) > 0:
                    importe = -abs(debito)
                
                movimientos.append({
                    "Fecha": fecha,
                    "Descripcion": desc_raw,
                    "Importe": importe,
                    "SaldoLinea": saldo,
                    "DebitoRaw": debito, 
                    "CreditoRaw": credito
                })
            else:
                # Append línea multilínea
                if l and movimientos and not re.match(r"^\d{2}/\d{2}/\d{4}", l):
                    if "Página" in l or "Banco Credicoop" in l or "Adherente:" in l:
                        continue
                    movimientos[-1]["Descripcion"] += " " + l

        if not movimientos:
            st.error("No se encontraron movimientos")
            return None
            
        # Orden Cronológico Ascendente (para Excel)
        # El PDF viene Descendente.
        
        # Saldo Final del Periodo = Saldo del primer movimiento leido (el mas reciente)
        saldo_final_reporte = movimientos[0]["SaldoLinea"]
        
        # Saldo Inicial del Periodo = Saldo del ultimo movimiento (mas antiguo) REVERTIDO.
        # SaldoInicial = SaldoFinalLineaAntigua - (Credito - Debito)
        ultimo_mov = movimientos[-1]
        
        # Ojo: Aquí debemos usar los valores RAW de la columna.
        # En el ejemplo:
        # Saldo: -6351.03
        # Debito: 15.45
        # Credito: 0.00
        # Saldo Previo = -6351.03 - (0.00 - 15.45) = -6351.03 + 15.45 = -6335.58
        # Verificar logica: SaldoNuevo = SaldoViejo + Cred - Deb
        # => SaldoViejo = SaldoNuevo - Cred + Deb
        # Correcto.
        saldo_inicial_reporte = ultimo_mov["SaldoLinea"] - ultimo_mov["CreditoRaw"] + ultimo_mov["DebitoRaw"]
        
        movimientos.reverse()
        
        # DataFrame
        df = pd.DataFrame(movimientos)
        
        # Separar DFs
        creditos = df[df["Importe"] > 0].copy()
        debitos = df[df["Importe"] < 0].copy()
        debitos["Importe"] = debitos["Importe"].abs()

        # --- GENERACIÓN EXCEL (DASHBOARD) ---
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Credicoop"
        ws.sheet_view.showGridLines = False
        
        # Paleta
        color_bg_main = "003366" 
        color_txt_main = "FFFFFF"
        
        thin_border = Border(left=Side(style='thin', color="A6A6A6"), 
                             right=Side(style='thin', color="A6A6A6"), 
                             top=Side(style='thin', color="A6A6A6"), 
                             bottom=Side(style='thin', color="A6A6A6"))
                             
        fill_head_deb = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        fill_col_deb = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
        fill_row_deb = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")

        fill_head_cred = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        fill_col_cred = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")
        fill_row_cred = PatternFill(start_color="F2F9F1", end_color="F2F9F1", fill_type="solid")

        # 1. Header
        ws.merge_cells("A1:G1")
        tit = ws["A1"]
        tit.value = f"REPORTE CREDICOOP - {clean_for_excel(titular)}"
        tit.font = Font(size=14, bold=True, color=color_txt_main)
        tit.fill = PatternFill(start_color=color_bg_main, end_color=color_bg_main, fill_type="solid")
        tit.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25

        # 2. Metadata
        ws["A3"] = "SALDO INICIAL"
        ws["A3"].font = Font(bold=True, size=10, color="666666")
        ws["B3"] = saldo_inicial_reporte
        ws["B3"].number_format = '"$ "#,##0.00'
        ws["B3"].font = Font(bold=True, size=11)
        ws["B3"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

        ws["A4"] = "SALDO FINAL"
        ws["A4"].font = Font(bold=True, size=10, color="666666")
        ws["B4"] = saldo_final_reporte
        ws["B4"].number_format = '"$ "#,##0.00'
        ws["B4"].font = Font(bold=True, size=11)
        ws["B4"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

        ws["D3"] = "CUENTA"
        ws["D3"].alignment = Alignment(horizontal='right')
        ws["D3"].font = Font(bold=True, color="666666", size=10)
        ws["E3"] = clean_for_excel(cuenta)
        ws["E3"].font = Font(bold=True, size=11)
        ws["E3"].alignment = Alignment(horizontal='center')
        ws.merge_cells("E3:G3")
        for c in ["E","F","G"]: ws[f"{c}3"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

        ws["D4"] = "PERÍODO"
        ws["D4"].alignment = Alignment(horizontal='right')
        ws["D4"].font = Font(bold=True, color="666666", size=10)
        ws["E4"] = clean_for_excel(periodo)
        ws["E4"].font = Font(bold=True, size=11)
        ws["E4"].alignment = Alignment(horizontal='center')
        ws.merge_cells("E4:G4")
        for c in ["E","F","G"]: ws[f"{c}4"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

        ws["D6"] = "CONTROL DE SALDOS"
        ws["D6"].font = Font(bold=True, size=10, color="666666")
        ws["D6"].alignment = Alignment(horizontal='center')
        
        cell_ctl = ws["D7"]
        cell_ctl.font = Font(bold=True, size=12)
        cell_ctl.alignment = Alignment(horizontal='center')
        cell_ctl.border = thin_border

        # 3. Tables
        f_header = 10
        
        # Creditos
        ws.merge_cells(f"A{f_header}:C{f_header}")
        ws[f"A{f_header}"] = "CRÉDITOS" 
        ws[f"A{f_header}"].fill = fill_head_cred
        ws[f"A{f_header}"].font = Font(bold=True, color="FFFFFF")
        ws[f"A{f_header}"].alignment = Alignment(horizontal='center')
        
        headers = ["Fecha", "Descripción", "Importe"]
        cols_cred = ["A", "B", "C"]
        f_sub = f_header + 1
        for i, h in enumerate(headers):
            c = ws[f"{cols_cred[i]}{f_sub}"]
            c.value = h
            c.fill = fill_col_cred
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal='center')
            c.border = thin_border

        # Debitos
        ws.merge_cells(f"E{f_header}:G{f_header}")
        ws[f"E{f_header}"] = "DÉBITOS" 
        ws[f"E{f_header}"].fill = fill_head_deb
        ws[f"E{f_header}"].font = Font(bold=True, color="FFFFFF")
        ws[f"E{f_header}"].alignment = Alignment(horizontal='center')
        
        cols_deb = ["E", "F", "G"]
        for i, h in enumerate(headers):
            c = ws[f"{cols_deb[i]}{f_sub}"]
            c.value = h
            c.fill = fill_col_deb
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal='center')
            c.border = thin_border

        # Data
        f_cred = f_sub + 1
        if creditos.empty:
            ws.merge_cells(f"A{f_cred}:C{f_cred}")
            ws[f"A{f_cred}"] = "SIN MOVIMIENTOS"
            ws[f"A{f_cred}"].alignment = Alignment(horizontal='center')
            ws[f"A{f_cred}"].font = Font(italic=True, color="666666")
            ws[f"A{f_cred}"].border = thin_border
            f_cred += 1
        else:
            start_c = f_cred
            for _, r in creditos.iterrows():
                ws[f"A{f_cred}"] = clean_for_excel(r["Fecha"])
                ws[f"A{f_cred}"].fill = fill_row_cred
                ws[f"A{f_cred}"].alignment = Alignment(horizontal='center')
                ws[f"A{f_cred}"].border = thin_border
                
                ws[f"B{f_cred}"] = clean_for_excel(r["Descripcion"])
                ws[f"B{f_cred}"].fill = fill_row_cred
                ws[f"B{f_cred}"].border = thin_border
                
                ws[f"C{f_cred}"] = r["Importe"]
                ws[f"C{f_cred}"].number_format = '"$ "#,##0.00'
                ws[f"C{f_cred}"].fill = fill_row_cred
                ws[f"C{f_cred}"].border = thin_border
                f_cred += 1
            
            # Total Cred
            ws.merge_cells(f"A{f_cred}:B{f_cred}")
            ws[f"A{f_cred}"] = "TOTAL"
            ws[f"A{f_cred}"].alignment = Alignment(horizontal='right')
            ws[f"A{f_cred}"].font = Font(bold=True)
            ws[f"A{f_cred}"].border = thin_border
            ws[f"C{f_cred}"] = f"=SUM(C{start_c}:C{f_cred-1})"
            ws[f"C{f_cred}"].number_format = '"$ "#,##0.00'
            ws[f"C{f_cred}"].font = Font(bold=True)
            ws[f"C{f_cred}"].border = thin_border
            f_cred += 1

        f_deb = f_sub + 1
        if debitos.empty:
            ws.merge_cells(f"E{f_deb}:G{f_deb}")
            ws[f"E{f_deb}"] = "SIN MOVIMIENTOS"
            ws[f"E{f_deb}"].alignment = Alignment(horizontal='center')
            ws[f"E{f_deb}"].font = Font(italic=True, color="666666")
            ws[f"E{f_deb}"].border = thin_border
            f_deb += 1
        else:
            start_d = f_deb
            for _, r in debitos.iterrows():
                ws[f"E{f_deb}"] = clean_for_excel(r["Fecha"])
                ws[f"E{f_deb}"].fill = fill_row_deb
                ws[f"E{f_deb}"].alignment = Alignment(horizontal='center')
                ws[f"E{f_deb}"].border = thin_border
                
                ws[f"F{f_deb}"] = clean_for_excel(r["Descripcion"])
                ws[f"F{f_deb}"].fill = fill_row_deb
                ws[f"F{f_deb}"].border = thin_border
                
                ws[f"G{f_deb}"] = r["Importe"]
                ws[f"G{f_deb}"].number_format = '"$ "#,##0.00'
                ws[f"G{f_deb}"].fill = fill_row_deb
                ws[f"G{f_deb}"].border = thin_border
                f_deb += 1
            
            # Total Deb
            ws.merge_cells(f"E{f_deb}:F{f_deb}")
            ws[f"E{f_deb}"] = "TOTAL"
            ws[f"E{f_deb}"].alignment = Alignment(horizontal='right')
            ws[f"E{f_deb}"].font = Font(bold=True)
            ws[f"E{f_deb}"].border = thin_border
            ws[f"G{f_deb}"] = f"=SUM(G{start_d}:G{f_deb-1})"
            ws[f"G{f_deb}"].number_format = '"$ "#,##0.00'
            ws[f"G{f_deb}"].font = Font(bold=True)
            ws[f"G{f_deb}"].border = thin_border
            f_deb += 1

        # Control
        f_ini = "B3"
        f_tot_cred = f"C{f_cred-1}" if not creditos.empty else "0"
        f_tot_deb = f"G{f_deb-1}" if not debitos.empty else "0"
        f_fin = "B4"
        
        ws["D7"] = f"=ROUND({f_ini}+{f_tot_cred}-{f_tot_deb}-{f_fin}, 2)"
        ws["D7"].number_format = '"$ "#,##0.00'
        
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        red_font = Font(color='9C0006', bold=True)
        ws.conditional_formatting.add('D7', CellIsRule(operator='notEqual', formula=['0'], stopIfTrue=True, fill=red_fill, font=red_font))

        # Widths
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 25
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 40
        ws.column_dimensions["G"].width = 18

        wb.save(output)
        output.seek(0)
        return output.getvalue()
        
    except Exception as e:
        import traceback
        st.error(f"Error procesando Credicoop F2: {e}")
        print(traceback.format_exc())
        return None

    except Exception as e:
        import traceback
        st.error(f"Error procesando Credicoop F2: {e}")
        print(traceback.format_exc())
        return None


